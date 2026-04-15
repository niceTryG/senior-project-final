from datetime import datetime, date, timedelta
from io import BytesIO
import os

from flask import Blueprint, render_template, session, redirect, url_for, flash, request, send_file, current_app, abort
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import joinedload

from ..extensions import db
from ..forms import (
    OperationalTaskForm,
    ProfileUpdateForm,
    ChangePasswordForm,
    TelegramLinkCodeForm,
    WorkspaceProfileForm,
    WorkspaceOwnershipTransferForm,
    WorkspaceTeamMemberForm,
)
from ..models import CashRecord, CuttingOrder, CuttingOrderMaterial, OperationalTask, Production, ProductionPlan, Product, ProductComposition, ShopOrder, ShopOrderItem, ShopStock, Sale, StockMovement, User, TelegramLinkCode, WholesaleSale, WholesaleSaleItem, Factory, Shop, ShopFactoryLink, Fabric, SupplierReceipt
from ..services.product_service import ProductService
from ..translations import t as translate
from ..user_identity import build_login_username, normalize_phone
from ..user_display import (
    display_value,
    get_user_display_name,
    get_user_initials,
    get_workspace_name,
)


main_bp = Blueprint("main", __name__)
product_service = ProductService()


SYSTEM_TASK_SOURCE_META = {
    "shop_low_stock": {
        "group": "Stock",
        "group_key": "task_group_stock",
        "action_label": "Transfer stock",
        "action_label_key": "task_action_transfer_stock",
    },
    "no_production_today": {
        "group": "Production",
        "group_key": "task_group_production",
        "action_label": "Record production",
        "action_label_key": "task_action_record_production",
    },
    "workspace_setup": {
        "group": "Workspace",
        "group_key": "task_group_workspace",
        "action_label": "Finish setup",
        "action_label_key": "task_action_finish_setup",
    },
    "no_shop_linked": {
        "group": "Expansion",
        "group_key": "task_group_expansion",
        "action_label": "Review workspace",
        "action_label_key": "task_action_review_workspace",
    },
    "owner_only_account": {
        "group": "Team",
        "group_key": "task_group_team",
        "action_label": "Add teammate",
        "action_label_key": "task_action_add_teammate",
    },
    "no_products": {
        "group": "Catalog",
        "group_key": "task_group_catalog",
        "action_label": "Open products",
        "action_label_key": "task_action_open_products",
    },
    "factory_stock_heavy": {
        "group": "Stock flow",
        "group_key": "task_group_stock_flow",
        "action_label": "Review transfer",
        "action_label_key": "task_action_review_transfer",
    },
    "owner_telegram_missing": {
        "group": "Notifications",
        "group_key": "task_group_notifications",
        "action_label": "Link Telegram",
        "action_label_key": "task_action_link_telegram",
    },
    "no_cash_records": {
        "group": "Cash",
        "group_key": "task_group_cash",
        "action_label": "Open cash",
        "action_label_key": "task_action_open_cash",
    },
    "sales_without_cash_today": {
        "group": "Cash",
        "group_key": "task_group_cash",
        "action_label": "Review cash",
        "action_label_key": "task_action_review_cash",
    },
    "pending_shop_orders": {
        "group": "Orders",
        "group_key": "task_group_orders",
        "action_label": "Open factory queue",
        "action_label_key": "task_action_open_factory_queue",
    },
    "ready_shop_orders": {
        "group": "Orders",
        "group_key": "task_group_orders",
        "action_label": "Review ready orders",
        "action_label_key": "task_action_review_ready_orders",
    },
    "daily_close_review": {
        "group": "Daily close",
        "group_key": "task_group_daily_close",
        "action_label": "Review close",
        "action_label_key": "task_action_review_close",
    },
}

OPS_STAGE_FLOW = [
    {
        "key": "cutting",
        "label_key": "ops_stage_cutting_label",
        "task_type": "ops_stage_cutting",
        "description_key": "ops_stage_cutting_description",
        "priority": "high",
        "target_role": "viewer",
        "action_label_key": "ops_stage_cutting_action_label",
    },
    {
        "key": "sewing",
        "label_key": "ops_stage_sewing_label",
        "task_type": "ops_stage_sewing",
        "description_key": "ops_stage_sewing_description",
        "priority": "high",
        "target_role": "viewer",
        "action_label_key": "ops_stage_sewing_action_label",
    },
    {
        "key": "packing",
        "label_key": "ops_stage_packing_label",
        "task_type": "ops_stage_packing",
        "description_key": "ops_stage_packing_description",
        "priority": "medium",
        "target_role": "viewer",
        "action_label_key": "ops_stage_packing_action_label",
    },
    {
        "key": "ready",
        "label_key": "ops_stage_ready_label",
        "task_type": "ops_stage_ready",
        "description_key": "ops_stage_ready_description",
        "priority": "medium",
        "target_role": "viewer",
        "action_label_key": "ops_stage_ready_action_label",
    },
]

OPS_STAGE_INDEX = {row["key"]: idx for idx, row in enumerate(OPS_STAGE_FLOW)}
OPS_STAGE_BY_TASK_TYPE = {row["task_type"]: row for row in OPS_STAGE_FLOW}


def _find_default_linked_shop(factory_id: int | None):
    if not factory_id:
        return None

    row = (
        ShopFactoryLink.query
        .join(Shop, Shop.id == ShopFactoryLink.shop_id)
        .filter(
            ShopFactoryLink.factory_id == factory_id,
            Shop.name == "Main Shop",
        )
        .first()
    )
    return row.shop if row and row.shop else None


def _resolve_telegram_link_factory_id(user) -> int | None:
    factory_id = getattr(user, "factory_id", None)
    if factory_id:
        return int(factory_id)

    shop = getattr(user, "shop", None)
    if shop and getattr(shop, "factory_id", None):
        return int(shop.factory_id)

    session_factory_id = session.get("factory_id")
    try:
        return int(session_factory_id) if session_factory_id else None
    except (TypeError, ValueError):
        return None


RU_MONTHS = {
    "January": "января",
    "February": "февраля",
    "March": "марта",
    "April": "апреля",
    "May": "мая",
    "June": "июня",
    "July": "июля",
    "August": "августа",
    "September": "сентября",
    "October": "октября",
    "November": "ноября",
    "December": "декабря",
}

UZ_MONTHS = {
    "January": "yanvar",
    "February": "fevral",
    "March": "mart",
    "April": "aprel",
    "May": "may",
    "June": "iyun",
    "July": "iyul",
    "August": "avgust",
    "September": "sentabr",
    "October": "oktyabr",
    "November": "noyabr",
    "December": "dekabr",
}


def _get_current_date_for_lang():
    now = datetime.now()
    day = now.strftime("%d")
    year = now.strftime("%Y")
    eng_month = now.strftime("%B")

    lang = session.get("lang_code", "ru")

    if lang == "ru":
        month = RU_MONTHS.get(eng_month, eng_month)
    elif lang == "uz":
        month = UZ_MONTHS.get(eng_month, eng_month)
    else:
        month = eng_month

    return f"{day} {month} {year}"


def _calc_cash_totals(factory_id: int):
    records = CashRecord.query.filter_by(factory_id=factory_id).all()
    total_uzs = sum(float(r.amount or 0) for r in records if (r.currency or "UZS").upper() == "UZS")
    total_usd = sum(float(r.amount or 0) for r in records if (r.currency or "UZS").upper() == "USD")
    return total_uzs, total_usd


def _get_production_today_summary(factory_id: int):
    today = date.today()

    rows = (
        db.session.query(Product.name, func.coalesce(func.sum(Production.quantity), 0))
        .join(Production, Production.product_id == Product.id)
        .filter(Product.factory_id == factory_id)
        .filter(Production.date == today)
        .group_by(Product.name)
        .order_by(func.sum(Production.quantity).desc())
        .all()
    )

    produced_today_total = int(sum(int(qty or 0) for _, qty in rows))
    produced_today_models = int(len(rows))
    return produced_today_total, produced_today_models, rows


def _get_production_week_total(factory_id: int):
    today = date.today()
    week_start = today - timedelta(days=6)

    total = (
        db.session.query(func.coalesce(func.sum(Production.quantity), 0))
        .join(Product, Product.id == Production.product_id)
        .filter(Product.factory_id == factory_id)
        .filter(Production.date >= week_start)
        .filter(Production.date <= today)
        .scalar()
    )
    return int(total or 0)


def _get_shop_low_stock(factory_id: int, threshold: int = 5, limit: int = 3):
    rows = (
        ShopStock.query
        .join(Product, Product.id == ShopStock.product_id)
        .filter(Product.factory_id == factory_id)
        .filter(ShopStock.quantity < threshold)
        .order_by(ShopStock.quantity.asc(), Product.name.asc())
        .all()
    )

    top_items = [
        {"id": r.product_id, "name": r.product.name, "qty": int(r.quantity or 0)}
        for r in rows[:limit]
    ]
    return len(rows), top_items


def _get_yesterday_transfer_total(factory_id: int):
    y = date.today() - timedelta(days=1)
    start = datetime(y.year, y.month, y.day, 0, 0, 0)
    end = start + timedelta(days=1)

    total = (
        db.session.query(func.coalesce(func.sum(StockMovement.qty_change), 0))
        .filter(StockMovement.factory_id == factory_id)
        .filter(
            StockMovement.movement_type.in_(("factory_to_shop", "factory_to_shop_for_order"))
        )
        .filter(StockMovement.timestamp >= start)
        .filter(StockMovement.timestamp < end)
        .scalar()
    )

    return int(total or 0)


def _sale_amount_uzs(sale, product) -> float:
    if hasattr(sale, "total_sell") and sale.total_sell is not None:
        try:
            return float(sale.total_sell or 0)
        except Exception:
            return 0.0

    qty = getattr(sale, "quantity", 0) or 0
    price = getattr(sale, "sell_price_per_item", None)
    if price is None:
        price = getattr(product, "sell_price_per_item", 0) or 0

    try:
        return float(qty) * float(price)
    except Exception:
        return 0.0


def _normalize_to_date(raw_value):
    if not raw_value:
        return None
    if isinstance(raw_value, datetime):
        return raw_value.date()
    return raw_value


def _get_sales_dashboard_stats(factory_id: int):
    today = date.today()
    yesterday = today - timedelta(days=1)
    week_start = today - timedelta(days=6)

    retail_rows = (
        db.session.query(Sale, Product)
        .join(Product, Product.id == Sale.product_id)
        .filter(Product.factory_id == factory_id)
        .all()
    )

    wholesale_rows = (
        db.session.query(WholesaleSaleItem, WholesaleSale)
        .join(WholesaleSale, WholesaleSale.id == WholesaleSaleItem.wholesale_sale_id)
        .filter(WholesaleSaleItem.source_factory_id == factory_id)
        .all()
    )

    today_sales_uzs = 0.0
    yesterday_sales_uzs = 0.0
    week_sales_uzs = 0.0

    by_product = {}

    for sale, product in retail_rows:
        s_date = _normalize_to_date(getattr(sale, "date", None))
        if not s_date:
            continue

        amount = _sale_amount_uzs(sale, product)

        if s_date == today:
            today_sales_uzs += amount
        if s_date == yesterday:
            yesterday_sales_uzs += amount
        if week_start <= s_date <= today:
            week_sales_uzs += amount

        pid = product.id
        if pid not in by_product:
            by_product[pid] = {
                "product_id": pid,
                "name": product.name,
                "qty": 0,
                "amount_uzs": 0.0,
            }

        by_product[pid]["qty"] += int(getattr(sale, "quantity", 0) or 0)
        by_product[pid]["amount_uzs"] += amount

    for item, wholesale_sale in wholesale_rows:
        s_date = _normalize_to_date(getattr(wholesale_sale, "sale_date", None))
        if not s_date:
            continue

        amount = float(getattr(item, "line_total", 0) or 0)

        if s_date == today:
            today_sales_uzs += amount
        if s_date == yesterday:
            yesterday_sales_uzs += amount
        if week_start <= s_date <= today:
            week_sales_uzs += amount

        product = getattr(item, "product", None)
        if not product:
            continue

        pid = product.id
        if pid not in by_product:
            by_product[pid] = {
                "product_id": pid,
                "name": product.name,
                "qty": 0,
                "amount_uzs": 0.0,
            }

        by_product[pid]["qty"] += int(getattr(item, "quantity", 0) or 0)
        by_product[pid]["amount_uzs"] += amount

    top_selling_models = sorted(
        by_product.values(),
        key=lambda x: (x["qty"], x["amount_uzs"]),
        reverse=True,
    )[:5]

    return {
        "today_sales_uzs": today_sales_uzs,
        "yesterday_sales_uzs": yesterday_sales_uzs,
        "week_sales_uzs": week_sales_uzs,
        "top_selling_models": top_selling_models,
    }


def _empty_currency_totals():
    return {"UZS": 0.0, "USD": 0.0}


def _summary_currency_key(currency_value) -> str:
    return "USD" if str(currency_value or "").strip().upper() == "USD" else "UZS"


def _add_currency_amount(bucket, amount, currency_value) -> None:
    key = _summary_currency_key(currency_value)
    bucket[key] = float(bucket.get(key, 0.0) or 0.0) + float(amount or 0.0)


def _format_summary_money(amount, currency_value: str) -> str:
    currency_code = _summary_currency_key(currency_value)
    precision = 2 if currency_code == "USD" else 0
    return f"{float(amount or 0.0):,.{precision}f} {currency_code}"


def _format_summary_money_pair(bucket, *, zero_label: str | None = None):
    uzs_amount = float(bucket.get("UZS", 0.0) or 0.0)
    usd_amount = float(bucket.get("USD", 0.0) or 0.0)

    if abs(uzs_amount) >= 0.01 or abs(usd_amount) < 0.01:
        primary = _format_summary_money(uzs_amount, "UZS")
        secondary = _format_summary_money(usd_amount, "USD") if abs(usd_amount) >= 0.01 else None
    else:
        primary = _format_summary_money(usd_amount, "USD")
        secondary = _format_summary_money(uzs_amount, "UZS") if abs(uzs_amount) >= 0.01 else None

    if zero_label and abs(uzs_amount) < 0.01 and abs(usd_amount) < 0.01:
        primary = zero_label

    return {
        "primary": primary,
        "secondary": secondary,
    }


def _format_change_hint(current_value: float, previous_value: float) -> dict:
    current_amount = float(current_value or 0.0)
    previous_amount = float(previous_value or 0.0)

    if abs(previous_amount) < 0.01:
        if abs(current_amount) < 0.01:
            return {"label": translate("flat_vs_prior_period"), "tone": "flat"}
        return {"label": translate("new_activity_vs_prior_period"), "tone": "up"}

    delta = current_amount - previous_amount
    if abs(delta) < 0.01:
        return {"label": translate("flat_vs_prior_period"), "tone": "flat"}

    pct_change = abs((delta / previous_amount) * 100)
    tone = "up" if delta > 0 else "down"
    return {
        "label": translate("percent_" + tone + "_vs_prior_period").format(percent=int(pct_change)),
        "tone": tone,
    }


def _summary_export_filename(workspace_name: str | None, suffix: str, extension: str) -> str:
    slug = "".join(
        ch.lower() if ch.isalnum() else "_"
        for ch in str(workspace_name or "workspace")
    ).strip("_")
    while "__" in slug:
        slug = slug.replace("__", "_")
    if not slug:
        slug = "workspace"
    return f"{slug}_{suffix}.{extension}"


def _summary_pdf_fonts():
    regular_candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        os.path.join(current_app.root_path, "static", "fonts", "DejaVuSans.ttf"),
    ]
    bold_candidates = [
        r"C:\Windows\Fonts\arialbd.ttf",
        os.path.join(current_app.root_path, "static", "fonts", "DejaVuSans-Bold.ttf"),
    ]

    regular_path = next((path for path in regular_candidates if os.path.exists(path)), None)
    bold_path = next((path for path in bold_candidates if os.path.exists(path)), None)

    if regular_path and "AdrasSummaryRegular" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("AdrasSummaryRegular", regular_path))
    if bold_path and "AdrasSummaryBold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("AdrasSummaryBold", bold_path))

    return (
        "AdrasSummaryRegular" if regular_path else "Helvetica",
        "AdrasSummaryBold" if bold_path else "Helvetica-Bold",
    )


def _summary_pdf_table(rows, col_widths=None):
    table = Table(rows, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef4ff")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbe3ef")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _summary_xlsx_style_sheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="1E293B")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _summary_xlsx_fit_columns(ws, *, max_width: int = 34):
    for column_cells in ws.columns:
        values = [
            len(str(cell.value or ""))
            for cell in column_cells
        ]
        if not values:
            continue
        width = min(max(max(values) + 2, 12), max_width)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def _build_business_summary_pdf(summary_state):
    regular_font, bold_font = _summary_pdf_fonts()
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        name="AdrasSummaryTitle",
        parent=styles["Title"],
        fontName=bold_font,
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_LEFT,
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        name="AdrasSummarySubtitle",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        name="AdrasSummarySection",
        parent=styles["Heading2"],
        fontName=bold_font,
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#0f172a"),
        spaceBefore=8,
        spaceAfter=8,
    )
    normal_style = ParagraphStyle(
        name="AdrasSummaryNormal",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#1f2937"),
    )
    small_style = ParagraphStyle(
        name="AdrasSummarySmall",
        parent=styles["Normal"],
        fontName=regular_font,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#475569"),
    )
    bold_style = ParagraphStyle(
        name="AdrasSummaryBold",
        parent=styles["Normal"],
        fontName=bold_font,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    story = [
        Paragraph("Adras Business Summary", title_style),
        Paragraph(
            f"{summary_state['workspace_page_name']} · generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            subtitle_style,
        ),
        Paragraph(
            (
                f"Workspace health: <b>{summary_state['health_score']}</b> ({summary_state['health_label']}). "
                f"Owner: <b>{summary_state['workspace_owner_name']}</b>. "
                f"Open issues: <b>{summary_state['command_center_open_count']}</b>."
            ),
            normal_style,
        ),
        Spacer(1, 8),
        Paragraph("Core metrics", section_style),
    ]

    metrics_rows = [[Paragraph("<b>Metric</b>", bold_style), Paragraph("<b>Value</b>", bold_style), Paragraph("<b>Note</b>", bold_style)]]
    for card in summary_state["summary_cards"]:
        metrics_rows.append(
            [
                Paragraph(str(card["label"]), small_style),
                Paragraph(str(card["value"]), normal_style),
                Paragraph(str(card["sub"]), small_style),
            ]
        )
    story.append(_summary_pdf_table(metrics_rows, col_widths=[48 * mm, 48 * mm, 82 * mm]))

    if summary_state["branch_rows_all"]:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Branches", section_style))
        branch_rows = [[
            Paragraph("<b>Branch</b>", bold_style),
            Paragraph("<b>Month sales</b>", bold_style),
            Paragraph("<b>Month profit</b>", bold_style),
            Paragraph("<b>Stock value</b>", bold_style),
            Paragraph("<b>Low stock</b>", bold_style),
        ]]
        for branch in summary_state["branch_rows_all"][:10]:
            branch_rows.append(
                [
                    Paragraph(str(branch["name"]), small_style),
                    Paragraph(str(branch["month_sales_display"]["primary"]), small_style),
                    Paragraph(str(branch["month_profit_display"]["primary"]), small_style),
                    Paragraph(str(branch["stock_value_display"]["primary"]), small_style),
                    Paragraph(str(branch["low_stock_count"]), small_style),
                ]
            )
        story.append(_summary_pdf_table(branch_rows, col_widths=[52 * mm, 36 * mm, 36 * mm, 36 * mm, 20 * mm]))

    if summary_state["top_product_rows_all"]:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Best sellers", section_style))
        product_rows = [[
            Paragraph("<b>Product</b>", bold_style),
            Paragraph("<b>Sold 30d</b>", bold_style),
            Paragraph("<b>Revenue</b>", bold_style),
            Paragraph("<b>Profit</b>", bold_style),
            Paragraph("<b>Stock split</b>", bold_style),
        ]]
        for product in summary_state["top_product_rows_all"][:10]:
            product_rows.append(
                [
                    Paragraph(str(product["name"]), small_style),
                    Paragraph(str(product["sold_qty_30"]), small_style),
                    Paragraph(str(product["revenue_display"]["primary"]), small_style),
                    Paragraph(str(product["profit_display"]["primary"]), small_style),
                    Paragraph(f"{product['factory_qty']} factory / {product['shop_qty']} shop", small_style),
                ]
            )
        story.append(_summary_pdf_table(product_rows, col_widths=[52 * mm, 24 * mm, 36 * mm, 36 * mm, 40 * mm]))

    if summary_state["attention_items"]:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Needs attention", section_style))
        attention_rows = [[
            Paragraph("<b>Task</b>", bold_style),
            Paragraph("<b>Priority</b>", bold_style),
            Paragraph("<b>Status</b>", bold_style),
            Paragraph("<b>Summary</b>", bold_style),
        ]]
        for item in summary_state["attention_items"][:8]:
            attention_rows.append(
                [
                    Paragraph(str(item["title"]), small_style),
                    Paragraph(str(item["priority_label"]), small_style),
                    Paragraph(str(item["status_label"]), small_style),
                    Paragraph(str(item["summary_line"]), small_style),
                ]
            )
        story.append(_summary_pdf_table(attention_rows, col_widths=[56 * mm, 22 * mm, 24 * mm, 78 * mm]))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def _build_business_summary_xlsx(summary_state):
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value", "Note"])
    for card in summary_state["summary_cards"]:
        ws.append([card["label"], card["value"], card["sub"]])
    ws.append([])
    ws.append(["Workspace health", summary_state["health_score"], summary_state["health_label"]])
    ws.append(["Owner", summary_state["workspace_owner_name"], summary_state["workspace_owner_login"]])
    ws.append(["Open issues", summary_state["command_center_open_count"], f"Overdue: {summary_state['command_center_overdue_count']}"])
    _summary_xlsx_style_sheet(ws)
    _summary_xlsx_fit_columns(ws)

    ws_branches = wb.create_sheet("Branches")
    ws_branches.append(["Branch", "Today sales", "Month sales", "Month profit", "Stock value", "Stock units", "Sold 30d", "Low stock count", "Note"])
    for branch in summary_state["branch_rows_all"]:
        ws_branches.append([
            branch["name"],
            branch["today_sales_display"]["primary"],
            branch["month_sales_display"]["primary"],
            branch["month_profit_display"]["primary"],
            branch["stock_value_display"]["primary"],
            branch["stock_units"],
            branch["sold_units_30"],
            branch["low_stock_count"],
            branch["note"],
        ])
    _summary_xlsx_style_sheet(ws_branches)
    _summary_xlsx_fit_columns(ws_branches)

    ws_products = wb.create_sheet("Best Sellers")
    ws_products.append(["Product", "Category", "Sold 30d", "On hand", "Factory qty", "Shop qty", "Revenue", "Profit", "Pace"])
    for product in summary_state["top_product_rows_all"]:
        ws_products.append([
            product["name"],
            product["category"],
            product["sold_qty_30"],
            product["on_hand_qty"],
            product["factory_qty"],
            product["shop_qty"],
            product["revenue_display"]["primary"],
            product["profit_display"]["primary"],
            product["pace_label"],
        ])
    _summary_xlsx_style_sheet(ws_products)
    _summary_xlsx_fit_columns(ws_products)

    ws_slow = wb.create_sheet("Slow Movers")
    ws_slow.append(["Product", "Category", "On hand", "Factory qty", "Shop qty", "Sold 30d", "Dormancy"])
    for product in summary_state["slow_mover_rows_all"]:
        ws_slow.append([
            product["name"],
            product["category"],
            product["on_hand_qty"],
            product["factory_qty"],
            product["shop_qty"],
            product["sold_qty_30"],
            product["dormant_label"],
        ])
    _summary_xlsx_style_sheet(ws_slow)
    _summary_xlsx_fit_columns(ws_slow)

    ws_customers = wb.create_sheet("Customers")
    ws_customers.append(["Customer", "Phone", "Orders", "Revenue", "Profit", "Last seen"])
    for customer in summary_state["customer_rows"]:
        ws_customers.append([
            customer["name"],
            customer["phone"],
            customer["orders_count"],
            customer["revenue_display"]["primary"],
            customer["profit_display"]["primary"],
            customer["last_seen"],
        ])
    _summary_xlsx_style_sheet(ws_customers)
    _summary_xlsx_fit_columns(ws_customers)

    ws_attention = wb.create_sheet("Attention")
    ws_attention.append(["Task", "Priority", "Status", "Origin", "Group", "Summary", "Link"])
    for item in summary_state["attention_items"]:
        ws_attention.append([
            item["title"],
            item["priority_label"],
            item["status_label"],
            item["origin_label"],
            item["group"],
            item["summary_line"],
            item["href"],
        ])
    _summary_xlsx_style_sheet(ws_attention)
    _summary_xlsx_fit_columns(ws_attention)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_business_summary(factory_id: int):
    workspace = Factory.query.get(factory_id) if factory_id else None
    if not workspace:
        return {}

    today = date.today()
    yesterday = today - timedelta(days=1)
    week_start = today - timedelta(days=6)
    month_start = date(today.year, today.month, 1)
    rolling_start = today - timedelta(days=29)
    customer_window_start = today - timedelta(days=89)

    workspace_owner = _get_workspace_owner_user(workspace)
    workspace_team = _get_workspace_team_users(factory_id)
    workspace_shops = _get_workspace_shops(factory_id)
    product_rows = (
        Product.query
        .filter(Product.factory_id == factory_id)
        .order_by(Product.name.asc())
        .all()
    )
    setup_items = _build_dashboard_setup_items(factory_id)
    setup_remaining = sum(1 for item in setup_items if not item["done"])
    command_center = _build_command_center_snapshot(factory_id)

    produced_today_total, produced_today_models, _prod_today_rows = _get_production_today_summary(factory_id)
    produced_week_total = _get_production_week_total(factory_id)
    yesterday_transfer_total = _get_yesterday_transfer_total(factory_id)
    shop_low_stock_count, shop_low_stock_items = _get_shop_low_stock(factory_id=factory_id, limit=5)
    cash_total_uzs, cash_total_usd = _calc_cash_totals(factory_id=factory_id)

    sales_totals = {
        "today": _empty_currency_totals(),
        "yesterday": _empty_currency_totals(),
        "week": _empty_currency_totals(),
        "month": _empty_currency_totals(),
        "month_profit": _empty_currency_totals(),
    }
    cash_totals = {
        "UZS": float(cash_total_uzs or 0.0),
        "USD": float(cash_total_usd or 0.0),
    }
    factory_stock_cost = _empty_currency_totals()
    shop_stock_value = _empty_currency_totals()

    product_summaries = {}
    product_shop_qty = {}
    customer_summaries = {}
    branch_summaries = {}

    def ensure_branch(branch_id, branch_name: str):
        normalized_id = int(branch_id or 0)
        if normalized_id not in branch_summaries:
            branch_summaries[normalized_id] = {
                "id": normalized_id,
                "name": branch_name or "Direct / unassigned",
                "today_sales": _empty_currency_totals(),
                "month_sales": _empty_currency_totals(),
                "month_profit": _empty_currency_totals(),
                "stock_value": _empty_currency_totals(),
                "stock_units": 0,
                "sold_units_30": 0,
                "low_stock_count": 0,
            }
        return branch_summaries[normalized_id]

    def ensure_product(product):
        if product.id not in product_summaries:
            product_summaries[product.id] = {
                "id": product.id,
                "name": product.name,
                "category": display_value(getattr(product, "category", None), fallback="Uncategorized"),
                "factory_qty": int(product.quantity or 0),
                "shop_qty": 0,
                "on_hand_qty": int(product.quantity or 0),
                "sold_qty_30": 0,
                "revenue_30": _empty_currency_totals(),
                "profit_30": _empty_currency_totals(),
                "last_sale_date": None,
            }
        return product_summaries[product.id]

    def ensure_customer(name_value, phone_value):
        customer_name = (name_value or "").strip() or "Unnamed customer"
        customer_phone = normalize_phone(phone_value) or (phone_value or "").strip() or "-"
        customer_key = normalize_phone(phone_value) or f"name:{customer_name.lower()}"
        if customer_key not in customer_summaries:
            customer_summaries[customer_key] = {
                "name": customer_name,
                "phone": customer_phone,
                "revenue": _empty_currency_totals(),
                "profit": _empty_currency_totals(),
                "orders": set(),
                "last_seen": None,
            }
        return customer_summaries[customer_key]

    for product in product_rows:
        ensure_product(product)
        _add_currency_amount(
            factory_stock_cost,
            (product.quantity or 0) * float(product.cost_price_per_item or 0.0),
            getattr(product, "currency", None),
        )

    stock_rows = (
        ShopStock.query
        .join(Product, Product.id == ShopStock.product_id)
        .join(Shop, Shop.id == ShopStock.shop_id)
        .filter(ShopStock.source_factory_id == factory_id)
        .all()
    )
    for stock_row in stock_rows:
        product = stock_row.product
        shop = stock_row.shop
        if not product or not shop:
            continue

        product_state = ensure_product(product)
        qty = int(stock_row.quantity or 0)
        product_state["shop_qty"] += qty
        product_state["on_hand_qty"] += qty
        product_shop_qty[product.id] = product_shop_qty.get(product.id, 0) + qty

        stock_value = qty * float(product.sell_price_per_item or 0.0)
        _add_currency_amount(shop_stock_value, stock_value, getattr(product, "currency", None))

        branch_state = ensure_branch(shop.id, shop.name)
        branch_state["stock_units"] += qty
        _add_currency_amount(branch_state["stock_value"], stock_value, getattr(product, "currency", None))
        if 0 < qty < 5:
            branch_state["low_stock_count"] += 1

    retail_sales_rows = (
        db.session.query(Sale, Product, Shop)
        .join(Product, Product.id == Sale.product_id)
        .outerjoin(Shop, Shop.id == Sale.shop_id)
        .filter(Product.factory_id == factory_id)
        .all()
    )

    for sale, product, shop in retail_sales_rows:
        if not sale or not product or not getattr(sale, "date", None):
            continue

        sale_date = _normalize_to_date(getattr(sale, "date", None))
        if not sale_date:
            continue

        revenue = float(sale.total_sell or 0.0)
        profit = float(sale.profit or 0.0)
        qty = int(sale.quantity or 0)
        currency_value = getattr(sale, "currency", None) or getattr(product, "currency", None)

        if sale_date == today:
            _add_currency_amount(sales_totals["today"], revenue, currency_value)
        if sale_date == yesterday:
            _add_currency_amount(sales_totals["yesterday"], revenue, currency_value)
        if week_start <= sale_date <= today:
            _add_currency_amount(sales_totals["week"], revenue, currency_value)
        if month_start <= sale_date <= today:
            _add_currency_amount(sales_totals["month"], revenue, currency_value)
            _add_currency_amount(sales_totals["month_profit"], profit, currency_value)

        branch_state = ensure_branch(getattr(shop, "id", None), getattr(shop, "name", None) or "Direct / unassigned")
        if sale_date == today:
            _add_currency_amount(branch_state["today_sales"], revenue, currency_value)
        if month_start <= sale_date <= today:
            _add_currency_amount(branch_state["month_sales"], revenue, currency_value)
            _add_currency_amount(branch_state["month_profit"], profit, currency_value)
        if rolling_start <= sale_date <= today:
            branch_state["sold_units_30"] += qty

        product_state = ensure_product(product)
        if rolling_start <= sale_date <= today:
            product_state["sold_qty_30"] += qty
            _add_currency_amount(product_state["revenue_30"], revenue, currency_value)
            _add_currency_amount(product_state["profit_30"], profit, currency_value)
        if not product_state["last_sale_date"] or sale_date > product_state["last_sale_date"]:
            product_state["last_sale_date"] = sale_date

        if sale_date >= customer_window_start and (sale.customer_name or sale.customer_phone):
            customer_state = ensure_customer(sale.customer_name, sale.customer_phone)
            _add_currency_amount(customer_state["revenue"], revenue, currency_value)
            _add_currency_amount(customer_state["profit"], profit, currency_value)
            customer_state["orders"].add(f"retail:{sale.id}")
            if not customer_state["last_seen"] or sale_date > customer_state["last_seen"]:
                customer_state["last_seen"] = sale_date

    wholesale_rows = (
        db.session.query(WholesaleSaleItem, WholesaleSale, Shop)
        .join(WholesaleSale, WholesaleSale.id == WholesaleSaleItem.wholesale_sale_id)
        .outerjoin(Shop, Shop.id == WholesaleSale.shop_id)
        .filter(WholesaleSaleItem.source_factory_id == factory_id)
        .all()
    )

    for item, wholesale_sale, shop in wholesale_rows:
        sale_date = _normalize_to_date(getattr(wholesale_sale, "sale_date", None))
        if not item or not wholesale_sale or not sale_date:
            continue

        revenue = float(item.line_total or 0.0)
        profit = revenue - (float(item.quantity or 0) * float(item.cost_price_per_item or 0.0))
        qty = int(item.quantity or 0)
        currency_value = getattr(item, "currency", None) or getattr(wholesale_sale, "currency", None)

        if sale_date == today:
            _add_currency_amount(sales_totals["today"], revenue, currency_value)
        if sale_date == yesterday:
            _add_currency_amount(sales_totals["yesterday"], revenue, currency_value)
        if week_start <= sale_date <= today:
            _add_currency_amount(sales_totals["week"], revenue, currency_value)
        if month_start <= sale_date <= today:
            _add_currency_amount(sales_totals["month"], revenue, currency_value)
            _add_currency_amount(sales_totals["month_profit"], profit, currency_value)

        branch_state = ensure_branch(getattr(shop, "id", None), getattr(shop, "name", None) or "Wholesale / unassigned")
        if sale_date == today:
            _add_currency_amount(branch_state["today_sales"], revenue, currency_value)
        if month_start <= sale_date <= today:
            _add_currency_amount(branch_state["month_sales"], revenue, currency_value)
            _add_currency_amount(branch_state["month_profit"], profit, currency_value)
        if rolling_start <= sale_date <= today:
            branch_state["sold_units_30"] += qty

        product = getattr(item, "product", None)
        if product and getattr(product, "factory_id", None) == factory_id:
            product_state = ensure_product(product)
            if rolling_start <= sale_date <= today:
                product_state["sold_qty_30"] += qty
                _add_currency_amount(product_state["revenue_30"], revenue, currency_value)
                _add_currency_amount(product_state["profit_30"], profit, currency_value)
            if not product_state["last_sale_date"] or sale_date > product_state["last_sale_date"]:
                product_state["last_sale_date"] = sale_date

        if sale_date >= customer_window_start and (wholesale_sale.customer_name or wholesale_sale.customer_phone):
            customer_state = ensure_customer(wholesale_sale.customer_name, wholesale_sale.customer_phone)
            _add_currency_amount(customer_state["revenue"], revenue, currency_value)
            _add_currency_amount(customer_state["profit"], profit, currency_value)
            customer_state["orders"].add(f"wholesale:{wholesale_sale.id}")
            if not customer_state["last_seen"] or sale_date > customer_state["last_seen"]:
                customer_state["last_seen"] = sale_date

    for product in product_rows:
        product_state = ensure_product(product)
        product_state["shop_qty"] = int(product_shop_qty.get(product.id, product_state["shop_qty"]) or 0)
        product_state["on_hand_qty"] = int(product_state["factory_qty"] + product_state["shop_qty"])

    setup_penalty = setup_remaining * 7
    low_stock_penalty = min(shop_low_stock_count * 3, 18)
    open_task_penalty = min(int(command_center["open_count"] or 0) * 2, 16)
    overdue_penalty = min(int(command_center["overdue_count"] or 0) * 5, 20)
    production_penalty = 6 if produced_today_total == 0 else 0
    telegram_penalty = 4 if not bool(workspace_owner and getattr(workspace_owner, "telegram_links", None)) else 0
    health_score = max(8, 100 - setup_penalty - low_stock_penalty - open_task_penalty - overdue_penalty - production_penalty - telegram_penalty)

    if health_score >= 86:
        health_label = "Strong"
        health_copy = "The workspace looks stable, and the owner view has room to focus on growth rather than cleanup."
        health_tone = "strong"
    elif health_score >= 68:
        health_label = "Stable"
        health_copy = "The business is running, but there are still a few open issues that deserve owner attention."
        health_tone = "stable"
    elif health_score >= 48:
        health_label = "Needs focus"
        health_copy = "The workspace is active, but stock, setup, or overdue issues are starting to drag on the operating picture."
        health_tone = "watch"
    else:
        health_label = "At risk"
        health_copy = "Important tasks are piling up, so the owner should step in before branch, stock, or cash discipline slips further."
        health_tone = "risk"

    top_product_rows_all = sorted(
        product_summaries.values(),
        key=lambda row: (
            int(row["sold_qty_30"] or 0),
            float(row["revenue_30"]["UZS"] or 0.0),
            float(row["revenue_30"]["USD"] or 0.0),
        ),
        reverse=True,
    )
    for row in top_product_rows_all:
        row["revenue_display"] = _format_summary_money_pair(row["revenue_30"], zero_label="0 UZS")
        row["profit_display"] = _format_summary_money_pair(row["profit_30"], zero_label="0 UZS")
        row["pace_label"] = (
            "Strong" if row["sold_qty_30"] >= 20
            else "Moving" if row["sold_qty_30"] >= 6
            else "Slow"
        )
    top_product_rows = top_product_rows_all[:6]

    slow_mover_rows = []
    for row in product_summaries.values():
        if int(row["on_hand_qty"] or 0) <= 0:
            continue

        last_sale_date = row["last_sale_date"]
        dormant_days = (today - last_sale_date).days if last_sale_date else None
        if int(row["sold_qty_30"] or 0) <= 2:
            slow_mover_rows.append(
                {
                    **row,
                    "dormant_days": dormant_days,
                    "dormant_label": (
                        f"{dormant_days} day{'s' if dormant_days != 1 else ''} since last sale"
                        if dormant_days is not None
                        else "No recorded sale yet"
                    ),
                }
            )

    slow_mover_rows_all = sorted(
        slow_mover_rows,
        key=lambda row: (
            int(row["sold_qty_30"] or 0),
            -(int(row["on_hand_qty"] or 0)),
            -(row["dormant_days"] if row["dormant_days"] is not None else 9999),
        ),
    )
    slow_mover_rows = slow_mover_rows_all[:6]

    branch_rows_all = []
    for branch in branch_summaries.values():
        total_month = float(branch["month_sales"]["UZS"] or 0.0) + float(branch["month_sales"]["USD"] or 0.0)
        if branch["low_stock_count"] > 0:
            branch_note = f"{branch['low_stock_count']} low-stock item{'s' if branch['low_stock_count'] != 1 else ''}"
        elif branch["stock_units"] == 0:
            branch_note = "No stock placed in this branch yet"
        elif total_month > 0:
            branch_note = "Active branch with movement this month"
        else:
            branch_note = "Stock is sitting without recent sales"

        branch_rows_all.append(
            {
                **branch,
                "today_sales_display": _format_summary_money_pair(branch["today_sales"], zero_label="0 UZS"),
                "month_sales_display": _format_summary_money_pair(branch["month_sales"], zero_label="0 UZS"),
                "month_profit_display": _format_summary_money_pair(branch["month_profit"], zero_label="0 UZS"),
                "stock_value_display": _format_summary_money_pair(branch["stock_value"], zero_label="0 UZS"),
                "note": branch_note,
            }
        )

    branch_rows_all = sorted(
        branch_rows_all,
        key=lambda row: (
            float(row["month_sales"]["UZS"] or 0.0),
            float(row["month_sales"]["USD"] or 0.0),
            int(row["sold_units_30"] or 0),
        ),
        reverse=True,
    )
    branch_rows = branch_rows_all[:6]

    customer_rows = []
    for customer in customer_summaries.values():
        customer_rows.append(
            {
                "name": customer["name"],
                "phone": customer["phone"],
                "orders_count": len(customer["orders"]),
                "last_seen": customer["last_seen"].strftime("%Y-%m-%d") if customer["last_seen"] else "-",
                "revenue_display": _format_summary_money_pair(customer["revenue"], zero_label="0 UZS"),
                "profit_display": _format_summary_money_pair(customer["profit"], zero_label="0 UZS"),
                "sort_uzs": float(customer["revenue"]["UZS"] or 0.0),
                "sort_usd": float(customer["revenue"]["USD"] or 0.0),
            }
        )

    customer_rows = sorted(
        customer_rows,
        key=lambda row: (row["sort_uzs"], row["sort_usd"], row["orders_count"]),
        reverse=True,
    )[:6]

    today_sales_display = _format_summary_money_pair(sales_totals["today"], zero_label="0 UZS")
    yesterday_sales_display = _format_summary_money_pair(sales_totals["yesterday"], zero_label="0 UZS")
    week_sales_display = _format_summary_money_pair(sales_totals["week"], zero_label="0 UZS")
    month_sales_display = _format_summary_money_pair(sales_totals["month"], zero_label="0 UZS")
    month_profit_display = _format_summary_money_pair(sales_totals["month_profit"], zero_label="0 UZS")
    cash_display = _format_summary_money_pair(cash_totals, zero_label="0 UZS")
    factory_stock_display = _format_summary_money_pair(factory_stock_cost, zero_label="0 UZS")
    shop_stock_display = _format_summary_money_pair(shop_stock_value, zero_label="0 UZS")

    today_change = _format_change_hint(
        float(sales_totals["today"]["UZS"] or 0.0) + float(sales_totals["today"]["USD"] or 0.0),
        float(sales_totals["yesterday"]["UZS"] or 0.0) + float(sales_totals["yesterday"]["USD"] or 0.0),
    )
    month_change = _format_change_hint(
        float(sales_totals["month"]["UZS"] or 0.0) + float(sales_totals["month"]["USD"] or 0.0),
        float(sales_totals["week"]["UZS"] or 0.0) + float(sales_totals["week"]["USD"] or 0.0),
    )

    hero_metrics = [
        {
            "label": translate("revenue_today"),
            "value": today_sales_display["primary"],
            "sub": yesterday_sales_display["primary"],
            "hint": today_change["label"],
            "tone": today_change["tone"],
        },
        {
            "label": translate("revenue_this_month"),
            "value": month_sales_display["primary"],
            "sub": week_sales_display["primary"],
            "hint": month_change["label"],
            "tone": month_change["tone"],
        },
        {
            "label": translate("profit_this_month"),
            "value": month_profit_display["primary"],
            "sub": month_profit_display["secondary"] or translate("profit_this_month_sub"),
            "hint": translate("profit_this_month_hint"),
            "tone": "up",
        },
        {
            "label": translate("open_issues"),
            "value": str(int(command_center["open_count"] or 0)),
            "sub": f"{int(command_center['overdue_count'] or 0)} {translate('overdue')}",
            "hint": translate("pulled_from_live_command_center"),
            "tone": "down" if int(command_center["overdue_count"] or 0) else "flat",
        },
    ]

    summary_cards = [
        {
            "label": translate("factory_stock_cost"),
            "value": factory_stock_display["primary"],
            "sub": factory_stock_display["secondary"] or translate("ready_goods_still_in_factory"),
        },
        {
            "label": translate("shop_stock_sell_value"),
            "value": shop_stock_display["primary"],
            "sub": shop_stock_display["secondary"] or translate("goods_sitting_in_branches"),
        },
        {
            "label": translate("cash_on_hand"),
            "value": cash_display["primary"],
            "sub": cash_display["secondary"] or translate("recorded_cash_balances"),
        },
        {
            "label": translate("produced_today"),
            "value": f"{int(produced_today_total or 0):,}",
            "sub": translate("produced_today_models").format(count=int(produced_today_models or 0)),
        },
        {
            "label": translate("produced_this_week"),
            "value": f"{int(produced_week_total or 0):,}",
            "sub": translate("current_seven_day_factory_output"),
        },
        {
            "label": translate("transferred_yesterday"),
            "value": f"{int(yesterday_transfer_total or 0):,}",
            "sub": translate("factory_to_shop_movement_yesterday"),
        },
        {
            "label": translate("team_footprint"),
            "value": f"{len(workspace_team):,}",
            "sub": translate("team_footprint_sub").format(shops=len(workspace_shops), products=len(product_rows)),
        },
        {
            "label": translate("shop_low_stock"),
            "value": f"{int(shop_low_stock_count or 0):,}",
            "sub": translate("items_below_5_pcs_across_linked_branches"),
        },
    ]

    quick_actions = [
        {
            "title": translate("open_command_center"),
            "subtitle": translate("open_command_center_sub"),
            "href": url_for("main.command_center"),
        },
        {
            "title": translate("manage_workspace"),
            "subtitle": translate("manage_workspace_sub"),
            "href": url_for("main.workspace_details"),
        },
        {
            "title": translate("review_products"),
            "subtitle": translate("review_products_sub"),
            "href": url_for("main.business_summary_products"),
        },
        {
            "title": translate("workspace_review_branches"),
            "subtitle": translate("workspace_review_branches_sub"),
            "href": url_for("main.business_summary_branches"),
        },
        {
            "title": translate("more_tools"),
            "subtitle": translate("more_tools_sub"),
            "href": url_for("main.more"),
        },
    ]

    return {
        "workspace_page_name": display_value(getattr(workspace, "name", None), fallback="Workspace"),
        "workspace_role_label": _format_role_label(getattr(current_user, "role", None)),
        "workspace_owner_name": get_user_display_name(workspace_owner) if workspace_owner else "-",
        "workspace_owner_login": display_value(getattr(workspace_owner, "username", None)),
        "health_score": int(health_score),
        "health_label": health_label,
        "health_copy": health_copy,
        "health_tone": health_tone,
        "hero_metrics": hero_metrics,
        "summary_cards": summary_cards,
        "attention_items": command_center["attention_items"][:4],
        "command_center_open_count": int(command_center["open_count"] or 0),
        "command_center_overdue_count": int(command_center["overdue_count"] or 0),
        "setup_items": setup_items,
        "setup_remaining": setup_remaining,
        "top_product_rows": top_product_rows,
        "top_product_rows_all": top_product_rows_all,
        "slow_mover_rows": slow_mover_rows,
        "slow_mover_rows_all": slow_mover_rows_all,
        "branch_rows": branch_rows,
        "branch_rows_all": branch_rows_all,
        "customer_rows": customer_rows,
        "shop_low_stock_items": shop_low_stock_items,
        "today_sales_display": today_sales_display,
        "month_sales_display": month_sales_display,
        "month_profit_display": month_profit_display,
        "cash_display": cash_display,
        "quick_actions": quick_actions,
        "branch_summary_href": url_for("main.business_summary_branches"),
        "product_summary_href": url_for("main.business_summary_products"),
    }


def _build_manager_dashboard(factory_id: int):
    factory_uzs, factory_usd = product_service.total_stock_value(factory_id=factory_id)
    shop_uzs, shop_usd = product_service.shop_stock_totals(factory_id=factory_id)

    shop_low_stock_count, shop_low_stock_items = _get_shop_low_stock(factory_id=factory_id)
    yesterday_transfer_total = _get_yesterday_transfer_total(factory_id=factory_id)

    produced_today_total, produced_today_models, prod_today_rows = _get_production_today_summary(factory_id)
    produced_week_total = _get_production_week_total(factory_id)

    cash_total_uzs, cash_total_usd = _calc_cash_totals(factory_id=factory_id)
    sales_stats = _get_sales_dashboard_stats(factory_id=factory_id)
    setup_items = _build_dashboard_setup_items(factory_id)
    command_center = _build_command_center_snapshot(factory_id)

    return {
        "factory_uzs": factory_uzs,
        "shop_uzs": shop_uzs,
        "total_uzs": factory_uzs + shop_uzs,

        "factory_usd": factory_usd,
        "shop_usd": shop_usd,
        "total_usd": factory_usd + shop_usd,

        "cash_total_uzs": cash_total_uzs,
        "cash_total_usd": cash_total_usd,

        "shop_low_stock_count": shop_low_stock_count,
        "shop_low_stock_items": shop_low_stock_items,

        "yesterday_transfer_total": yesterday_transfer_total,

        "produced_today_total": produced_today_total,
        "produced_today_models": produced_today_models,
        "produced_week_total": produced_week_total,
        "prod_today_rows": prod_today_rows,

        "today_sales_uzs": sales_stats["today_sales_uzs"],
        "yesterday_sales_uzs": sales_stats["yesterday_sales_uzs"],
        "week_sales_uzs": sales_stats["week_sales_uzs"],
        "top_selling_models": sales_stats["top_selling_models"],
        "setup_items": setup_items,
        "setup_remaining": sum(1 for item in setup_items if not item["done"]),
        "command_center_attention_items": command_center["attention_items"],
        "command_center_open_count": command_center["open_count"],
        "command_center_live_count": command_center["live_count"],
        "command_center_manual_count": command_center["manual_count"],
        "command_center_my_count": command_center["my_count"],
        "command_center_urgent_count": command_center["urgent_count"],
        "command_center_overdue_count": command_center["overdue_count"],
        "command_center_done_today_count": command_center["done_today_count"],
        "command_center_manual_items": command_center["manual_items"][:3],
        "command_center_my_items": command_center["my_items"][:3],
    }


def _format_role_label(role_value) -> str:
    role_text = str(role_value or "").strip().replace("_", " ")
    return role_text.title() if role_text else "-"


def _build_workspace_registration_rows(workspace):
    return [
        {
            "label": "Business name",
            "value": display_value(getattr(workspace, "name", None), fallback="Workspace"),
            "is_ready": True,
        },
        {
            "label": "Owner name",
            "value": display_value(getattr(workspace, "owner_name", None)),
            "is_ready": bool(display_value(getattr(workspace, "owner_name", None), fallback="")),
        },
        {
            "label": "Location",
            "value": display_value(getattr(workspace, "location", None)),
            "is_ready": bool(display_value(getattr(workspace, "location", None), fallback="")),
        },
        {
            "label": "Phone",
            "value": display_value(getattr(workspace, "phone", None)),
            "is_ready": bool(display_value(getattr(workspace, "phone", None), fallback="")),
        },
        {
            "label": "Internal note",
            "value": display_value(getattr(workspace, "note", None)),
            "is_ready": bool(display_value(getattr(workspace, "note", None), fallback="")),
        },
    ]


def _build_workspace_owner_points(team_count: int, linked_shop_count: int, missing_fields):
    team_label = "team member" if team_count == 1 else "team members"
    shop_label = "shop" if linked_shop_count == 1 else "shops"

    points = [
        f"{team_count} {team_label} currently belong to this workspace.",
        (
            f"{linked_shop_count} linked {shop_label} are connected to this business."
            if linked_shop_count
            else "No linked shops are connected yet, so this workspace still looks centralized."
        ),
    ]

    if missing_fields:
        points.append(
            "The registration card still has missing basics: "
            + ", ".join(missing_fields)
            + "."
        )
    else:
        points.append("Core registration details are filled, which is what I want before scaling access.")

    return points


def _build_dashboard_setup_items(factory_id: int):
    workspace = Factory.query.get(factory_id) if factory_id else None
    linked_shop_count = 0
    if factory_id:
        linked_shop_count = (
            db.session.query(func.count(func.distinct(ShopFactoryLink.shop_id)))
            .filter(ShopFactoryLink.factory_id == factory_id)
            .scalar()
            or 0
        )

    team_count = User.query.filter(User.factory_id == factory_id).count() if factory_id else 0
    product_count = Product.query.filter(Product.factory_id == factory_id).count() if factory_id else 0
    telegram_linked = bool(getattr(current_user, "telegram_links", None))

    return [
        {
            "label": translate("dashboard_setup_business_profile_title"),
            "done": bool(
                workspace
                and getattr(workspace, "name", None)
                and getattr(workspace, "owner_name", None)
                and getattr(workspace, "location", None)
                and getattr(workspace, "phone", None)
            ),
            "copy": translate("dashboard_setup_business_profile_copy"),
            "href": url_for("main.workspace_details"),
            "action": translate("open_workspace"),
        },
        {
            "label": translate("dashboard_setup_first_shop_connected_title"),
            "done": linked_shop_count > 0,
            "copy": translate("dashboard_setup_first_shop_connected_copy"),
            "href": url_for("main.workspace_details"),
            "action": translate("dashboard_action_review_shops"),
        },
        {
            "label": translate("dashboard_setup_team_added_title"),
            "done": team_count > 1,
            "copy": translate("dashboard_setup_team_added_copy"),
            "href": url_for("main.workspace_details"),
            "action": translate("dashboard_action_manage_team"),
        },
        {
            "label": translate("dashboard_setup_first_products_added_title"),
            "done": product_count > 0,
            "copy": translate("dashboard_setup_first_products_added_copy"),
            "href": url_for("products.list_products"),
            "action": translate("open_products"),
        },
        {
            "label": translate("dashboard_setup_telegram_linked_title"),
            "done": telegram_linked,
            "copy": translate("dashboard_setup_telegram_linked_copy"),
            "href": url_for("main.profile_overview"),
            "action": translate("workspace_open_profile"),
        },
    ]


def _task_priority_sort_key(priority_value: str | None) -> int:
    order = {
        "urgent": 0,
        "high": 1,
        "medium": 2,
        "low": 3,
    }
    return order.get(str(priority_value or "").strip().lower(), 99)


def _task_priority_label(priority_value: str | None) -> str:
    text = str(priority_value or "").strip().lower()
    key = f"priority_{text}"
    if not text:
        return translate("priority_medium")
    label = translate(key)
    return label if label != key else text.title()


def _task_status_label(status_value: str | None) -> str:
    text = str(status_value or "").strip().replace("_", " ")
    key = f"status_{text.replace(' ', '_')}"
    if not text:
        return translate("status_open")
    label = translate(key)
    return label if label != key else text.title()


def _task_due_label(due_date) -> str:
    if not due_date:
        return translate("no_due_date")

    try:
        days_left = (due_date - date.today()).days
    except Exception:
        return str(due_date)

    if days_left < 0:
        return translate("overdue_by_days").format(days=abs(days_left))
    if days_left == 0:
        return translate("due_today")
    if days_left == 1:
        return translate("due_tomorrow")
    return translate("due_in_days").format(days=days_left)


def _can_manage_command_center(workspace) -> bool:
    return _can_manage_workspace_team(workspace)


def _can_view_operational_task(task, workspace) -> bool:
    if not task or not workspace or not current_user.is_authenticated:
        return False

    if getattr(current_user, "is_superadmin", False):
        return True

    if _can_manage_command_center(workspace):
        return True

    if getattr(task, "assigned_user_id", None) == getattr(current_user, "id", None):
        return True

    task_target_role = str(getattr(task, "target_role", None) or "").strip().lower()
    if task_target_role and task_target_role == str(getattr(current_user, "role", "")).strip().lower():
        return True

    return False


def _can_update_operational_task(task, workspace) -> bool:
    if not _can_view_operational_task(task, workspace):
        return False

    if getattr(current_user, "is_superadmin", False) or _can_manage_command_center(workspace):
        return True

    current_role = str(getattr(current_user, "role", None) or "").strip().lower()
    task_target_role = str(getattr(task, "target_role", None) or "").strip().lower()
    return bool(
        getattr(task, "assigned_user_id", None) == getattr(current_user, "id", None)
        or (not getattr(task, "assigned_user_id", None) and task_target_role and task_target_role == current_role)
    )


def _system_task_meta(source_type):
    return SYSTEM_TASK_SOURCE_META.get(str(source_type or "").strip().lower(), {})


def _default_operational_task_due_date(priority_value: str):
    offsets = {
        "urgent": 0,
        "high": 1,
        "medium": 3,
        "low": 7,
    }
    return date.today() + timedelta(days=offsets.get(str(priority_value or "").strip().lower(), 3))


def _suggest_operational_task_assignee(workspace_id: int | None, target_role: str | None):
    role = str(target_role or "").strip().lower()
    if not workspace_id or not role:
        return None

    if role == "admin":
        return _get_workspace_owner_user(workspace_id)

    candidates = [
        user
        for user in _get_workspace_team_users(workspace_id)
        if str(getattr(user, "role", None) or "").strip().lower() == role
    ]
    return candidates[0] if len(candidates) == 1 else None


def _build_system_task_blueprints(factory_id: int):
    workspace = Factory.query.get(factory_id) if factory_id else None
    if not workspace:
        return []

    today = date.today()
    start_of_day = datetime.combine(today, datetime.min.time())
    end_of_day = start_of_day + timedelta(days=1)

    setup_items = _build_dashboard_setup_items(factory_id)
    setup_remaining = sum(1 for item in setup_items if not item["done"])
    linked_shop_count = (
        db.session.query(func.count(func.distinct(ShopFactoryLink.shop_id)))
        .filter(ShopFactoryLink.factory_id == factory_id)
        .scalar()
        or 0
    )
    team_count = User.query.filter(User.factory_id == factory_id).count()
    product_count = Product.query.filter(Product.factory_id == factory_id).count()
    shop_low_stock_count, _items = _get_shop_low_stock(factory_id=factory_id)
    produced_today_total, _produced_today_models, _rows = _get_production_today_summary(factory_id)
    factory_uzs, _factory_usd = product_service.total_stock_value(factory_id=factory_id)
    shop_uzs, _shop_usd = product_service.shop_stock_totals(factory_id=factory_id)
    total_stock_uzs = float(factory_uzs or 0) + float(shop_uzs or 0)
    factory_pct = int((float(factory_uzs or 0) * 100) / total_stock_uzs) if total_stock_uzs else 0
    cash_total_uzs, cash_total_usd = _calc_cash_totals(factory_id)
    workspace_owner = _get_workspace_owner_user(workspace)
    owner_telegram_linked = bool(workspace_owner and getattr(workspace_owner, "telegram_links", None))
    sales_today_count = (
        Sale.query
        .join(Product, Product.id == Sale.product_id)
        .filter(
            Product.factory_id == factory_id,
            Sale.date == today,
        )
        .count()
    )
    cash_records_today_count = (
        CashRecord.query
        .filter(
            CashRecord.factory_id == factory_id,
            CashRecord.date == today,
        )
        .count()
    )
    today_transfer_count = (
        StockMovement.query
        .filter(
            StockMovement.factory_id == factory_id,
            StockMovement.timestamp >= start_of_day,
            StockMovement.timestamp < end_of_day,
            StockMovement.movement_type.in_(("factory_to_shop", "factory_to_shop_for_order")),
        )
        .count()
    )
    pending_orders_query = ShopOrder.query.filter(
        ShopOrder.factory_id == factory_id,
        ShopOrder.status == "pending",
    )
    pending_order_count = pending_orders_query.count()
    oldest_pending_order = pending_orders_query.order_by(ShopOrder.created_at.asc()).first()
    oldest_pending_days = 0
    if oldest_pending_order and getattr(oldest_pending_order, "created_at", None):
        oldest_pending_days = max((today - oldest_pending_order.created_at.date()).days, 0)
    ready_order_count = (
        ShopOrder.query
        .filter(
            ShopOrder.factory_id == factory_id,
            ShopOrder.status == "ready",
        )
        .count()
    )

    blueprints = []

    def add_rule(
        source_key: str,
        *,
        title: str,
        description: str,
        href: str,
        priority: str,
        target_role: str,
        due_days: int | None = None,
    ):
        assignee = _suggest_operational_task_assignee(factory_id, target_role)
        blueprints.append(
            {
                "source_key": source_key,
                "source_id": 0,
                "title": title,
                "description": description,
                "action_url": href,
                "priority": priority,
                "target_role": target_role,
                "assigned_user_id": getattr(assignee, "id", None),
                "due_date": today + timedelta(days=due_days) if due_days is not None else _default_operational_task_due_date(priority),
            }
        )

    def _plural_title(key_singular: str, key_plural: str, count: int) -> str:
        return translate(key_singular if count == 1 else key_plural).format(count=count)

    if shop_low_stock_count > 0:
        add_rule(
            "shop_low_stock",
            title=_plural_title(
                "dashboard_live_shop_low_stock_title_singular",
                "dashboard_live_shop_low_stock_title_plural",
                shop_low_stock_count,
            ),
            description=translate("dashboard_live_shop_low_stock_copy"),
            href=url_for("shop.transfer_to_shop"),
            priority="urgent" if shop_low_stock_count > 3 else "high",
            target_role="manager",
            due_days=0,
        )

    if produced_today_total == 0:
        add_rule(
            "no_production_today",
            title=translate("dashboard_live_no_production_today_title"),
            description=translate("dashboard_live_no_production_today_copy"),
            href=url_for("factory.produce_today"),
            priority="high",
            target_role="manager",
            due_days=0,
        )

    if setup_remaining > 0:
        add_rule(
            "workspace_setup",
            title=_plural_title(
                "dashboard_live_workspace_setup_title_singular",
                "dashboard_live_workspace_setup_title_plural",
                setup_remaining,
            ),
            description=translate("dashboard_live_workspace_setup_copy"),
            href=url_for("main.workspace_details"),
            priority="high" if setup_remaining >= 3 else "medium",
            target_role="admin",
        )

    if linked_shop_count == 0:
        add_rule(
            "no_shop_linked",
            title=translate("dashboard_live_no_shop_linked_title"),
            description=translate("dashboard_live_no_shop_linked_copy"),
            href=url_for("main.workspace_details", panel="workspace"),
            priority="medium",
            target_role="admin",
        )

    if team_count <= 1:
        add_rule(
            "owner_only_account",
            title=translate("dashboard_live_only_owner_active_title"),
            description=translate("dashboard_live_only_owner_active_copy"),
            href=url_for("main.workspace_details", panel="team"),
            priority="medium",
            target_role="admin",
        )

    if product_count == 0:
        add_rule(
            "no_products",
            title=translate("dashboard_live_no_products_registered_title"),
            description=translate("dashboard_live_no_products_registered_copy"),
            href=url_for("products.list_products"),
            priority="high",
            target_role="manager",
        )

    if factory_pct >= 70 and total_stock_uzs > 0 and linked_shop_count > 0:
        add_rule(
            "factory_stock_heavy",
            title=translate("dashboard_live_factory_stock_heavy_title").format(percent=factory_pct),
            description=translate("dashboard_live_factory_stock_heavy_copy"),
            href=url_for("shop.transfer_to_shop"),
            priority="medium",
            target_role="manager",
        )

    if not owner_telegram_linked:
        add_rule(
            "owner_telegram_missing",
            title=translate("dashboard_live_owner_telegram_missing_title"),
            description=translate("dashboard_live_owner_telegram_missing_copy"),
            href=url_for("main.profile_overview", panel="telegram"),
            priority="low",
            target_role="admin",
        )

    if product_count > 0 and cash_total_uzs == 0 and cash_total_usd == 0:
        add_rule(
            "no_cash_records",
            title=translate("dashboard_live_no_cash_records_title"),
            description=translate("dashboard_live_no_cash_records_copy"),
            href=url_for("cash.list_cash"),
            priority="low",
            target_role="accountant",
        )

    if sales_today_count > 0 and cash_records_today_count == 0:
        add_rule(
            "sales_without_cash_today",
            title=_plural_title(
                "dashboard_live_sales_without_cash_title_singular",
                "dashboard_live_sales_without_cash_title_plural",
                sales_today_count,
            ),
            description=translate("dashboard_live_sales_without_cash_copy"),
            href=url_for("cash.list_cash"),
            priority="high",
            target_role="accountant",
            due_days=0,
        )

    if pending_order_count > 0:
        add_rule(
            "pending_shop_orders",
            title=_plural_title(
                "dashboard_live_pending_shop_orders_title_singular",
                "dashboard_live_pending_shop_orders_title_plural",
                pending_order_count,
            ),
            description=translate("dashboard_live_pending_shop_orders_copy"),
            href=url_for("shop.factory_pending_orders"),
            priority="high" if oldest_pending_days >= 2 or pending_order_count >= 5 else "medium",
            target_role="manager",
            due_days=0 if oldest_pending_days >= 2 else 1,
        )

    if ready_order_count > 0:
        add_rule(
            "ready_shop_orders",
            title=_plural_title(
                "dashboard_live_ready_shop_orders_title_singular",
                "dashboard_live_ready_shop_orders_title_plural",
                ready_order_count,
            ),
            description=translate("dashboard_live_ready_shop_orders_copy"),
            href=url_for("shop.list_shop_orders", status="ready"),
            priority="high" if ready_order_count >= 4 else "medium",
            target_role="shop",
            due_days=0,
        )

    if datetime.now().hour >= 18 and (sales_today_count > 0 or produced_today_total > 0 or today_transfer_count > 0):
        add_rule(
            "daily_close_review",
            title=f"Run the close review for {today.strftime('%d %b %Y')}",
            description="Check that production, orders, transfers, and cash all match the business day before everyone leaves it behind.",
            href=url_for("main.command_center"),
            priority="medium",
            target_role="admin",
            due_days=0,
        )

    return blueprints


def _sync_system_generated_operational_tasks(factory_id: int | None):
    if not factory_id:
        return {"created": 0, "updated": 0, "reopened": 0, "resolved": 0}

    workspace = Factory.query.get(factory_id)
    if not workspace:
        return {"created": 0, "updated": 0, "reopened": 0, "resolved": 0}

    blueprints = _build_system_task_blueprints(factory_id)
    today = date.today()
    now = datetime.utcnow()
    changed = False
    stats = {"created": 0, "updated": 0, "reopened": 0, "resolved": 0}

    system_rows = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.is_system_generated.is_(True),
        )
        .all()
    )
    rows_by_key = {}
    for row in sorted(
        system_rows,
        key=lambda task: (
            getattr(task, "updated_at", None) or getattr(task, "created_at", None) or datetime.min,
            getattr(task, "id", 0),
        ),
        reverse=True,
    ):
        key = (str(getattr(row, "source_type", None) or "").strip().lower(), int(getattr(row, "source_id", None) or 0))
        rows_by_key.setdefault(key, []).append(row)

    active_keys = set()

    for blueprint in blueprints:
        key = (str(blueprint["source_key"]).strip().lower(), int(blueprint.get("source_id") or 0))
        active_keys.add(key)
        history = rows_by_key.get(key, [])
        active_task = next((task for task in history if getattr(task, "status", None) in {"open", "in_progress"}), None)
        dismissed_today = next(
            (
                task for task in history
                if getattr(task, "status", None) == "dismissed"
                and getattr(task, "closed_at", None)
                and task.closed_at.date() == today
            ),
            None,
        )

        if active_task is None and dismissed_today is not None:
            continue

        if active_task is None:
            reusable = next((task for task in history if getattr(task, "status", None) in {"done", "dismissed"}), None)
            if reusable is not None:
                active_task = reusable
                active_task.status = "open"
                active_task.closed_at = None
                active_task.closed_by_id = None
                active_task.updated_at = now
                stats["reopened"] += 1
                changed = True
            else:
                active_task = OperationalTask(
                    factory_id=factory_id,
                    task_type="system_rule",
                    source_type=blueprint["source_key"],
                    source_id=int(blueprint.get("source_id") or 0),
                    status="open",
                    is_system_generated=True,
                )
                db.session.add(active_task)
                stats["created"] += 1
                changed = True

        duplicate_open_rows = [
            task for task in history
            if task is not active_task and getattr(task, "status", None) in {"open", "in_progress"}
        ]
        for duplicate in duplicate_open_rows:
            duplicate.status = "done"
            duplicate.closed_at = now
            duplicate.closed_by_id = None
            duplicate.updated_at = now
            stats["resolved"] += 1
            changed = True

        target_fields = {
            "title": blueprint["title"],
            "description": blueprint["description"],
            "action_url": blueprint["action_url"],
            "priority": blueprint["priority"],
            "target_role": blueprint["target_role"],
            "assigned_user_id": blueprint.get("assigned_user_id"),
            "due_date": blueprint["due_date"],
            "task_type": "system_rule",
            "source_type": blueprint["source_key"],
            "source_id": int(blueprint.get("source_id") or 0),
            "is_system_generated": True,
        }

        field_changed = False
        for field_name, field_value in target_fields.items():
            if getattr(active_task, field_name, None) != field_value:
                setattr(active_task, field_name, field_value)
                field_changed = True

        if field_changed:
            active_task.updated_at = now
            if stats["created"] == 0 or getattr(active_task, "id", None):
                stats["updated"] += 1
            changed = True

    for row in system_rows:
        key = (str(getattr(row, "source_type", None) or "").strip().lower(), int(getattr(row, "source_id", None) or 0))
        if getattr(row, "status", None) in {"open", "in_progress"} and key not in active_keys:
            row.status = "done"
            row.closed_at = now
            row.closed_by_id = None
            row.updated_at = now
            stats["resolved"] += 1
            changed = True

    if changed:
        db.session.commit()

    return stats


def _build_live_command_items(factory_id: int):
    workspace = Factory.query.get(factory_id) if factory_id else None
    if not workspace:
        return []

    setup_items = _build_dashboard_setup_items(factory_id)
    setup_remaining = sum(1 for item in setup_items if not item["done"])
    linked_shop_count = (
        db.session.query(func.count(func.distinct(ShopFactoryLink.shop_id)))
        .filter(ShopFactoryLink.factory_id == factory_id)
        .scalar()
        or 0
    )
    team_count = User.query.filter(User.factory_id == factory_id).count()
    product_count = Product.query.filter(Product.factory_id == factory_id).count()
    shop_low_stock_count, _items = _get_shop_low_stock(factory_id=factory_id)
    produced_today_total, _produced_today_models, _rows = _get_production_today_summary(factory_id)
    factory_uzs, _factory_usd = product_service.total_stock_value(factory_id=factory_id)
    shop_uzs, _shop_usd = product_service.shop_stock_totals(factory_id=factory_id)
    total_stock_uzs = float(factory_uzs or 0) + float(shop_uzs or 0)
    factory_pct = int((float(factory_uzs or 0) * 100) / total_stock_uzs) if total_stock_uzs else 0
    cash_total_uzs, cash_total_usd = _calc_cash_totals(factory_id)
    workspace_owner = _get_workspace_owner_user(workspace)
    owner_telegram_linked = bool(workspace_owner and getattr(workspace_owner, "telegram_links", None))

    def _plural_title(key_singular: str, key_plural: str, count: int) -> str:
        return translate(key_singular if count == 1 else key_plural).format(count=count)

    items = []

    if shop_low_stock_count > 0:
        items.append(
            {
                "kind": "live",
                "priority": "urgent" if shop_low_stock_count > 3 else "high",
                "title": _plural_title(
                    "dashboard_live_shop_low_stock_title_singular",
                    "dashboard_live_shop_low_stock_title_plural",
                    shop_low_stock_count,
                ),
                "description": translate("dashboard_live_shop_low_stock_copy"),
                "href": url_for("shop.transfer_to_shop"),
                "action_label": translate("dashboard_live_shop_low_stock_action_label"),
                "target_roles": {"admin", "manager"},
                "group": "Stock",
            }
        )

    if produced_today_total == 0:
        items.append(
            {
                "kind": "live",
                "priority": "high",
                "title": translate("dashboard_live_no_production_today_title"),
                "description": translate("dashboard_live_no_production_today_copy"),
                "href": url_for("factory.produce_today"),
                "action_label": translate("dashboard_live_no_production_today_action_label"),
                "target_roles": {"admin", "manager"},
                "group": "Production",
            }
        )

    if setup_remaining > 0:
        items.append(
            {
                "kind": "live",
                "priority": "high" if setup_remaining >= 3 else "medium",
                "title": _plural_title(
                    "dashboard_live_workspace_setup_title_singular",
                    "dashboard_live_workspace_setup_title_plural",
                    setup_remaining,
                ),
                "description": translate("dashboard_live_workspace_setup_copy"),
                "href": url_for("main.workspace_details"),
                "action_label": translate("dashboard_live_workspace_setup_action_label"),
                "target_roles": {"admin"},
                "group": "Workspace",
            }
        )

    if linked_shop_count == 0:
        items.append(
            {
                "kind": "live",
                "priority": "medium",
                "title": translate("dashboard_live_no_shop_linked_title"),
                "description": translate("dashboard_live_no_shop_linked_copy"),
                "href": url_for("main.workspace_details", panel="workspace"),
                "action_label": translate("dashboard_live_no_shop_linked_action_label"),
                "target_roles": {"admin"},
                "group": "Expansion",
            }
        )

    if team_count <= 1:
        items.append(
            {
                "kind": "live",
                "priority": "medium",
                "title": translate("dashboard_live_only_owner_active_title"),
                "description": translate("dashboard_live_only_owner_active_copy"),
                "href": url_for("main.workspace_details", panel="team"),
                "action_label": translate("dashboard_live_only_owner_active_action_label"),
                "target_roles": {"admin"},
                "group": "Team",
            }
        )

    if product_count == 0:
        items.append(
            {
                "kind": "live",
                "priority": "high",
                "title": translate("dashboard_live_no_products_registered_title"),
                "description": translate("dashboard_live_no_products_registered_copy"),
                "href": url_for("products.list_products"),
                "action_label": translate("dashboard_live_no_products_registered_action_label"),
                "target_roles": {"admin", "manager"},
                "group": "Catalog",
            }
        )

    if factory_pct >= 70 and total_stock_uzs > 0 and linked_shop_count > 0:
        items.append(
            {
                "kind": "live",
                "priority": "medium",
                "title": translate("dashboard_live_factory_stock_heavy_title").format(percent=factory_pct),
                "description": translate("dashboard_live_factory_stock_heavy_copy"),
                "href": url_for("shop.transfer_to_shop"),
                "action_label": translate("dashboard_live_factory_stock_heavy_action_label"),
                "target_roles": {"admin", "manager"},
                "group": "Stock flow",
            }
        )

    if not owner_telegram_linked:
        items.append(
            {
                "kind": "live",
                "priority": "low",
                "title": translate("dashboard_live_owner_telegram_missing_title"),
                "description": translate("dashboard_live_owner_telegram_missing_copy"),
                "href": url_for("main.profile_overview", panel="telegram"),
                "action_label": translate("dashboard_live_owner_telegram_missing_action_label"),
                "target_roles": {"admin"},
                "group": "Notifications",
            }
        )

    if product_count > 0 and cash_total_uzs == 0 and cash_total_usd == 0:
        items.append(
            {
                "kind": "live",
                "priority": "low",
                "title": translate("dashboard_live_no_cash_records_title"),
                "description": translate("dashboard_live_no_cash_records_copy"),
                "href": url_for("cash.list_cash"),
                "action_label": translate("dashboard_live_no_cash_records_action_label"),
                "target_roles": {"admin", "accountant"},
                "group": "Cash",
            }
        )

    current_role = str(getattr(current_user, "role", None) or "").strip().lower()
    if not (_can_manage_command_center(workspace) or getattr(current_user, "is_superadmin", False)):
        items = [
            item
            for item in items
            if not item.get("target_roles") or current_role in item["target_roles"]
        ]

    return sorted(
        items,
        key=lambda item: (
            _task_priority_sort_key(item.get("priority")),
            item.get("title", ""),
        ),
    )


def _serialize_operational_task(task):
    assignee_name = None
    if getattr(task, "assigned_user", None):
        assignee_name = get_user_display_name(task.assigned_user)
    elif getattr(task, "target_role", None):
        assignee_name = _format_role_label(task.target_role)
    else:
        assignee_name = "Workspace team"

    origin = "system" if bool(getattr(task, "is_system_generated", False)) else "manual"
    system_meta = _system_task_meta(getattr(task, "source_type", None)) if origin == "system" else {}
    if origin == "system":
        group_label = translate(system_meta.get("group_key")) if system_meta.get("group_key") else system_meta.get("group") or translate("system_source")
    elif origin == "manual":
        group_label = translate("manual_source")
    else:
        group_label = translate("workspace")
    due_label = _task_due_label(getattr(task, "due_date", None))
    summary_parts = []
    if origin == "system" and group_label:
        summary_parts.append(group_label)
    if assignee_name:
        summary_parts.append(assignee_name)
    if due_label != translate("no_due_date"):
        summary_parts.append(due_label)

    return {
        "kind": "task",
        "origin": origin,
        "origin_label": translate("system_task") if origin == "system" else translate("manual_task"),
        "id": task.id,
        "title": display_value(getattr(task, "title", None), fallback=translate("task_fallback_title")),
        "description": display_value(getattr(task, "description", None), fallback=translate("task_fallback_description")),
        "priority": str(getattr(task, "priority", None) or "medium").strip().lower(),
        "priority_label": _task_priority_label(getattr(task, "priority", None)),
        "status": str(getattr(task, "status", None) or "open").strip().lower(),
        "status_label": _task_status_label(getattr(task, "status", None)),
        "assignee_name": assignee_name,
        "due_label": due_label,
        "summary_line": " - ".join(summary_parts) if summary_parts else due_label,
        "group": group_label,
        "href": (getattr(task, "action_url", None) or "").strip() or url_for("main.command_center"),
        "action_label": (
            translate(system_meta.get("action_label_key"))
            if origin == "system" and system_meta.get("action_label_key")
            else (translate("open_link") if getattr(task, "action_url", None) else translate("view_inbox"))
        ),
        "created_at": getattr(task, "created_at", None),
        "is_system_generated": bool(getattr(task, "is_system_generated", False)),
        "is_overdue": bool(getattr(task, "due_date", None) and getattr(task, "due_date", None) < date.today()),
    }


def _get_visible_operational_tasks(
    factory_id: int | None,
    *,
    statuses: tuple[str, ...] = ("open", "in_progress"),
    limit: int | None = None,
    is_system_generated: bool | None = None,
):
    if not factory_id:
        return []

    workspace = Factory.query.get(factory_id)
    query = OperationalTask.query.filter(OperationalTask.factory_id == factory_id)
    if statuses:
        query = query.filter(OperationalTask.status.in_(statuses))
    if is_system_generated is True:
        query = query.filter(OperationalTask.is_system_generated.is_(True))
    elif is_system_generated is False:
        query = query.filter(OperationalTask.is_system_generated.is_(False))

    rows = query.all()
    rows = [row for row in rows if _can_view_operational_task(row, workspace)]
    rows = sorted(
        rows,
        key=lambda task: (
            _task_priority_sort_key(getattr(task, "priority", None)),
            0 if getattr(task, "status", None) == "in_progress" else 1,
            getattr(task, "due_date", None) or date.max,
            getattr(task, "created_at", None) or datetime.min,
            getattr(task, "id", 0),
        ),
    )
    if limit is not None:
        rows = rows[:limit]
    return rows


def _get_my_operational_tasks(factory_id: int | None, *, limit: int | None = None):
    if not factory_id or not current_user.is_authenticated:
        return []

    workspace = Factory.query.get(factory_id)
    rows = _get_visible_operational_tasks(factory_id)
    current_role = str(getattr(current_user, "role", None) or "").strip().lower()
    direct_rows = []
    fallback_rows = []

    for row in rows:
        task_target_role = str(getattr(row, "target_role", None) or "").strip().lower()
        if getattr(row, "assigned_user_id", None) == getattr(current_user, "id", None):
            direct_rows.append(row)
            continue
        if task_target_role and task_target_role == current_role:
            direct_rows.append(row)
            continue
        fallback_rows.append(row)

    if _can_manage_command_center(workspace) or getattr(current_user, "is_superadmin", False):
        seen_ids = {getattr(task, "id", None) for task in direct_rows}
        target_rows = fallback_rows if not direct_rows else fallback_rows
        for row in target_rows:
            row_id = getattr(row, "id", None)
            if row_id in seen_ids:
                continue
            direct_rows.append(row)
            seen_ids.add(row_id)
            if limit is not None and len(direct_rows) >= limit:
                break

    if limit is not None:
        direct_rows = direct_rows[:limit]
    return direct_rows


def _build_command_center_backlog_rows(factory_id: int | None):
    if not factory_id:
        return []

    workspace = Factory.query.get(factory_id)
    if not workspace or not (_can_manage_command_center(workspace) or getattr(current_user, "is_superadmin", False)):
        return []

    rows = _get_visible_operational_tasks(factory_id)
    if not rows:
        return []

    role_rows = {}
    today = date.today()

    for task in rows:
        role_key = str(getattr(task, "target_role", None) or "").strip().lower()
        if not role_key and getattr(task, "assigned_user", None):
            role_key = str(getattr(task.assigned_user, "role", None) or "").strip().lower()
        if not role_key:
            role_key = "unassigned"

        bucket = role_rows.setdefault(
            role_key,
            {
                "role": role_key,
                "label": "Unassigned" if role_key == "unassigned" else _format_role_label(role_key),
                "open_count": 0,
                "urgent_count": 0,
                "overdue_count": 0,
            },
        )
        bucket["open_count"] += 1
        if str(getattr(task, "priority", None) or "").strip().lower() in {"urgent", "high"}:
            bucket["urgent_count"] += 1
        if getattr(task, "due_date", None) and task.due_date < today:
            bucket["overdue_count"] += 1

    return sorted(
        role_rows.values(),
        key=lambda row: (
            999 if row["role"] == "unassigned" else _workspace_role_sort_key(row["role"]),
            row["label"].lower(),
        ),
    )


def _get_done_operational_tasks(factory_id: int | None, *, limit: int = 6):
    if not factory_id:
        return []

    workspace = Factory.query.get(factory_id)
    rows = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.status.in_(("done", "dismissed")),
        )
        .all()
    )
    rows = [row for row in rows if _can_view_operational_task(row, workspace)]
    rows = sorted(
        rows,
        key=lambda task: (
            getattr(task, "closed_at", None) or datetime.min,
            getattr(task, "updated_at", None) or datetime.min,
            getattr(task, "id", 0),
        ),
        reverse=True,
    )
    return rows[:limit]


def _build_command_center_snapshot(factory_id: int | None):
    if not factory_id:
        return {
            "live_items": [],
            "manual_items": [],
            "my_items": [],
            "attention_items": [],
            "open_count": 0,
            "live_count": 0,
            "manual_count": 0,
            "my_count": 0,
            "urgent_count": 0,
            "overdue_count": 0,
            "unassigned_count": 0,
            "backlog_rows": [],
            "done_today_count": 0,
        }

    _sync_system_generated_operational_tasks(factory_id)

    system_rows_all = _get_visible_operational_tasks(factory_id, is_system_generated=True)
    manual_rows_all = _get_visible_operational_tasks(factory_id, is_system_generated=False)
    my_rows_all = _get_my_operational_tasks(factory_id)

    live_items = [_serialize_operational_task(task) for task in system_rows_all[:10]]
    manual_items = [_serialize_operational_task(task) for task in manual_rows_all[:10]]
    my_items = [_serialize_operational_task(task) for task in my_rows_all[:8]]
    combined_items = [_serialize_operational_task(task) for task in (system_rows_all + manual_rows_all)]
    combined_items = sorted(
        combined_items,
        key=lambda item: (
            _task_priority_sort_key(item.get("priority")),
            0 if item.get("status") == "in_progress" else 1,
            item.get("title", ""),
        ),
    )

    today = date.today()
    done_today_count = 0
    for task in _get_done_operational_tasks(factory_id, limit=30):
        closed_at = getattr(task, "closed_at", None)
        if closed_at and closed_at.date() == today:
            done_today_count += 1

    all_open_rows = system_rows_all + manual_rows_all

    return {
        "live_items": live_items,
        "manual_items": manual_items,
        "my_items": my_items,
        "attention_items": combined_items[:4],
        "open_count": len(all_open_rows),
        "live_count": len(system_rows_all),
        "manual_count": len(manual_rows_all),
        "my_count": len(my_rows_all),
        "urgent_count": sum(1 for item in combined_items if item.get("priority") in {"urgent", "high"}),
        "overdue_count": sum(1 for task in all_open_rows if getattr(task, "due_date", None) and task.due_date < today),
        "unassigned_count": sum(1 for task in all_open_rows if not getattr(task, "assigned_user_id", None)),
        "backlog_rows": _build_command_center_backlog_rows(factory_id),
        "done_today_count": done_today_count,
    }


def _build_role_focus(role_value):
    role = str(role_value or "").strip().lower()
    mapping = {
        "admin": {
            "title": "Owner-level control",
            "summary": "This account can usually manage people, branches, inventory flow, and reporting across the workspace.",
            "points": [
                "Review team access and keep permissions clean.",
                "Watch branch coverage so shops stay linked to the right factory.",
                "Use reports and cash tools to keep the operation aligned.",
            ],
        },
        "manager": {
            "title": "Operations manager focus",
            "summary": "This account is centered on production, transfers, stock movement, and day-to-day operational control.",
            "points": [
                "Keep production and transfer records current.",
                "Watch low stock and move goods before shops slow down.",
                "Use reports to catch issues before they become delays.",
            ],
        },
        "accountant": {
            "title": "Finance and control focus",
            "summary": "This account is best used for cash visibility, reporting, and financial oversight around business activity.",
            "points": [
                "Keep cash records accurate and timely.",
                "Review reporting pages for operational anomalies.",
                "Coordinate with operations when stock movement affects financial reporting.",
            ],
        },
        "viewer": {
            "title": "Read-only visibility",
            "summary": "This account is mainly for oversight and monitoring, with minimal risk of operational changes.",
            "points": [
                "Use it for status checks and visibility.",
                "Confirm workspace details stay current for reporting clarity.",
                "Escalate data corrections to managers or admins when needed.",
            ],
        },
        "shop": {
            "title": "Frontline branch account",
            "summary": "This account is designed for store-side work such as sales flow, branch stock, and order follow-through.",
            "points": [
                "Keep shop stock and sales data updated.",
                "Follow through on branch orders and transfers.",
                "Raise low-stock needs early so the factory can react quickly.",
            ],
        },
    }

    return mapping.get(
        role,
        {
            "title": "Workspace account",
            "summary": "This account belongs to the business workspace and should keep its identity and assignment information clear.",
            "points": [
                "Keep account details easy for the team to recognize.",
                "Confirm workspace links are still correct.",
                "Use the account only within its intended business scope.",
            ],
        },
    )


def _build_profile_next_steps(user, active_telegram_link):
    steps = []
    full_name = display_value(getattr(user, "full_name", None), fallback="")
    workspace = getattr(user, "factory", None)
    shop = getattr(user, "shop", None)

    if not full_name:
        steps.append("Add a full name so activity and approvals are easier to recognize.")

    if not active_telegram_link:
        steps.append("Connect Telegram if you want faster alerts and follow-up on business activity.")

    if not workspace and not shop:
        steps.append("Link this account to a workspace or shop so ownership and permissions stay clear.")

    if not steps:
        steps.append("Your account setup looks healthy. Use Account settings only when you need to change credentials or notification links.")

    return steps[:3]


def _resolve_current_workspace():
    workspace = getattr(current_user, "factory", None)
    if workspace:
        return workspace

    linked_shop = getattr(current_user, "shop", None)
    if linked_shop and getattr(linked_shop, "factory", None):
        return linked_shop.factory

    workspace_id = session.get("factory_id")
    if workspace_id:
        try:
            return Factory.query.get(int(workspace_id))
        except (TypeError, ValueError):
            return None

    return None


def _can_edit_workspace_settings(workspace) -> bool:
    if not workspace or not current_user.is_authenticated:
        return False

    if getattr(current_user, "is_superadmin", False):
        return True

    if bool(getattr(current_user, "is_admin", False) and getattr(current_user, "factory_id", None) == workspace.id):
        return True

    return _is_workspace_owner(workspace)


def _can_manage_workspace_team(workspace) -> bool:
    return _can_edit_workspace_settings(workspace)


def _get_workspace_shops(workspace_id: int | None):
    if not workspace_id:
        return []

    return (
        Shop.query
        .join(ShopFactoryLink, ShopFactoryLink.shop_id == Shop.id)
        .filter(ShopFactoryLink.factory_id == workspace_id)
        .order_by(Shop.name.asc())
        .distinct()
        .all()
    )


def _workspace_role_sort_key(role_value: str) -> int:
    order = {
        "admin": 0,
        "manager": 1,
        "accountant": 2,
        "viewer": 3,
        "shop": 4,
    }
    return order.get(str(role_value or "").strip().lower(), 99)


def _get_workspace_team_users(workspace_id: int | None):
    if not workspace_id:
        return []

    rows = User.query.filter(User.factory_id == workspace_id).all()
    owner_user = _get_workspace_owner_user(workspace_id)
    owner_user_id = getattr(owner_user, "id", None)
    return sorted(
        rows,
        key=lambda user: (
            0 if getattr(user, "id", None) == owner_user_id else 1,
            _workspace_role_sort_key(getattr(user, "role", None)),
            (display_value(getattr(user, "full_name", None), fallback="") or "").lower(),
            (display_value(getattr(user, "username", None), fallback="") or "").lower(),
            getattr(user, "id", 0),
        ),
    )


def _get_workspace_owner_user(workspace_or_id):
    workspace = workspace_or_id

    if isinstance(workspace_or_id, int) or workspace_or_id is None:
        workspace = Factory.query.get(workspace_or_id) if workspace_or_id else None

    if not workspace:
        return None

    workspace_id = getattr(workspace, "id", None)
    owner_user_id = getattr(workspace, "owner_user_id", None)

    if owner_user_id:
        owner_user = (
            User.query
            .filter(
                User.id == owner_user_id,
                User.factory_id == workspace_id,
            )
            .first()
        )
        if owner_user:
            return owner_user

    admin_owner = (
        User.query
        .filter(
            User.factory_id == workspace_id,
            User.role == "admin",
        )
        .order_by(User.id.asc())
        .first()
    )
    if admin_owner:
        return admin_owner

    return (
        User.query
        .filter(User.factory_id == workspace_id)
        .order_by(User.id.asc())
        .first()
    )


def _is_workspace_owner(workspace, user=None) -> bool:
    if not workspace:
        return False

    owner_user = _get_workspace_owner_user(getattr(workspace, "id", None))
    target_user = user or current_user

    return bool(
        owner_user
        and target_user
        and getattr(target_user, "is_authenticated", False)
        and getattr(owner_user, "id", None) == getattr(target_user, "id", None)
    )


def _can_assign_workspace_admin_role(workspace) -> bool:
    if not workspace or not current_user.is_authenticated:
        return False

    return bool(getattr(current_user, "is_superadmin", False) or _is_workspace_owner(workspace))


def _can_transfer_workspace_ownership(workspace) -> bool:
    if not workspace or not current_user.is_authenticated:
        return False

    return bool(getattr(current_user, "is_superadmin", False) or _is_workspace_owner(workspace))


def _can_reset_workspace_member_password(workspace, member=None) -> bool:
    if not workspace or not current_user.is_authenticated:
        return False

    if getattr(current_user, "is_superadmin", False):
        return True

    if not _is_workspace_owner(workspace):
        return False

    if member and getattr(member, "id", None) == getattr(current_user, "id", None):
        return False

    return True


def _get_transferable_workspace_members(workspace_id: int | None):
    owner_user = _get_workspace_owner_user(workspace_id)
    owner_user_id = getattr(owner_user, "id", None)
    return [
        user
        for user in _get_workspace_team_users(workspace_id)
        if getattr(user, "id", None) != owner_user_id
    ]


def _get_workspace_role_choices(workspace):
    role_choices = [
        ("manager", "Manager"),
        ("viewer", "Viewer"),
        ("shop", "Shop"),
        ("accountant", "Accountant"),
    ]

    if _can_assign_workspace_admin_role(workspace):
        role_choices.insert(0, ("admin", "Admin"))

    return role_choices


def _username_exists(username: str, *, exclude_user_id: int | None = None) -> bool:
    if not username:
        return False

    clauses = [func.lower(User.username) == func.lower(username)]
    normalized_phone = normalize_phone(username)
    if normalized_phone:
        clauses.append(User.phone == normalized_phone)

    q = User.query.filter(or_(*clauses))
    if exclude_user_id is not None:
        q = q.filter(User.id != exclude_user_id)
    return q.first() is not None


def _phone_exists(phone: str | None, *, exclude_user_id: int | None = None) -> bool:
    normalized_phone = normalize_phone(phone)
    if not normalized_phone:
        return False

    q = User.query.filter(
        or_(
            User.phone == normalized_phone,
            func.lower(User.username) == func.lower(normalized_phone),
        )
    )
    if exclude_user_id is not None:
        q = q.filter(User.id != exclude_user_id)
    return q.first() is not None


def _log_workspace_activity(action: str, entity: str, entity_id: int | None, *, comment: str, before=None, after=None):
    try:
        from ..services.activity_log_service import activity_log

        activity_log.log(
            action=action,
            entity=entity,
            entity_id=entity_id,
            before=before,
            after=after,
            comment=comment,
        )
    except Exception:
        pass


def _log_profile_event(action: str, comment: str, before=None, after=None):
    try:
        from ..services.activity_log_service import activity_log

        activity_log.log(
            action=action,
            entity="user",
            entity_id=current_user.id,
            before=before,
            after=after,
            comment=comment,
        )
    except Exception:
        pass


def _prepare_profile_center_state(success_endpoint: str):
    profile_form = ProfileUpdateForm(prefix="profile")
    password_form = ChangePasswordForm(prefix="password")
    telegram_form = TelegramLinkCodeForm(prefix="telegram")

    if request.method == "GET":
        profile_form.username.data = current_user.username
        profile_form.full_name.data = getattr(current_user, "full_name", None)
        profile_form.phone.data = getattr(current_user, "phone", None)

    if profile_form.submit_profile.data:
        if profile_form.validate_on_submit():
            new_full_name = (profile_form.full_name.data or "").strip() or None
            new_phone = normalize_phone(profile_form.phone.data)
            new_username = build_login_username(profile_form.username.data, new_phone)

            if not new_username:
                flash("Username or phone is required.", "danger")
                return redirect(url_for(success_endpoint))

            if _username_exists(new_username, exclude_user_id=current_user.id):
                flash("This username is already taken.", "danger")
            elif _phone_exists(new_phone, exclude_user_id=current_user.id):
                flash("This phone number is already taken.", "danger")
            else:
                old_username = current_user.username
                old_full_name = getattr(current_user, "full_name", None)
                old_phone = getattr(current_user, "phone", None)
                current_user.username = new_username
                current_user.full_name = new_full_name
                current_user.phone = new_phone
                db.session.commit()
                _log_profile_event(
                    action="profile_updated",
                    comment="Username updated from account center.",
                    before={
                        "username": old_username,
                        "full_name": old_full_name,
                        "phone": old_phone,
                    },
                    after={
                        "username": current_user.username,
                        "full_name": current_user.full_name,
                        "phone": current_user.phone,
                    },
                )
                flash("Profile updated successfully.", "success")
                return redirect(url_for(success_endpoint))
        elif request.method == "POST":
            flash("Please review the account settings fields and try again.", "danger")

    if password_form.submit_password.data:
        if password_form.validate_on_submit():
            current_password = password_form.current_password.data or ""
            new_password = password_form.new_password.data or ""

            if not current_user.check_password(current_password):
                flash("Current password is incorrect.", "danger")
            else:
                current_user.set_password(new_password)
                current_user.must_change_password = False
                current_user.clear_login_lock()
                db.session.commit()
                _log_profile_event(
                    action="password_updated",
                    comment="Password updated from account center.",
                    after={"password_changed": True},
                )
                flash("Password updated successfully.", "success")
                return redirect(url_for(success_endpoint))
        elif request.method == "POST":
            flash("Please review the password form and try again.", "danger")

    if telegram_form.submit_telegram_code.data:
        if telegram_form.validate_on_submit():
            factory_id = _resolve_telegram_link_factory_id(current_user)

            if not factory_id:
                flash(translate("profile_telegram_code_missing_factory"), "danger")
            else:
                (
                    TelegramLinkCode.query
                    .filter(
                        TelegramLinkCode.user_id == current_user.id,
                        TelegramLinkCode.used_at.is_(None),
                    )
                    .delete(synchronize_session=False)
                )

                link_code = TelegramLinkCode.generate(
                    user_id=current_user.id,
                    factory_id=factory_id,
                    minutes=10,
                )
                db.session.add(link_code)
                db.session.commit()
                _log_profile_event(
                    action="telegram_code_generated",
                    comment="Telegram link code generated from account center.",
                    after={"factory_id": factory_id},
                )
                flash(translate("profile_telegram_code_generated"), "success")
                return redirect(url_for(success_endpoint))

    active_telegram_link = None
    if getattr(current_user, "telegram_links", None):
        active_telegram_link = sorted(
            current_user.telegram_links,
            key=lambda x: getattr(x, "created_at", None) or datetime.min,
            reverse=True,
        )[0]

    active_telegram_code = (
        TelegramLinkCode.query
        .filter(
            TelegramLinkCode.user_id == current_user.id,
            TelegramLinkCode.used_at.is_(None),
            TelegramLinkCode.expires_at > datetime.utcnow(),
        )
        .order_by(TelegramLinkCode.created_at.desc())
        .first()
    )

    return {
        "profile_form": profile_form,
        "password_form": password_form,
        "telegram_form": telegram_form,
        "active_telegram_link": active_telegram_link,
        "active_telegram_code": active_telegram_code,
    }


def _format_activity_label(value, fallback: str) -> str:
    text = str(value or "").strip().replace("_", " ")
    return text.title() if text else fallback


def _get_recent_user_activity(user_id: int, limit: int = 6):
    try:
        from ..activity_log import ActivityLog

        rows = (
            ActivityLog.query
            .filter(ActivityLog.user_id == user_id)
            .order_by(ActivityLog.timestamp.desc(), ActivityLog.id.desc())
            .limit(limit)
            .all()
        )
    except Exception:
        db.session.rollback()
        return []

    items = []
    for row in rows:
        items.append(
            {
                "title": _format_activity_label(getattr(row, "action", None), "Activity"),
                "entity": _format_activity_label(getattr(row, "entity", None), "Record"),
                "comment": display_value(getattr(row, "comment", None), fallback="No extra detail recorded."),
                "timestamp": (
                    getattr(row, "timestamp", None).strftime("%Y-%m-%d %H:%M")
                    if getattr(row, "timestamp", None)
                    else "-"
                ),
            }
        )

    return items


WORKSPACE_ACTIVITY_GROUPS = (
    ("all", "All activity"),
    ("operations", "Operations"),
    ("workspace", "Workspace"),
    ("team", "Team"),
    ("security", "Security"),
    ("profile", "Profile"),
)


def _activity_group_for_action(action_value: str | None) -> str:
    action = str(action_value or "").strip().lower()

    if action.startswith("operational_task_"):
        return "operations"
    if action in {"workspace_updated", "workspace_owner_transferred"}:
        return "workspace"
    if action in {"workspace_user_created", "workspace_user_updated", "workspace_user_deleted"}:
        return "team"
    if action in {"workspace_user_password_reset", "password_updated", "telegram_code_generated"}:
        return "security"
    if action in {"profile_updated"}:
        return "profile"

    return "workspace"


def _activity_group_label(group_key: str | None) -> str:
    mapping = {key: label for key, label in WORKSPACE_ACTIVITY_GROUPS}
    return mapping.get(str(group_key or "").strip().lower(), "Activity")


def _get_workspace_activity_bundle(workspace_id: int | None, *, group: str = "all", search: str = "", limit: int = 24):
    if not workspace_id:
        return {
            "items": [],
            "counts": {key: 0 for key, _label in WORKSPACE_ACTIVITY_GROUPS},
        }

    try:
        from ..activity_log import ActivityLog

        selected_group = str(group or "all").strip().lower()
        allowed_groups = {key for key, _label in WORKSPACE_ACTIVITY_GROUPS}
        if selected_group not in allowed_groups:
            selected_group = "all"

        query = (
            db.session.query(ActivityLog, User)
            .outerjoin(User, User.id == ActivityLog.user_id)
            .filter(
                or_(
                    User.factory_id == workspace_id,
                    and_(
                        ActivityLog.entity == "factory",
                        ActivityLog.entity_id == workspace_id,
                    ),
                )
            )
        )

        if search:
            like = f"%{search}%"
            query = query.filter(
                or_(
                    ActivityLog.action.ilike(like),
                    ActivityLog.entity.ilike(like),
                    ActivityLog.comment.ilike(like),
                    User.username.ilike(like),
                    User.full_name.ilike(like),
                )
            )

        raw_rows = (
            query
            .order_by(ActivityLog.timestamp.desc(), ActivityLog.id.desc())
            .limit(max(limit * 5, 120))
            .all()
        )
    except Exception:
        db.session.rollback()
        return {
            "items": [],
            "counts": {key: 0 for key, _label in WORKSPACE_ACTIVITY_GROUPS},
        }

    counts = {key: 0 for key, _label in WORKSPACE_ACTIVITY_GROUPS}
    counts["all"] = len(raw_rows)
    normalized_rows = []

    for log_row, actor in raw_rows:
        group_key = _activity_group_for_action(getattr(log_row, "action", None))
        counts[group_key] = counts.get(group_key, 0) + 1
        normalized_rows.append((log_row, actor, group_key))

    if selected_group != "all":
        normalized_rows = [row for row in normalized_rows if row[2] == selected_group]

    items = []
    for log_row, actor, group_key in normalized_rows[:limit]:
        actor_name = get_user_display_name(actor) if actor else "System"
        actor_initials = get_user_initials(actor) if actor else "SY"
        items.append(
            {
                "id": getattr(log_row, "id", None),
                "title": _format_activity_label(getattr(log_row, "action", None), "Activity"),
                "entity": _format_activity_label(getattr(log_row, "entity", None), "Record"),
                "comment": display_value(getattr(log_row, "comment", None), fallback="No extra detail recorded."),
                "timestamp": (
                    getattr(log_row, "timestamp", None).strftime("%Y-%m-%d %H:%M")
                    if getattr(log_row, "timestamp", None)
                    else "-"
                ),
                "group_key": group_key,
                "group_label": _activity_group_label(group_key),
                "actor_name": actor_name,
                "actor_initials": actor_initials,
                "actor_login": display_value(getattr(actor, "username", None)),
            }
        )

    return {
        "items": items,
        "counts": counts,
    }


@main_bp.route("/dashboard")
@login_required
def dashboard():
    role = getattr(current_user, "role", "manager")

    if role == "shop" or getattr(current_user, "is_shop", False):
        return redirect(url_for("shop.dashboard_shop"))

    factory_id = getattr(current_user, "factory_id", None)
    current_date = _get_current_date_for_lang()

    if not factory_id:
        return render_template("dashboard.html", current_date=current_date)

    data = _build_manager_dashboard(factory_id=factory_id)
    data["current_date"] = current_date

    manager_like_roles = {"manager", "admin", "accountant", "viewer"}

    if role in manager_like_roles or getattr(current_user, "is_manager", False):
        return render_template("dashboard_manager.html", **data)

    # Keep one consistent dashboard data contract for non-shop roles with a workspace.
    return render_template("dashboard_manager.html", **data)


@main_bp.route("/command-center", methods=["GET", "POST"])
@main_bp.route("/inbox", methods=["GET", "POST"])
@login_required
def command_center():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    workspace_id = getattr(workspace, "id", None)
    task_form = OperationalTaskForm(prefix="task")
    workspace_team_users = _get_workspace_team_users(workspace_id)
    task_form.assigned_user_id.choices = [(0, "No specific assignee")] + [
        (user.id, get_user_display_name(user))
        for user in workspace_team_users
    ]

    if request.method == "GET":
        task_form.priority.data = "medium"
        current_role = str(getattr(current_user, "role", None) or "").strip().lower()
        task_form.target_role.data = current_role if current_role in {"admin", "manager", "accountant", "viewer", "shop"} else ""

    if request.method == "POST" and request.form.get("command_action") == "update_task_status":
        task_id = request.form.get("task_id", type=int)
        next_status = (request.form.get("next_status") or "").strip().lower()
        task = (
            OperationalTask.query
            .filter(
                OperationalTask.id == task_id,
                OperationalTask.factory_id == workspace_id,
            )
            .first()
        )

        if not task:
            flash("Task not found in this workspace.", "danger")
            return redirect(url_for("main.command_center"))

        if not _can_update_operational_task(task, workspace):
            flash("You do not have permission to update this task.", "danger")
            return redirect(url_for("main.command_center"))

        allowed_statuses = {"open", "in_progress", "done", "dismissed"}
        if next_status not in allowed_statuses:
            flash("Invalid task status requested.", "danger")
            return redirect(url_for("main.command_center"))

        before = {
            "status": task.status,
            "assigned_user_id": task.assigned_user_id,
        }

        task.status = next_status
        task.updated_at = datetime.utcnow()
        if next_status in {"done", "dismissed"}:
            task.closed_at = datetime.utcnow()
            task.closed_by_id = current_user.id
        else:
            task.closed_at = None
            task.closed_by_id = None

        db.session.commit()
        _log_workspace_activity(
            action=f"operational_task_{next_status}",
            entity="operational_task",
            entity_id=task.id,
            before=before,
            after={
                "status": task.status,
                "assigned_user_id": task.assigned_user_id,
            },
            comment=f"Task '{task.title}' moved to {task.status} from the command center.",
        )
        flash("Task updated successfully.", "success")
        return redirect(url_for("main.command_center"))

    if task_form.submit_task.data:
        if task_form.validate_on_submit():
            if not _can_manage_command_center(workspace):
                flash("Only workspace admins or the workspace owner can create manual tasks.", "danger")
                return redirect(url_for("main.command_center"))

            title = (task_form.title.data or "").strip()
            description = (task_form.description.data or "").strip() or None
            priority = (task_form.priority.data or "medium").strip().lower()
            assigned_user_id = task_form.assigned_user_id.data or 0
            target_role = (task_form.target_role.data or "").strip().lower() or None
            due_date = task_form.due_date.data
            action_url = (task_form.action_url.data or "").strip() or None

            allowed_user_ids = {user.id for user in workspace_team_users}
            if assigned_user_id and assigned_user_id not in allowed_user_ids:
                flash("Selected assignee is not part of this workspace.", "danger")
                return redirect(url_for("main.command_center"))

            if action_url and not action_url.startswith("/"):
                flash("Action link must start with / so it stays inside the app.", "danger")
                return redirect(url_for("main.command_center"))

            task = OperationalTask(
                factory_id=workspace_id,
                assigned_user_id=assigned_user_id or None,
                created_by_id=current_user.id,
                task_type="manual",
                title=title,
                description=description,
                action_url=action_url,
                target_role=None if assigned_user_id else target_role,
                priority=priority,
                status="open",
                due_date=due_date,
                is_system_generated=False,
            )
            db.session.add(task)
            db.session.commit()

            _log_workspace_activity(
                action="operational_task_created",
                entity="operational_task",
                entity_id=task.id,
                after={
                    "title": task.title,
                    "priority": task.priority,
                    "assigned_user_id": task.assigned_user_id,
                    "target_role": task.target_role,
                },
                comment=f"Manual task '{task.title}' created from the command center.",
            )
            flash("Manual task created successfully.", "success")
            return redirect(url_for("main.command_center"))
        elif request.method == "POST":
            flash("Please review the task fields and try again.", "danger")

    command_center = _build_command_center_snapshot(workspace_id)
    setup_items = _build_dashboard_setup_items(workspace_id)
    setup_remaining = sum(1 for item in setup_items if not item["done"])
    done_items = [_serialize_operational_task(task) for task in _get_done_operational_tasks(workspace_id, limit=6)]

    return render_template(
        "dashboard/command_center.html",
        workspace_page_name=display_value(getattr(workspace, "name", None), fallback="Workspace"),
        workspace_role_label=_format_role_label(getattr(current_user, "role", None)),
        command_center_attention_items=command_center["attention_items"],
        command_center_live_items=command_center["live_items"],
        command_center_manual_items=command_center["manual_items"],
        command_center_my_items=command_center["my_items"],
        command_center_open_count=command_center["open_count"],
        command_center_live_count=command_center["live_count"],
        command_center_manual_count=command_center["manual_count"],
        command_center_my_count=command_center["my_count"],
        command_center_urgent_count=command_center["urgent_count"],
        command_center_overdue_count=command_center["overdue_count"],
        command_center_unassigned_count=command_center["unassigned_count"],
        command_center_backlog_rows=command_center["backlog_rows"],
        command_center_done_today_count=command_center["done_today_count"],
        command_center_done_items=done_items,
        command_center_can_manage=_can_manage_command_center(workspace),
        setup_items=setup_items,
        setup_remaining=setup_remaining,
        task_form=task_form,
    )


@main_bp.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    state = _prepare_profile_center_state("main.profile")
    if not isinstance(state, dict):
        return state

    return render_template("profile/index.html", **state)


@main_bp.route("/workspace", methods=["GET", "POST"])
@login_required
def workspace_details():
    workspace = _resolve_current_workspace()
    workspace_form = WorkspaceProfileForm(prefix="workspace")
    member_form = WorkspaceTeamMemberForm(prefix="member")
    transfer_form = WorkspaceOwnershipTransferForm(prefix="transfer")
    workspace_can_edit = _can_edit_workspace_settings(workspace)
    workspace_can_manage_team = _can_manage_workspace_team(workspace)
    workspace_can_assign_admin = _can_assign_workspace_admin_role(workspace)
    workspace_can_transfer_owner = _can_transfer_workspace_ownership(workspace)
    workspace_can_reset_passwords = _can_reset_workspace_member_password(workspace)
    workspace_id = getattr(workspace, "id", None)
    requested_panel = (request.args.get("panel") or "").strip().lower()
    workspace_shops = _get_workspace_shops(workspace_id)
    workspace_owner_user = _get_workspace_owner_user(workspace)
    transfer_candidates = _get_transferable_workspace_members(workspace_id)
    member_form.role.choices = _get_workspace_role_choices(workspace)
    member_form.shop_id.choices = [(0, "No shop")] + [(shop.id, shop.name) for shop in workspace_shops]
    transfer_form.new_owner_id.choices = [(0, "Select a team member")] + [
        (user.id, f"{get_user_display_name(user)} ({_format_role_label(getattr(user, 'role', None))})")
        for user in transfer_candidates
    ]

    if request.method == "GET" and workspace:
        workspace_form.name.data = getattr(workspace, "name", None)
        workspace_form.owner_name.data = getattr(workspace, "owner_name", None)
        workspace_form.location.data = getattr(workspace, "location", None)
        workspace_form.phone.data = getattr(workspace, "phone", None)
        workspace_form.note.data = getattr(workspace, "note", None)
        member_form.role.data = "manager"
        if transfer_candidates:
            transfer_form.new_owner_id.data = transfer_candidates[0].id

    if request.method == "POST" and request.form.get("workspace_action") == "delete_member":
        member_id = request.form.get("member_id", type=int)

        if not workspace or not workspace_can_manage_team:
            flash("You do not have permission to manage workspace users.", "danger")
            return redirect(url_for("main.workspace_details"))

        member = (
            User.query
            .filter(User.id == member_id, User.factory_id == workspace_id)
            .first()
        )

        if not member:
            flash("User not found in this workspace.", "danger")
        elif member.id == current_user.id:
            flash("You cannot delete your own account from this workspace.", "danger")
        elif workspace_owner_user and member.id == workspace_owner_user.id and not current_user.is_superadmin:
            flash("The workspace owner account cannot be deleted here.", "danger")
        elif member.role == "admin" and not workspace_can_assign_admin:
            flash("Only the workspace owner can remove another admin account.", "danger")
        else:
            deleted_label = get_user_display_name(member)
            db.session.delete(member)
            db.session.commit()
            _log_workspace_activity(
                action="workspace_user_deleted",
                entity="user",
                entity_id=member_id,
                after={"workspace_id": workspace_id},
                comment=f"User {deleted_label} deleted from workspace page.",
            )
            flash("User deleted successfully.", "success")

        return redirect(url_for("main.workspace_details"))

    if request.method == "POST" and request.form.get("workspace_action") == "reset_member_password":
        member_id = request.form.get("member_id", type=int)
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not workspace or not workspace_can_manage_team:
            flash("You do not have permission to manage workspace users.", "danger")
            return redirect(url_for("main.workspace_details"))

        member = (
            User.query
            .filter(User.id == member_id, User.factory_id == workspace_id)
            .first()
        )

        if not member:
            flash("User not found in this workspace.", "danger")
        elif not _can_reset_workspace_member_password(workspace, member):
            flash("Only the workspace owner can reset teammate passwords from this page.", "danger")
        elif not new_password:
            flash("Enter a temporary password for this member.", "danger")
        elif len(new_password) < 6:
            flash("Temporary password must be at least 6 characters.", "danger")
        elif new_password != confirm_password:
            flash("Password confirmation does not match.", "danger")
        else:
            member.set_password(new_password)
            member.must_change_password = True
            member.clear_login_lock()
            db.session.commit()
            _log_workspace_activity(
                action="workspace_user_password_reset",
                entity="user",
                entity_id=member.id,
                after={"workspace_id": workspace_id},
                comment=f"Temporary password reset for {get_user_display_name(member)} from workspace page.",
            )
            flash("Temporary password updated. The user will be forced to create a private password on next login.", "success")

        return redirect(url_for("main.workspace_details"))

    if request.method == "POST" and request.form.get("workspace_action") == "update_member":
        member_id = request.form.get("member_id", type=int)

        if not workspace or not workspace_can_manage_team:
            flash("You do not have permission to manage workspace users.", "danger")
            return redirect(url_for("main.workspace_details"))

        member = (
            User.query
            .filter(User.id == member_id, User.factory_id == workspace_id)
            .first()
        )
        selected_full_name = (request.form.get("full_name") or "").strip() or None
        selected_phone = normalize_phone(request.form.get("phone"))
        selected_username_input = (request.form.get("username") or "").strip()
        selected_username = build_login_username(selected_username_input, selected_phone)

        if not member:
            flash("User not found in this workspace.", "danger")
            return redirect(url_for("main.workspace_details"))

        selected_role = (request.form.get("role") or "").strip()
        allowed_roles = {value for value, _label in _get_workspace_role_choices(workspace)}
        selected_shop_id = request.form.get("shop_id", type=int) or 0
        selected_shop = selected_shop_id or None
        allowed_shop_ids = {shop.id for shop in workspace_shops}

        if workspace_owner_user and member.id == workspace_owner_user.id and not current_user.is_superadmin:
            flash("The workspace owner account keeps owner access and cannot be reassigned here.", "danger")
            return redirect(url_for("main.workspace_details"))

        if member.role == "admin" and not workspace_can_assign_admin:
            flash("Only the workspace owner can change another admin account.", "danger")
            return redirect(url_for("main.workspace_details"))

        if not selected_username:
            flash("Username or phone is required for this member.", "danger")
            return redirect(url_for("main.workspace_details"))

        if _username_exists(selected_username, exclude_user_id=member.id):
            flash("That username already exists.", "danger")
            return redirect(url_for("main.workspace_details"))

        if _phone_exists(selected_phone, exclude_user_id=member.id):
            flash("That phone number already exists.", "danger")
            return redirect(url_for("main.workspace_details"))

        if selected_role not in allowed_roles:
            flash("Invalid role selected.", "danger")
            return redirect(url_for("main.workspace_details"))

        if selected_role == "shop":
            if not selected_shop:
                flash("Shop users must be linked to a shop.", "danger")
                return redirect(url_for("main.workspace_details"))
            if selected_shop not in allowed_shop_ids:
                flash("Selected shop is not linked to this workspace.", "danger")
                return redirect(url_for("main.workspace_details"))
        else:
            selected_shop = None

        before = {
            "username": member.username,
            "full_name": member.full_name,
            "phone": member.phone,
            "role": member.role,
            "shop_id": member.shop_id,
        }

        member.username = selected_username
        member.full_name = selected_full_name
        member.phone = selected_phone
        member.role = selected_role
        member.shop_id = selected_shop
        member.factory_id = workspace_id

        db.session.commit()

        _log_workspace_activity(
            action="workspace_user_updated",
            entity="user",
            entity_id=member.id,
            before=before,
            after={
                "username": member.username,
                "full_name": member.full_name,
                "phone": member.phone,
                "role": member.role,
                "shop_id": member.shop_id,
            },
            comment=f"Identity and access updated for {get_user_display_name(member)} from workspace page.",
        )

        flash("Team member details updated successfully.", "success")
        return redirect(url_for("main.workspace_details"))

    if workspace_form.submit_workspace.data:
        if workspace_form.validate_on_submit():
            if not workspace or not workspace_can_edit:
                flash("You do not have permission to update this workspace.", "danger")
                return redirect(url_for("main.workspace_details"))

            before = {
                "name": workspace.name,
                "owner_name": workspace.owner_name,
                "location": workspace.location,
                "phone": workspace.phone,
                "note": workspace.note,
            }

            workspace.name = (workspace_form.name.data or "").strip()
            workspace.owner_name = (workspace_form.owner_name.data or "").strip() or None
            workspace.location = (workspace_form.location.data or "").strip() or None
            workspace.phone = (workspace_form.phone.data or "").strip() or None
            workspace.note = (workspace_form.note.data or "").strip() or None
            db.session.commit()

            _log_workspace_activity(
                action="workspace_updated",
                entity="factory",
                entity_id=workspace.id,
                before=before,
                after={
                    "name": workspace.name,
                    "owner_name": workspace.owner_name,
                    "location": workspace.location,
                    "phone": workspace.phone,
                    "note": workspace.note,
                },
                comment="Workspace settings updated from dashboard workspace page.",
            )

            flash("Workspace details updated successfully.", "success")
            return redirect(url_for("main.workspace_details"))
        elif request.method == "POST":
            flash("Please review the workspace form and try again.", "danger")

    if member_form.submit_member.data:
        if member_form.validate_on_submit():
            if not workspace or not workspace_can_manage_team:
                flash("You do not have permission to manage workspace users.", "danger")
                return redirect(url_for("main.workspace_details"))

            full_name = (member_form.full_name.data or "").strip()
            phone = normalize_phone(member_form.phone.data)
            username = build_login_username(member_form.username.data, phone)
            password = member_form.password.data or ""
            role = (member_form.role.data or "").strip()
            selected_shop_id = member_form.shop_id.data or 0
            shop_id = selected_shop_id or None

            if not username:
                flash("Enter a username or a phone number for the new user.", "danger")
                return redirect(url_for("main.workspace_details"))

            if _username_exists(username):
                flash("That username already exists.", "danger")
                return redirect(url_for("main.workspace_details"))

            if _phone_exists(phone):
                flash("That phone number already exists.", "danger")
                return redirect(url_for("main.workspace_details"))

            allowed_roles = {value for value, _label in member_form.role.choices}
            if role not in allowed_roles:
                flash("Invalid role selected.", "danger")
                return redirect(url_for("main.workspace_details"))

            allowed_shop_ids = {shop.id for shop in workspace_shops}
            if role == "shop":
                if not shop_id:
                    flash("Shop users must be linked to a shop.", "danger")
                    return redirect(url_for("main.workspace_details"))
                if shop_id not in allowed_shop_ids:
                    flash("Selected shop is not linked to this workspace.", "danger")
                    return redirect(url_for("main.workspace_details"))
            else:
                shop_id = None

            new_user = User(
                username=username,
                full_name=full_name,
                phone=phone,
                role=role,
                factory_id=workspace_id,
                shop_id=shop_id,
                must_change_password=True,
            )
            new_user.set_password(password)
            new_user.clear_login_lock()
            db.session.add(new_user)
            db.session.commit()

            if workspace and not getattr(workspace, "owner_user_id", None):
                resolved_owner_id = getattr(workspace_owner_user, "id", None)
                if not resolved_owner_id and getattr(current_user, "factory_id", None) == workspace_id:
                    resolved_owner_id = current_user.id
                if resolved_owner_id:
                    workspace.owner_user_id = resolved_owner_id
                    db.session.commit()

            _log_workspace_activity(
                action="workspace_user_created",
                entity="user",
                entity_id=new_user.id,
                after={
                    "workspace_id": workspace_id,
                    "username": new_user.username,
                    "role": new_user.role,
                },
                comment=f"User {new_user.username} created from workspace page.",
            )

            flash(f"Team member created successfully. Login: {new_user.username}. First login will require a private password update.", "success")
            return redirect(url_for("main.workspace_details"))
        elif request.method == "POST":
            flash("Please review the team member form and try again.", "danger")

    if transfer_form.submit_transfer.data:
        if transfer_form.validate_on_submit():
            if not workspace or not workspace_can_transfer_owner:
                flash("Only the workspace owner can transfer ownership from this page.", "danger")
                return redirect(url_for("main.workspace_details"))

            next_owner = (
                User.query
                .filter(
                    User.id == transfer_form.new_owner_id.data,
                    User.factory_id == workspace_id,
                )
                .first()
            )
            current_owner_id = getattr(workspace_owner_user, "id", None)

            if not next_owner or next_owner.id == current_owner_id:
                flash("Select a different teammate to become the new owner.", "danger")
                return redirect(url_for("main.workspace_details"))

            if not current_user.is_superadmin and not current_user.check_password(transfer_form.current_password.data or ""):
                flash("Current password is incorrect.", "danger")
                return redirect(url_for("main.workspace_details"))

            before = {
                "owner_user_id": current_owner_id,
            }

            next_owner.role = "admin"
            next_owner.shop_id = None
            next_owner.factory_id = workspace_id
            workspace.owner_user_id = next_owner.id
            db.session.commit()

            _log_workspace_activity(
                action="workspace_owner_transferred",
                entity="factory",
                entity_id=workspace_id,
                before=before,
                after={
                    "owner_user_id": workspace.owner_user_id,
                    "owner_username": next_owner.username,
                },
                comment=f"Workspace ownership transferred to {get_user_display_name(next_owner)}.",
            )

            flash("Workspace ownership transferred successfully.", "success")
            return redirect(url_for("main.workspace_details"))
        elif request.method == "POST":
            flash("Please review the ownership transfer form and try again.", "danger")

    workspace_rows = _build_workspace_registration_rows(workspace)
    workspace_ready_rows = sum(1 for row in workspace_rows if row["is_ready"])
    workspace_total_rows = len(workspace_rows)
    workspace_completion_pct = int((workspace_ready_rows / workspace_total_rows) * 100) if workspace_total_rows else 0
    workspace_missing_fields = [row["label"] for row in workspace_rows if not row["is_ready"]]

    workspace_team_users = _get_workspace_team_users(workspace_id)
    workspace_owner_user = _get_workspace_owner_user(workspace)
    workspace_activity_bundle = _get_workspace_activity_bundle(workspace_id, limit=4)
    team_count = len(workspace_team_users)
    linked_shop_count = 0
    if workspace_id:
        linked_shop_count = (
            db.session.query(func.count(func.distinct(ShopFactoryLink.shop_id)))
            .filter(ShopFactoryLink.factory_id == workspace_id)
            .scalar()
            or 0
        )

    phone_present = bool(display_value(getattr(workspace, "phone", None), fallback=""))
    location_present = bool(display_value(getattr(workspace, "location", None), fallback=""))
    owner_present = bool(display_value(getattr(workspace, "owner_name", None), fallback=""))
    contact_status = "Ready" if phone_present and location_present and owner_present else "Needs attention"
    workspace_disclosure_state = {
        "workspace_settings": requested_panel == "workspace" or bool(workspace_form.errors),
        "add_member": requested_panel == "team" or bool(member_form.errors),
        "owner_controls": requested_panel == "owner" or bool(transfer_form.errors),
    }

    return render_template(
        "dashboard/workspace_details.html",
        workspace_page_name=display_value(getattr(workspace, "name", None), fallback="Workspace"),
        workspace_owner_name=display_value(getattr(workspace, "owner_name", None)),
        workspace_location=display_value(getattr(workspace, "location", None)),
        workspace_phone=display_value(getattr(workspace, "phone", None)),
        workspace_note=display_value(getattr(workspace, "note", None)),
        workspace_role=display_value(getattr(current_user, "role", None)),
        workspace_role_label=_format_role_label(getattr(current_user, "role", None)),
        workspace_rows=workspace_rows,
        workspace_team_count=team_count,
        workspace_shop_count=int(linked_shop_count),
        workspace_ready_rows=workspace_ready_rows,
        workspace_total_rows=workspace_total_rows,
        workspace_completion_pct=workspace_completion_pct,
        workspace_missing_fields=workspace_missing_fields,
        workspace_contact_status=contact_status,
        workspace_owner_points=_build_workspace_owner_points(
            team_count=team_count,
            linked_shop_count=int(linked_shop_count),
            missing_fields=workspace_missing_fields,
        ),
        workspace_owner_user=workspace_owner_user,
        workspace_owner_display_name=get_user_display_name(workspace_owner_user) if workspace_owner_user else "-",
        workspace_owner_login=display_value(getattr(workspace_owner_user, "username", None)),
        workspace_form=workspace_form,
        workspace_can_edit=workspace_can_edit,
        workspace_is_owner=_is_workspace_owner(workspace),
        member_form=member_form,
        workspace_can_manage_team=workspace_can_manage_team,
        workspace_can_assign_admin=workspace_can_assign_admin,
        workspace_can_reset_passwords=workspace_can_reset_passwords,
        workspace_can_transfer_owner=workspace_can_transfer_owner,
        workspace_team_users=workspace_team_users,
        workspace_owner_user_id=getattr(workspace_owner_user, "id", None),
        workspace_shop_options=workspace_shops,
        transfer_form=transfer_form,
        workspace_transfer_candidate_count=len(transfer_candidates),
        workspace_disclosure_state=workspace_disclosure_state,
        workspace_activity_preview=workspace_activity_bundle["items"],
        workspace_activity_counts=workspace_activity_bundle["counts"],
        workspace_can_view_activity=workspace_can_manage_team,
    )


@main_bp.route("/workspace/activity")
@login_required
def workspace_activity():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    if not _can_manage_workspace_team(workspace):
        flash("Only workspace admins or the workspace owner can view the activity center.", "danger")
        return redirect(url_for("main.workspace_details"))

    selected_group = (request.args.get("group") or "all").strip().lower()
    search_query = (request.args.get("q") or "").strip()
    limit = request.args.get("limit", type=int) or 24
    limit = max(12, min(limit, 60))

    activity_bundle = _get_workspace_activity_bundle(
        getattr(workspace, "id", None),
        group=selected_group,
        search=search_query,
        limit=limit,
    )

    return render_template(
        "dashboard/workspace_activity.html",
        workspace_page_name=display_value(getattr(workspace, "name", None), fallback="Workspace"),
        workspace_role_label=_format_role_label(getattr(current_user, "role", None)),
        activity_items=activity_bundle["items"],
        activity_counts=activity_bundle["counts"],
        activity_group_options=[
            {
                "key": key,
                "label": label,
                "count": activity_bundle["counts"].get(key, 0),
                "href": url_for("main.workspace_activity", group=key, q=search_query or None, limit=limit),
            }
            for key, label in WORKSPACE_ACTIVITY_GROUPS
        ],
        activity_selected_group=selected_group if selected_group in {key for key, _ in WORKSPACE_ACTIVITY_GROUPS} else "all",
        activity_search_query=search_query,
        activity_limit=limit,
    )


@main_bp.route("/profile/overview", methods=["GET", "POST"])
@login_required
def profile_overview():
    state = _prepare_profile_center_state("main.profile_overview")
    if not isinstance(state, dict):
        return state

    requested_panel = (request.values.get("panel") or "").strip().lower()
    requested_tab = (request.values.get("tab") or "").strip().lower()
    linked_shop = getattr(current_user, "shop", None)
    linked_workspace = _resolve_current_workspace()
    active_telegram_link = state["active_telegram_link"]

    role_focus = _build_role_focus(getattr(current_user, "role", None))
    assignment_summary = "Workspace + shop linked"
    if linked_workspace and not linked_shop:
        assignment_summary = "Workspace-level account"
    elif linked_shop and not linked_workspace:
        assignment_summary = "Shop-linked account"
    elif not linked_workspace and not linked_shop:
        assignment_summary = "Assignment missing"

    profile_disclosure_state = {
        "account": requested_panel == "account" or bool(state["profile_form"].errors),
        "password": requested_panel == "password" or bool(state["password_form"].errors),
        "telegram": requested_panel == "telegram" or bool(state["telegram_form"].errors) or bool(state["active_telegram_code"]),
        "activity": requested_panel == "activity",
    }
    settings_panel_requested = requested_panel in {"account", "password", "telegram", "activity"}
    profile_active_tab = "settings" if requested_tab == "settings" or settings_panel_requested else "profile"
    can_view_workspace_activity = bool(linked_workspace and _can_manage_workspace_team(linked_workspace))

    return render_template(
        "dashboard/profile_overview.html",
        user_chip_name=get_user_display_name(current_user),
        user_chip_initials=get_user_initials(current_user),
        profile_username=display_value(getattr(current_user, "username", None)),
        profile_full_name=display_value(getattr(current_user, "full_name", None)),
        profile_phone=display_value(getattr(current_user, "phone", None)),
        profile_role=display_value(getattr(current_user, "role", None)),
        profile_role_label=_format_role_label(getattr(current_user, "role", None)),
        profile_workspace_name=display_value(getattr(linked_workspace, "name", None)),
        profile_shop_name=display_value(getattr(linked_shop, "name", None)),
        profile_assignment_summary=assignment_summary,
        profile_telegram_status=translate("profile_connected") if active_telegram_link else translate("profile_not_connected"),
        profile_telegram_chat_id=display_value(getattr(active_telegram_link, "telegram_chat_id", None)),
        profile_role_focus_title=role_focus["title"],
        profile_role_focus_summary=role_focus["summary"],
        profile_role_focus_points=role_focus["points"],
        profile_next_steps=_build_profile_next_steps(current_user, active_telegram_link),
        recent_activity=_get_recent_user_activity(current_user.id),
        profile_disclosure_state=profile_disclosure_state,
        profile_active_tab=profile_active_tab,
        profile_can_view_workspace_activity=can_view_workspace_activity,
        **state,
    )

@main_bp.route("/inventory")
@login_required
def inventory():
    inventory_cards = [
        {
            "id": "cutting_batch",
            "title": translate("inventory_card_cutting_batch_title"),
            "subtitle": translate("inventory_card_cutting_batch_subtitle"),
            "href": url_for("cutting.cutting_order_list"),
            "status": "live",
        },
        {
            "id": "operations_board",
            "title": translate("inventory_card_operations_board_title"),
            "subtitle": translate("inventory_card_operations_board_subtitle"),
            "href": url_for("main.operations_board"),
            "status": "live",
        },
        {
            "id": "my_work",
            "title": translate("inventory_card_my_work_title"),
            "subtitle": translate("inventory_card_my_work_subtitle"),
            "href": url_for("main.worker_floor"),
            "status": "live",
        },
        {
            "id": "materials",
            "title": translate("inventory_card_materials_title"),
            "subtitle": translate("inventory_card_materials_subtitle"),
            "href": url_for("fabrics.list"),
            "status": "live",
        },
        {
            "id": "products",
            "title": translate("inventory_card_products_title"),
            "subtitle": translate("inventory_card_products_subtitle"),
            "href": url_for("products.list_products"),
            "status": "live",
        },
        {
            "id": "factory_stock",
            "title": translate("inventory_card_factory_stock_title"),
            "subtitle": translate("inventory_card_factory_stock_subtitle"),
            "href": url_for("products.list_products"),
            "status": "live",
        },
        {
            "id": "shop_stock",
            "title": translate("inventory_card_shop_stock_title"),
            "subtitle": translate("inventory_card_shop_stock_subtitle"),
            "href": url_for("products.shop_stock_products"),
            "status": "live",
        },
        {
            "id": "transfers",
            "title": translate("inventory_card_transfers_title"),
            "subtitle": translate("inventory_card_transfers_subtitle"),
            "href": url_for("shop.transfer_to_shop"),
            "status": "live" if current_user.role in ("admin", "manager") else "locked",
        },
        {
            "id": "low_stock",
            "title": translate("inventory_card_low_stock_title"),
            "subtitle": translate("inventory_card_low_stock_subtitle"),
            "href": url_for("main.low_stock"),
            "status": "live",
        },
        {
            "id": "production_planning",
            "title": translate("inventory_card_production_planning_title"),
            "subtitle": translate("inventory_card_production_planning_subtitle"),
            "href": url_for("main.production_plan"),
            "status": "live",
        },
        {
            "id": "saved_plans",
            "title": translate("inventory_card_saved_plans_title"),
            "subtitle": translate("inventory_card_saved_plans_subtitle"),
            "href": url_for("main.production_plan_history"),
            "status": "live",
        },
        {
            "id": "composition",
            "title": translate("inventory_card_composition_title"),
            "subtitle": translate("inventory_card_composition_subtitle"),
            "href": url_for("main.composition"),
            "status": "soon",
        },
        {
            "id": "activity",
            "title": translate("inventory_card_activity_title"),
            "subtitle": translate("inventory_card_activity_subtitle"),
            "href": url_for("shop.movement_history"),
            "status": "live",
        },
    ]

    return render_template("inventory/hub.html", inventory_cards=inventory_cards)


def _latest_supplier_cost_map(factory_id: int | None):
    if not factory_id:
        return {}

    rows = (
        SupplierReceipt.query
        .filter(
            SupplierReceipt.factory_id == factory_id,
            SupplierReceipt.fabric_id.isnot(None),
            SupplierReceipt.unit_cost.isnot(None),
        )
        .order_by(
            SupplierReceipt.received_at.desc(),
            SupplierReceipt.created_at.desc(),
            SupplierReceipt.id.desc(),
        )
        .all()
    )

    result = {}
    for row in rows:
        fabric_id = getattr(row, "fabric_id", None)
        if not fabric_id or fabric_id in result:
            continue
        result[fabric_id] = {
            "unit_cost": float(row.unit_cost or 0),
            "currency": (row.currency or "UZS").upper(),
            "invoice_number": row.invoice_number or "",
            "received_at": row.received_at,
        }
    return result


def _active_ops_stage_map(factory_id: int | None, order_item_ids: list[int] | None):
    if not factory_id or not order_item_ids:
        return {}

    rows = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "ops_stage_flow",
            OperationalTask.source_id.in_(order_item_ids),
            OperationalTask.status.in_(("open", "in_progress")),
        )
        .order_by(OperationalTask.updated_at.desc(), OperationalTask.id.desc())
        .all()
    )

    result = {}
    for row in rows:
        source_id = getattr(row, "source_id", None)
        if not source_id or source_id in result:
            continue
        stage_meta = OPS_STAGE_BY_TASK_TYPE.get(getattr(row, "task_type", None) or "")
        if not stage_meta:
            continue
        result[source_id] = {
            "task_id": row.id,
            "stage_key": stage_meta["key"],
            "stage_label": stage_meta["label"],
            "task_type": stage_meta["task_type"],
            "status": (row.status or "open").strip().lower(),
        }
    return result


def _append_operational_task_note(task, note: str | None, *, prefix: str):
    clean_note = str(note or "").strip()
    if not clean_note:
        return
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
    actor = get_user_display_name(current_user) if current_user.is_authenticated else "User"
    line = f"[{timestamp}] {prefix} by {actor}: {clean_note}"
    existing = (getattr(task, "description", None) or "").strip()
    task.description = f"{existing}\n{line}".strip() if existing else line


def _start_ops_stage_for_order(factory_id: int | None, order_item_id: int):
    item = _operations_board_stage_order_item(factory_id, order_item_id)
    if not item:
        return False, "That order line is no longer available."

    existing = _active_ops_stage_map(factory_id, [order_item_id]).get(order_item_id)
    if existing:
        return False, "A workflow stage is already active for this order line."

    plan = (
        ProductionPlan.query
        .filter(
            ProductionPlan.factory_id == factory_id,
            ProductionPlan.order_item_id == order_item_id,
        )
        .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
        .first()
    )
    if not plan or not bool(plan.can_fulfill_plan):
        return False, translate("ops_error_start_stage_no_plan")

    stage = OPS_STAGE_FLOW[0]
    task = OperationalTask(
        factory_id=factory_id,
        created_by_id=current_user.id,
        task_type=stage["task_type"],
        source_type="ops_stage_flow",
        source_id=order_item_id,
        title=f"{translate(stage['label_key'])}: {item.product.name if item.product else f'Product #{item.product_id}'}",
        description=f"{translate(stage['description_key'])} Remaining qty: {int(item.qty_remaining or 0)}.",
        action_url=url_for("main.production_plan_detail", plan_id=plan.id),
        target_role=stage["target_role"],
        priority=stage["priority"],
        status="open",
        is_system_generated=False,
    )
    db.session.add(task)
    db.session.commit()
    return True, translate("ops_stage_started_message")


def _advance_ops_stage_for_task(factory_id: int | None, task, *, note: str | None = None):
    workspace = Factory.query.get(factory_id) if factory_id else None
    if not task or not workspace or not _can_update_operational_task(task, workspace):
        return False, translate("ops_error_cannot_update_task")

    order_item_id = getattr(task, "source_id", None)
    item = _operations_board_stage_order_item(factory_id, order_item_id)
    if not item:
        return False, translate("ops_error_order_line_unavailable")

    stage_meta = OPS_STAGE_BY_TASK_TYPE.get(getattr(task, "task_type", None) or "")
    if not stage_meta:
        return False, translate("ops_error_not_floor_stage_task")

    current_index = OPS_STAGE_INDEX.get(stage_meta["key"], -1)
    if current_index < 0:
        return False, translate("ops_error_stage_not_recognized")

    _append_operational_task_note(task, note, prefix=translate("ops_task_finished_prefix"))
    task.status = "done"
    task.assigned_user_id = getattr(current_user, "id", None)
    task.closed_by_id = getattr(current_user, "id", None)
    task.closed_at = datetime.utcnow()

    next_stage = OPS_STAGE_FLOW[current_index + 1] if current_index + 1 < len(OPS_STAGE_FLOW) else None
    if not next_stage:
        db.session.commit()
        return True, translate("ops_stage_final_completed_message")

    plan = (
        ProductionPlan.query
        .filter(
            ProductionPlan.factory_id == factory_id,
            ProductionPlan.order_item_id == order_item_id,
        )
        .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
        .first()
    )

    next_task = OperationalTask(
        factory_id=factory_id,
        created_by_id=current_user.id,
        task_type=next_stage["task_type"],
        source_type="ops_stage_flow",
        source_id=order_item_id,
        title=f"{translate(next_stage['label_key'])}: {item.product.name if item.product else f'Product #{item.product_id}'}",
        description=f"{translate(next_stage['description_key'])} Remaining qty: {int(item.qty_remaining or 0)}.",
        action_url=(
            url_for("main.production_plan_detail", plan_id=plan.id)
            if plan else url_for("main.production_plan", order_item_id=order_item_id)
        ),
        target_role=next_stage["target_role"],
        priority=next_stage["priority"],
        status="open",
        is_system_generated=False,
    )
    db.session.add(next_task)
    db.session.commit()
    return True, translate("ops_stage_moved_message").format(stage=translate(next_stage['label_key']).lower())


def _worker_stage_task_rows(factory_id: int | None):
    if not factory_id:
        return {"available": [], "mine": [], "recent_done": []}

    workspace = Factory.query.get(factory_id)
    if not workspace:
        return {"available": [], "mine": [], "recent_done": []}

    current_role = str(getattr(current_user, "role", None) or "").strip().lower()

    def build_row(task):
        item = _operations_board_stage_order_item(factory_id, getattr(task, "source_id", None))
        stage_meta = OPS_STAGE_BY_TASK_TYPE.get(getattr(task, "task_type", None) or "", {})
        plan = (
            ProductionPlan.query
            .filter(
                ProductionPlan.factory_id == factory_id,
                ProductionPlan.order_item_id == getattr(task, "source_id", None),
            )
            .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
            .first()
        )
        return {
            "task_id": task.id,
            "order_item_id": getattr(task, "source_id", None),
            "order_id": item.order_id if item else None,
            "product_name": item.product.name if item and item.product else display_value(getattr(task, "title", None), fallback="Task"),
            "customer_name": item.order.customer_name if item and item.order and item.order.customer_name else translate("ops_customer_walk_in"),
            "qty_remaining": int(item.qty_remaining or 0) if item else 0,
            "stage_key": stage_meta.get("key") or "stage",
            "stage_label": translate(stage_meta.get("label_key")) if stage_meta.get("label_key") else translate("ops_stage_fallback"),
            "status": (task.status or "open").strip().lower(),
            "assignee_name": get_user_display_name(task.assigned_user) if getattr(task, "assigned_user", None) else translate("ops_assignee_unassigned"),
            "description": task.description or "",
            "plan_href": (
                url_for("main.production_plan_detail", plan_id=plan.id)
                if plan else url_for("main.production_plan", order_item_id=getattr(task, "source_id", None))
            ),
            "order_href": url_for("shop.history_by_order", order_id=item.order_id) if item else None,
            "can_accept": (
                _can_update_operational_task(task, workspace)
                and not getattr(task, "assigned_user_id", None)
                and (getattr(task, "target_role", None) or "").strip().lower() == current_role
            ),
            "is_mine": getattr(task, "assigned_user_id", None) == getattr(current_user, "id", None),
        }

    open_rows = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "ops_stage_flow",
            OperationalTask.status.in_(("open", "in_progress")),
        )
        .order_by(OperationalTask.updated_at.desc(), OperationalTask.id.desc())
        .all()
    )
    open_rows = [task for task in open_rows if _can_view_operational_task(task, workspace)]

    available = []
    mine = []
    for task in open_rows:
        target_role = (getattr(task, "target_role", None) or "").strip().lower()
        if getattr(task, "assigned_user_id", None) == getattr(current_user, "id", None):
            mine.append(build_row(task))
        elif not getattr(task, "assigned_user_id", None) and target_role == current_role:
            available.append(build_row(task))

    done_rows = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "ops_stage_flow",
            OperationalTask.assigned_user_id == getattr(current_user, "id", None),
            OperationalTask.status == "done",
        )
        .order_by(OperationalTask.closed_at.desc(), OperationalTask.id.desc())
        .limit(8)
        .all()
    )

    recent_done = [build_row(task) for task in done_rows]
    return {"available": available, "mine": mine, "recent_done": recent_done}


def _ops_stage_history_map(factory_id: int | None, order_item_ids: list[int] | None, *, include_open: bool = True):
    if not factory_id or not order_item_ids:
        return {}

    statuses = ["done"]
    if include_open:
        statuses.extend(["open", "in_progress"])

    rows = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "ops_stage_flow",
            OperationalTask.source_id.in_(order_item_ids),
            OperationalTask.status.in_(tuple(statuses)),
        )
        .order_by(
            OperationalTask.source_id.asc(),
            OperationalTask.created_at.asc(),
            OperationalTask.id.asc(),
        )
        .all()
    )

    result = {order_item_id: [] for order_item_id in order_item_ids}
    for row in rows:
        stage_meta = OPS_STAGE_BY_TASK_TYPE.get(getattr(row, "task_type", None) or "", {})
        order_item_id = getattr(row, "source_id", None)
        if not order_item_id:
            continue
        result.setdefault(order_item_id, []).append({
            "task_id": row.id,
            "stage_key": stage_meta.get("key") or "stage",
            "stage_label": translate(stage_meta.get("label_key")) if stage_meta.get("label_key") else translate("ops_stage_fallback"),
            "status": (row.status or "open").strip().lower(),
            "assignee_name": get_user_display_name(row.assigned_user) if getattr(row, "assigned_user", None) else translate("ops_assignee_unassigned"),
            "created_at": getattr(row, "created_at", None),
            "updated_at": getattr(row, "updated_at", None),
            "closed_at": getattr(row, "closed_at", None),
            "description": row.description or "",
        })
    return result


def _build_operations_board_state(factory_id: int | None):
    state = {
        "queue_rows": [],
        "queue_counts": {
            "all": 0,
            "needs_plan": 0,
            "waiting_materials": 0,
            "ready_to_start": 0,
            "in_production": 0,
            "ready_to_ship": 0,
            "cutting": 0,
            "sewing": 0,
            "packing": 0,
            "ready": 0,
        },
        "cost_rows": [],
        "alert_rows": [],
        "open_task_rows": [],
        "unpaid_supplier_count": 0,
        "blocked_plan_count": 0,
        "receipt_pending_cost_count": 0,
    }
    if not factory_id:
        return state

    worker_options = [
        {
            "id": user.id,
            "label": get_user_display_name(user),
            "role": _format_role_label(getattr(user, "role", None)),
        }
        for user in _get_workspace_team_users(factory_id)
        if str(getattr(user, "role", None) or "").strip().lower() in {"viewer", "manager"}
    ]

    order_items = (
        ShopOrderItem.query
        .join(Product, Product.id == ShopOrderItem.product_id)
        .join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
        .filter(Product.factory_id == factory_id)
        .filter(ShopOrder.status == "pending")
        .filter(ShopOrderItem.qty_remaining > 0)
        .order_by(ShopOrder.created_at.asc(), ShopOrderItem.id.asc())
        .limit(24)
        .all()
    )

    order_item_ids = [int(row.id) for row in order_items]
    latest_plan_map = {}
    active_stage_map = _active_ops_stage_map(factory_id, order_item_ids)
    stage_history_map = _ops_stage_history_map(factory_id, order_item_ids)
    if order_item_ids:
        plan_rows = (
            ProductionPlan.query
            .filter(
                ProductionPlan.factory_id == factory_id,
                ProductionPlan.order_item_id.in_(order_item_ids),
            )
            .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
            .all()
        )
        for row in plan_rows:
            order_item_id = getattr(row, "order_item_id", None)
            if order_item_id and order_item_id not in latest_plan_map:
                latest_plan_map[order_item_id] = row

    for item in order_items:
        plan = latest_plan_map.get(item.id)
        run_summary = _production_run_summary(factory_id=factory_id, plan_id=plan.id) if plan else None
        target_qty = int(plan.target_qty or 0) if plan else int(item.qty_remaining or 0)
        progress = _plan_execution_progress(target_qty=target_qty, run_summary=run_summary)
        active_stage = active_stage_map.get(item.id)

        stage_key = "needs_plan"
        stage_label = translate("ops_queue_needs_plan_label")
        action_label = translate("ops_queue_plan_order_action")
        action_href = url_for("main.production_plan", order_item_id=item.id)
        tone = "secondary"

        if plan:
            action_label = translate("ops_queue_open_plan_action")
            action_href = url_for("main.production_plan_detail", plan_id=plan.id)
            if not bool(plan.can_fulfill_plan):
                stage_key = "waiting_materials"
                stage_label = translate("ops_queue_waiting_materials_label")
                tone = "danger"
            elif progress.get("status") == "completed":
                stage_key = "ready_to_ship"
                stage_label = translate("ops_queue_ready_to_ship_label")
                tone = "success"
            elif progress.get("status") == "in_progress":
                stage_key = "in_production"
                stage_label = translate("ops_queue_in_production_label")
                tone = "primary"
            else:
                stage_key = "ready_to_start"
                stage_label = translate("ops_queue_ready_to_start_label")
                tone = "warning"

        stage_action_label = None
        stage_action_href = None
        can_start_stage_flow = bool(plan) and bool(plan.can_fulfill_plan) and stage_key == "ready_to_start"

        if active_stage:
            stage_key = active_stage["stage_key"]
            stage_label = active_stage["stage_label"]
            tone = (
                "warning" if stage_key == "cutting"
                else "primary" if stage_key == "sewing"
                else "secondary" if stage_key == "packing"
                else "success"
            )
            stage_position = OPS_STAGE_INDEX.get(stage_key, 0)
            next_stage = OPS_STAGE_FLOW[stage_position + 1] if stage_position + 1 < len(OPS_STAGE_FLOW) else None
            if next_stage:
                stage_action_label = next_stage["label"]
                stage_action_href = url_for("main.operations_board_advance_stage", order_item_id=item.id)
        elif can_start_stage_flow:
            stage_action_label = translate("ops_queue_start_cutting_action")
            stage_action_href = url_for("main.operations_board_start_stage", order_item_id=item.id)

        state["queue_counts"]["all"] += 1
        state["queue_counts"][stage_key] += 1
        state["queue_rows"].append({
            "order_item_id": item.id,
            "order_id": item.order_id,
            "product_name": item.product.name if item.product else f"Product #{item.product_id}",
            "product_id": item.product_id,
            "customer_name": item.order.customer_name if item.order and item.order.customer_name else translate("ops_customer_walk_in"),
            "created_at": item.order.created_at if item.order else None,
            "qty_requested": int(item.qty_requested or 0),
            "qty_remaining": int(item.qty_remaining or 0),
            "qty_from_shop_now": int(item.qty_from_shop_now or 0),
            "factory_stock_now": int(getattr(item.product, "quantity", 0) or 0) if item.product else 0,
            "stage_key": stage_key,
            "stage_label": stage_label,
            "tone": tone,
            "action_label": action_label,
            "action_href": action_href,
            "plan_id": plan.id if plan else None,
            "progress": progress if plan else None,
            "active_stage": active_stage,
            "stage_history": stage_history_map.get(item.id, []),
            "stage_action_label": stage_action_label,
            "stage_action_href": stage_action_href,
            "can_start_stage_flow": can_start_stage_flow,
            "worker_options": worker_options,
            "task_exists": bool(
                OperationalTask.query
                .filter(
                    OperationalTask.factory_id == factory_id,
                    OperationalTask.source_type == "ops_order_queue",
                    OperationalTask.source_id == item.id,
                    OperationalTask.status.in_(("open", "in_progress")),
                )
                .first()
            ),
        })

    receipt_cost_map = _latest_supplier_cost_map(factory_id)
    product_rows = (
        Product.query
        .filter(Product.factory_id == factory_id)
        .order_by(Product.name.asc())
        .all()
    )

    for product in product_rows:
        composition_items = sorted(
            list(getattr(product, "composition_items", []) or []),
            key=lambda row: (
                getattr(getattr(row, "fabric", None), "material_type", "fabric"),
                getattr(getattr(row, "fabric", None), "name", ""),
            ),
        )
        if not composition_items:
            continue

        estimated_material_cost = 0.0
        estimated_currency = None
        missing_cost_count = 0
        cost_sources = []
        for row in composition_items:
            fabric = getattr(row, "fabric", None)
            if not fabric:
                continue
            source = receipt_cost_map.get(fabric.id)
            unit_cost = None
            currency = None
            source_label = None
            if source and source.get("unit_cost") is not None:
                unit_cost = float(source["unit_cost"] or 0)
                currency = source.get("currency") or "UZS"
                source_label = translate("cost_source_latest_receipt")
            elif getattr(fabric, "price_per_unit", None) is not None:
                unit_cost = float(fabric.price_per_unit or 0)
                currency = (fabric.price_currency or "UZS").upper()
                source_label = translate("cost_source_material_price")

            if unit_cost is None:
                missing_cost_count += 1
                continue

            if estimated_currency is None:
                estimated_currency = currency

            if estimated_currency != currency:
                missing_cost_count += 1
                continue

            estimated_material_cost += float(row.quantity_required or 0) * unit_cost
            cost_sources.append(source_label)

        if estimated_currency is None:
            estimated_currency = (product.currency or "UZS").upper()

        saved_cost = float(product.cost_price_per_item or 0)
        sell_price = float(product.sell_price_per_item or 0)
        gross_margin = sell_price - estimated_material_cost if estimated_material_cost else sell_price - saved_cost
        state["cost_rows"].append({
            "product_id": product.id,
            "product_name": product.name,
            "category": product.category or "",
            "materials_count": len(composition_items),
            "estimated_material_cost": estimated_material_cost,
            "saved_cost": saved_cost,
            "sell_price": sell_price,
            "gross_margin": gross_margin,
            "currency": estimated_currency,
            "missing_cost_count": missing_cost_count,
            "cost_source_label": " + ".join(sorted(set(cost_sources))) if cost_sources else "no material pricing",
            "cost_href": url_for("products.product_cost", product_id=product.id),
            "composition_href": url_for("main.composition", product_id=product.id),
        })

    state["cost_rows"].sort(
        key=lambda row: (
            -int(row["missing_cost_count"] or 0),
            -float(row["estimated_material_cost"] or 0),
            row["product_name"],
        )
    )
    state["cost_rows"] = state["cost_rows"][:8]

    low_material_count = 0
    for item in Fabric.query.filter(Fabric.factory_id == factory_id).all():
        threshold = float(getattr(item, "min_stock_quantity", 0) or 0) or 5.0
        quantity = float(getattr(item, "quantity", 0) or 0)
        if quantity < threshold:
            low_material_count += 1

    unpaid_supplier_count = (
        SupplierReceipt.query
        .filter(
            SupplierReceipt.factory_id == factory_id,
            SupplierReceipt.payment_status == "unpaid",
        )
        .count()
    )
    receipt_pending_cost_count = (
        SupplierReceipt.query
        .filter(
            SupplierReceipt.factory_id == factory_id,
            SupplierReceipt.unit_cost.is_(None),
        )
        .count()
    )
    blocked_plan_count = (
        ProductionPlan.query
        .filter(
            ProductionPlan.factory_id == factory_id,
            ProductionPlan.can_fulfill_plan.is_(False),
        )
        .count()
    )

    state["unpaid_supplier_count"] = unpaid_supplier_count
    state["blocked_plan_count"] = blocked_plan_count
    state["receipt_pending_cost_count"] = receipt_pending_cost_count

    state["alert_rows"] = [
        {
            "title": "Blocked plans",
            "value": blocked_plan_count,
            "copy": "Saved production plans that still cannot run from current material stock.",
            "href": url_for("main.production_plan_history", status="blocked"),
            "action_label": "Open saved plans",
            "tone": "danger",
        },
        {
            "title": "Unpaid supplier receipts",
            "value": unpaid_supplier_count,
            "copy": "Supplier deliveries already received but still not marked as paid.",
            "href": url_for("fabrics.supplier_receipts", payment_status="unpaid"),
            "action_label": "Open receipt history",
            "tone": "warning",
        },
        {
            "title": "Materials below minimum",
            "value": low_material_count,
            "copy": "Materials currently below their minimum stock threshold.",
            "href": url_for("main.low_stock"),
            "action_label": "Open low stock",
            "tone": "warning",
        },
        {
            "title": "Receipts missing cost",
            "value": receipt_pending_cost_count,
            "copy": "Supplier receipts without unit cost, which weakens product costing quality.",
            "href": url_for("fabrics.supplier_receipts"),
            "action_label": "Open receipts",
            "tone": "secondary",
        },
    ]

    open_tasks = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.status.in_(("open", "in_progress")),
        )
        .order_by(
            OperationalTask.priority.desc(),
            OperationalTask.created_at.asc(),
            OperationalTask.id.asc(),
        )
        .limit(6)
        .all()
    )
    for task in open_tasks:
        state["open_task_rows"].append({
            "id": task.id,
            "title": task.title,
            "description": task.description or "",
            "priority": task.priority or "medium",
            "status": task.status or "open",
            "href": (task.action_url or "").strip() or url_for("main.command_center"),
        })

    return state


@main_bp.route("/inventory/operations-board")
@login_required
def operations_board():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    state = _build_operations_board_state(factory_id)
    return render_template(
        "inventory/operations_board.html",
        workspace_page_name=display_value(getattr(workspace, "name", None), fallback="Workspace"),
        workspace_role_label=_format_role_label(getattr(current_user, "role", None)),
        **state,
    )


@main_bp.route("/inventory/operations-board/order-items/<int:order_item_id>/task", methods=["POST"])
@login_required
def operations_board_create_queue_task(order_item_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    if not factory_id:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.operations_board"))

    item = (
        ShopOrderItem.query
        .join(Product, Product.id == ShopOrderItem.product_id)
        .join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
        .filter(
            ShopOrderItem.id == order_item_id,
            Product.factory_id == factory_id,
            ShopOrder.status == "pending",
        )
        .first()
    )
    if not item:
        flash("That order line is no longer available.", "warning")
        return redirect(url_for("main.operations_board"))

    existing = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "ops_order_queue",
            OperationalTask.source_id == order_item_id,
            OperationalTask.status.in_(("open", "in_progress")),
        )
        .first()
    )
    if existing:
        flash("A follow-up task already exists for this order line.", "info")
        return redirect(url_for("main.command_center"))

    latest_plan = (
        ProductionPlan.query
        .filter(
            ProductionPlan.factory_id == factory_id,
            ProductionPlan.order_item_id == order_item_id,
        )
        .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
        .first()
    )
    action_href = (
        url_for("main.production_plan_detail", plan_id=latest_plan.id)
        if latest_plan else
        url_for("main.production_plan", order_item_id=order_item_id)
    )
    task = OperationalTask(
        factory_id=factory_id,
        created_by_id=current_user.id,
        task_type="order_queue_followup",
        source_type="ops_order_queue",
        source_id=order_item_id,
        title=f"Follow order queue: {item.product.name if item.product else f'Product #{item.product_id}'}",
        description=(
            f"Customer: {(item.order.customer_name if item.order and item.order.customer_name else 'not set')} | "
            f"Remaining: {int(item.qty_remaining or 0)}"
        ),
        action_url=action_href,
        target_role="manager",
        priority="high" if int(item.qty_remaining or 0) >= 10 else "medium",
        status="open",
        is_system_generated=False,
    )
    db.session.add(task)
    db.session.commit()
    flash("Queue follow-up task added to Command Center.", "success")
    return redirect(url_for("main.command_center"))


def _operations_board_stage_order_item(factory_id: int | None, order_item_id: int):
    if not factory_id:
        return None
    return (
        ShopOrderItem.query
        .join(Product, Product.id == ShopOrderItem.product_id)
        .join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
        .filter(
            ShopOrderItem.id == order_item_id,
            Product.factory_id == factory_id,
            ShopOrder.status == "pending",
        )
        .first()
    )


@main_bp.route("/inventory/operations-board/order-items/<int:order_item_id>/stage/start", methods=["POST"])
@login_required
def operations_board_start_stage(order_item_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    ok, message = _start_ops_stage_for_order(factory_id, order_item_id)
    flash(message, "success" if ok else "warning")
    return redirect(url_for("main.operations_board"))


@main_bp.route("/inventory/operations-board/order-items/<int:order_item_id>/stage/advance", methods=["POST"])
@login_required
def operations_board_advance_stage(order_item_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    active_stage = _active_ops_stage_map(factory_id, [order_item_id]).get(order_item_id)
    if not active_stage:
        flash("No active workflow stage was found for this order line.", "warning")
        return redirect(url_for("main.operations_board"))

    current_task = OperationalTask.query.get(active_stage["task_id"])
    ok, message = _advance_ops_stage_for_task(factory_id, current_task)
    flash(message, "success" if ok else "warning")
    return redirect(url_for("main.operations_board"))


@main_bp.route("/inventory/my-work")
@login_required
def worker_floor():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    task_state = _worker_stage_task_rows(factory_id)
    return render_template(
        "inventory/worker_floor.html",
        workspace_page_name=display_value(getattr(workspace, "name", None), fallback="Workspace"),
        workspace_role_label=_format_role_label(getattr(current_user, "role", None)),
        worker_available_tasks=task_state["available"],
        worker_my_tasks=task_state["mine"],
        worker_recent_done=task_state["recent_done"],
    )


@main_bp.route("/inventory/my-work/tasks/<int:task_id>/accept", methods=["POST"])
@login_required
def worker_floor_accept(task_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    workspace_row = Factory.query.get(factory_id) if factory_id else None
    task = OperationalTask.query.filter_by(id=task_id, factory_id=factory_id, source_type="ops_stage_flow").first()
    if not task or not workspace_row or not _can_update_operational_task(task, workspace_row):
        flash("You cannot accept this task.", "warning")
        return redirect(url_for("main.worker_floor"))

    task.assigned_user_id = current_user.id
    if (task.status or "open").strip().lower() == "open":
        task.status = "in_progress"
    db.session.commit()
    flash("Task accepted into your work queue.", "success")
    return redirect(url_for("main.worker_floor"))


@main_bp.route("/inventory/my-work/tasks/<int:task_id>/partial", methods=["POST"])
@login_required
def worker_floor_partial(task_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    workspace_row = Factory.query.get(factory_id) if factory_id else None
    task = OperationalTask.query.filter_by(id=task_id, factory_id=factory_id, source_type="ops_stage_flow").first()
    if not task or not workspace_row or not _can_update_operational_task(task, workspace_row):
        flash("You cannot update this task.", "warning")
        return redirect(url_for("main.worker_floor"))

    note = (request.form.get("progress_note") or "").strip()
    if not note:
        flash("Write a short partial progress note first.", "warning")
        return redirect(url_for("main.worker_floor"))

    task.assigned_user_id = current_user.id
    task.status = "in_progress"
    _append_operational_task_note(task, note, prefix=translate("ops_task_partial_update_prefix"))
    db.session.commit()
    flash("Partial progress saved.", "success")
    return redirect(url_for("main.worker_floor"))


@main_bp.route("/inventory/my-work/tasks/<int:task_id>/finish", methods=["POST"])
@login_required
def worker_floor_finish(task_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    task = OperationalTask.query.filter_by(id=task_id, factory_id=factory_id, source_type="ops_stage_flow").first()
    note = (request.form.get("finish_note") or "").strip()
    ok, message = _advance_ops_stage_for_task(factory_id, task, note=note)
    flash(message, "success" if ok else "warning")
    return redirect(url_for("main.worker_floor"))


@main_bp.route("/inventory/stage-history")
@login_required
def stage_history():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    if not factory_id:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.inventory"))

    order_rows = (
        ShopOrderItem.query
        .join(Product, Product.id == ShopOrderItem.product_id)
        .join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
        .filter(Product.factory_id == factory_id)
        .order_by(ShopOrder.created_at.desc(), ShopOrderItem.id.desc())
        .limit(30)
        .all()
    )

    order_item_ids = [int(row.id) for row in order_rows]
    history_map = _ops_stage_history_map(factory_id, order_item_ids)
    stage_history_rows = []
    for row in order_rows:
        stage_entries = history_map.get(row.id, [])
        if not stage_entries:
            continue
        done_count = sum(1 for entry in stage_entries if entry.get("status") == "done")
        active_entry = next((entry for entry in reversed(stage_entries) if entry.get("status") in {"open", "in_progress"}), None)
        stage_history_rows.append({
            "order_item_id": row.id,
            "order_id": row.order_id,
            "product_name": row.product.name if row.product else f"Product #{row.product_id}",
            "customer_name": row.order.customer_name if row.order and row.order.customer_name else "Walk-in / not set",
            "qty_requested": int(row.qty_requested or 0),
            "qty_remaining": int(row.qty_remaining or 0),
            "entries": stage_entries,
            "done_count": done_count,
            "active_entry": active_entry,
            "order_href": url_for("shop.history_by_order", order_id=row.order_id),
            "plan_href": url_for("main.production_plan", order_item_id=row.id),
        })

    return render_template(
        "inventory/stage_history.html",
        workspace_page_name=display_value(getattr(workspace, "name", None), fallback="Workspace"),
        workspace_role_label=_format_role_label(getattr(current_user, "role", None)),
        stage_history_rows=stage_history_rows,
    )


@main_bp.route("/inventory/operations-board/order-items/<int:order_item_id>/assign", methods=["POST"])
@login_required
def operations_board_assign_worker(order_item_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    if not factory_id:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.operations_board"))

    if not (_can_manage_command_center(workspace) or getattr(current_user, "is_superadmin", False)):
        flash("Only workspace managers can assign floor-stage work.", "danger")
        return redirect(url_for("main.operations_board"))

    worker_id = request.form.get("assigned_user_id", type=int)
    if not worker_id:
        flash("Select a worker first.", "warning")
        return redirect(url_for("main.operations_board"))

    worker = (
        User.query
        .filter(User.id == worker_id, User.factory_id == factory_id)
        .first()
    )
    if not worker:
        flash("That worker is not part of this workspace.", "warning")
        return redirect(url_for("main.operations_board"))

    active_stage = _active_ops_stage_map(factory_id, [order_item_id]).get(order_item_id)
    if active_stage:
        task = OperationalTask.query.filter_by(
            id=active_stage["task_id"],
            factory_id=factory_id,
            source_type="ops_stage_flow",
        ).first()
        if not task:
            flash("The active stage task could not be loaded.", "warning")
            return redirect(url_for("main.operations_board"))
        task.assigned_user_id = worker.id
        if (task.status or "open").strip().lower() == "open":
            task.status = "in_progress"
        db.session.commit()
        flash(f"Assigned {active_stage['stage_label'].lower()} to {get_user_display_name(worker)}.", "success")
        return redirect(url_for("main.operations_board"))

    ok, message = _start_ops_stage_for_order(factory_id, order_item_id)
    if not ok:
        flash(message, "warning")
        return redirect(url_for("main.operations_board"))

    active_stage = _active_ops_stage_map(factory_id, [order_item_id]).get(order_item_id)
    if not active_stage:
        flash("Stage started, but assignment could not be completed.", "warning")
        return redirect(url_for("main.operations_board"))

    task = OperationalTask.query.filter_by(
        id=active_stage["task_id"],
        factory_id=factory_id,
        source_type="ops_stage_flow",
    ).first()
    if not task:
        flash("Stage started, but assignment could not be completed.", "warning")
        return redirect(url_for("main.operations_board"))

    task.assigned_user_id = worker.id
    task.status = "in_progress"
    db.session.commit()
    flash(f"Started {active_stage['stage_label'].lower()} and assigned it to {get_user_display_name(worker)}.", "success")
    return redirect(url_for("main.operations_board"))

@main_bp.route("/more")
@login_required
def more():
    return render_template("dashboard/more.html", **_build_more_cards())


@main_bp.route("/business-summary")
@login_required
def business_summary():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    state = _build_reports_hub_state(getattr(workspace, "id", None))
    return render_template("dashboard/reports_hub.html", **state)


@main_bp.route("/business-summary/pdf")
@login_required
def business_summary_pdf():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    state = _build_business_summary(getattr(workspace, "id", None))
    pdf_bytes = _build_business_summary_pdf(state)
    buffer = BytesIO(pdf_bytes)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=_summary_export_filename(state.get("workspace_page_name"), "business_summary", "pdf"),
        mimetype="application/pdf",
        max_age=0,
    )


@main_bp.route("/business-summary/xlsx")
@login_required
def business_summary_xlsx():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    state = _build_business_summary(getattr(workspace, "id", None))
    workbook_buffer = _build_business_summary_xlsx(state)
    return send_file(
        workbook_buffer,
        as_attachment=True,
        download_name=_summary_export_filename(state.get("workspace_page_name"), "business_summary", "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@main_bp.route("/business-summary/branches")
@login_required
def business_summary_branches():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    state = _build_business_summary(getattr(workspace, "id", None))
    return render_template("dashboard/business_summary_branches.html", **state)


@main_bp.route("/business-summary/products")
@login_required
def business_summary_products():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return redirect(url_for("main.dashboard"))

    state = _build_business_summary(getattr(workspace, "id", None))
    return render_template("dashboard/business_summary_products.html", **state)


def _purchase_needs_session_key() -> str:
    return "dismissed_purchase_need_material_ids"


def _get_dismissed_purchase_need_ids() -> set[int]:
    raw_value = session.get(_purchase_needs_session_key(), [])
    dismissed_ids = set()
    for item in raw_value:
        try:
            dismissed_ids.add(int(item))
        except (TypeError, ValueError):
            continue
    return dismissed_ids


@main_bp.route("/inventory/purchase-needs/dismiss", methods=["POST"])
@login_required
def dismiss_purchase_need():
    material_id = request.form.get("material_id", type=int)
    if material_id:
        dismissed_ids = _get_dismissed_purchase_need_ids()
        dismissed_ids.add(material_id)
        session[_purchase_needs_session_key()] = sorted(dismissed_ids)
        session.modified = True
    return redirect(url_for("main.purchase_needs"))


@main_bp.route("/inventory/purchase-needs")
@login_required
def purchase_needs():
    from app.models import Fabric, Product, ShopStock, Factory, ProductComposition

    LOW_MATERIAL_THRESHOLD = 5
    dismissed_material_ids = _get_dismissed_purchase_need_ids()

    all_materials = (
        Fabric.query
        .order_by(Fabric.name.asc(), Fabric.color.asc())
        .all()
    )

    low_materials = []
    for item in all_materials:
        threshold = float(getattr(item, "min_stock_quantity", 0) or 0) or float(LOW_MATERIAL_THRESHOLD)
        qty = float(getattr(item, "quantity", 0) or 0)
        if qty < threshold:
            low_materials.append(item)

    low_materials.sort(key=lambda row: ((row.quantity or 0), row.name or ""))
    low_materials = [
        item for item in low_materials
        if int(getattr(item, "id", 0) or 0) not in dismissed_material_ids
    ]

    low_shop_rows = (
        ShopStock.query
        .join(Product, Product.id == ShopStock.product_id)
        .outerjoin(Factory, Factory.id == ShopStock.source_factory_id)
        .filter(ShopStock.quantity > 0, ShopStock.quantity < 5)
        .order_by(ShopStock.quantity.asc(), Product.name.asc())
        .all()
    )

    low_shop_items = []
    for row in low_shop_rows:
        low_shop_items.append({
            "name": row.product.name if row.product else f"Product #{row.product_id}",
            "qty": row.quantity or 0,
            "source_factory_name": row.source_factory.name if getattr(row, "source_factory", None) else None,
            "image_path": row.product.image_path if row.product else None,
        })

    composition_items = (
        ProductComposition.query
        .join(Product, Product.id == ProductComposition.product_id)
        .join(Fabric, Fabric.id == ProductComposition.fabric_id)
        .order_by(Product.name.asc(), Fabric.name.asc())
        .all()
    )

    composition_alerts = []
    product_blockers_map = {}
    material_dependency_map = {}
    blocked_products_total = 0
    risk_products_total = 0

    for item in composition_items:
        required = item.quantity_required or 0
        available = item.fabric.quantity or 0
        status = "ok"
        threshold = float(getattr(item.fabric, "min_stock_quantity", 0) or 0) or float(LOW_MATERIAL_THRESHOLD)
        shortage_qty = max(float(required or 0) - float(available or 0), 0.0)
        reorder_gap = max(threshold - float(available or 0), 0.0)

        if required > 0:
            max_units = int((available or 0) // required)
        else:
            max_units = 0

        if available <= 0:
            status = "blocked"
        elif available < required:
            status = "risk"

        material_key = item.fabric_id
        if material_key not in material_dependency_map:
            material_dependency_map[material_key] = {
                "material_id": item.fabric_id,
                "material_name": item.fabric.name if item.fabric else f"Material #{item.fabric_id}",
                "material_type": getattr(item.fabric, "material_type", "fabric") or "fabric",
                "available": float(available or 0),
                "min_stock_quantity": float(threshold or 0),
                "unit": item.fabric.unit if item.fabric and item.fabric.unit else (item.unit or ""),
                "category": getattr(item.fabric, "category", "") or "",
                "supplier_name": getattr(item.fabric, "supplier_name", "") or "",
                "reorder_gap": max(float(threshold or 0) - float(available or 0), 0.0),
                "dependent_products": set(),
                "blocked_products": set(),
                "risk_products": set(),
            }

        material_dependency_map[material_key]["dependent_products"].add(
            item.product.name if item.product else f"Product #{item.product_id}"
        )

        if status != "ok":
            alert_row = {
                "product_name": item.product.name if item.product else f"Product #{item.product_id}",
                "material_name": item.fabric.name if item.fabric else f"Material #{item.fabric_id}",
                "material_type": getattr(item.fabric, "material_type", "fabric") or "fabric",
                "required": required,
                "available": available,
                "unit": item.unit or item.fabric.unit or "",
                "status": status,
                "shortage_qty": shortage_qty,
                "reorder_gap": reorder_gap,
                "max_units_now": max_units,
            }
            composition_alerts.append(alert_row)

            product_name = alert_row["product_name"]
            if product_name not in product_blockers_map:
                product_blockers_map[product_name] = {
                    "product_name": product_name,
                    "blocked_count": 0,
                    "risk_count": 0,
                    "max_units_now": None,
                    "materials": [],
                }

            product_entry = product_blockers_map[product_name]
            product_entry["materials"].append(alert_row)
            if status == "blocked":
                product_entry["blocked_count"] += 1
                material_dependency_map[material_key]["blocked_products"].add(product_name)
                blocked_products_total += 1
            else:
                product_entry["risk_count"] += 1
                material_dependency_map[material_key]["risk_products"].add(product_name)
                risk_products_total += 1

            if product_entry["max_units_now"] is None:
                product_entry["max_units_now"] = max_units
            else:
                product_entry["max_units_now"] = min(product_entry["max_units_now"], max_units)

    product_blockers = sorted(
        product_blockers_map.values(),
        key=lambda row: (
            -int(row["blocked_count"] or 0),
            -int(row["risk_count"] or 0),
            int(row["max_units_now"] or 0),
            row["product_name"],
        ),
    )

    material_dependency_rows = []
    for row in material_dependency_map.values():
        material_dependency_rows.append({
            "material_id": row["material_id"],
            "material_name": row["material_name"],
            "material_type": row["material_type"],
            "available": row["available"],
            "min_stock_quantity": row["min_stock_quantity"],
            "unit": row["unit"],
            "category": row["category"],
            "supplier_name": row["supplier_name"],
            "reorder_gap": row["reorder_gap"],
            "dependent_products_count": len(row["dependent_products"]),
            "blocked_products_count": len(row["blocked_products"]),
            "risk_products_count": len(row["risk_products"]),
        })

    material_dependency_rows.sort(
        key=lambda row: (
            -int(row["blocked_products_count"] or 0),
            -int(row["risk_products_count"] or 0),
            row["available"],
            row["material_name"],
        )
    )
    material_dependency_rows = [
        row for row in material_dependency_rows
        if int(row.get("material_id") or 0) not in dismissed_material_ids
    ]

    buy_now_rows = []
    for row in material_dependency_rows:
        urgency_score = (
            int(row["blocked_products_count"] or 0) * 100
            + int(row["risk_products_count"] or 0) * 35
            + int(row["dependent_products_count"] or 0) * 10
            + (15 if float(row["available"] or 0) <= 0 else 0)
        )
        if urgency_score <= 0 and float(row["reorder_gap"] or 0) <= 0:
            continue

        buy_now_rows.append({
            **row,
            "urgency_score": urgency_score,
            "status": "out" if float(row["available"] or 0) <= 0 else "low",
            "priority_reason": (
                f"Unlocks {row['blocked_products_count']} blocked product"
                f"{'' if int(row['blocked_products_count'] or 0) == 1 else 's'}"
                if int(row["blocked_products_count"] or 0) > 0
                else (
                    f"Protects {row['risk_products_count']} at-risk product"
                    f"{'' if int(row['risk_products_count'] or 0) == 1 else 's'}"
                    if int(row["risk_products_count"] or 0) > 0
                    else "Restore minimum stock"
                )
            ),
        })

    buy_now_rows.sort(
        key=lambda row: (
            -int(row["urgency_score"] or 0),
            row["available"],
            row["material_name"],
        )
    )
    buy_now_rows = buy_now_rows[:5]

    priority_products = []
    for row in product_blockers[:5]:
        top_material = None
        materials = sorted(
            row.get("materials") or [],
            key=lambda material: (
                0 if material.get("status") == "blocked" else 1,
                -float(material.get("shortage_qty") or 0),
                material.get("material_name") or "",
            ),
        )
        if materials:
            top_material = materials[0]
        priority_products.append({
            **row,
            "top_material": top_material,
        })

    return render_template(
        "inventory/purchase_needs.html",
        low_materials=low_materials,
        low_shop_items=low_shop_items,
        composition_alerts=composition_alerts,
        product_blockers=product_blockers,
        material_dependency_rows=material_dependency_rows,
        buy_now_rows=buy_now_rows,
        priority_products=priority_products,
        blocked_products_total=blocked_products_total,
        risk_products_total=risk_products_total,
        dismissed_material_ids=dismissed_material_ids,
    )

def _build_production_plan_state(factory_id: int | None, selected_product_id: int | None, target_qty: int | None, order_item_id: int | None = None):
    from app.models import Product, ProductComposition, Fabric

    products = (
        Product.query
        .filter(Product.factory_id == factory_id)
        .order_by(Product.name.asc())
        .all()
        if factory_id else []
    )

    pending_order_items = []
    source_order_item = None
    default_shop = _find_default_linked_shop(factory_id)
    if factory_id:
        pending_items_query = (
            ShopOrderItem.query
            .join(Product, Product.id == ShopOrderItem.product_id)
            .join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
            .filter(Product.factory_id == factory_id)
            .filter(ShopOrder.status == "pending")
            .filter(ShopOrderItem.qty_remaining > 0)
            .order_by(ShopOrder.created_at.asc(), ShopOrderItem.id.asc())
            .limit(10)
        )

        pending_order_items = [
            {
                "item_id": item.id,
                "order_id": item.order_id,
                "product_id": item.product_id,
                "product_name": item.product.name if item.product else f"Product #{item.product_id}",
                "category": item.product.category if item.product else "",
                "qty_requested": int(item.qty_requested or 0),
                "qty_from_shop_now": int(item.qty_from_shop_now or 0),
                "qty_remaining": int(item.qty_remaining or 0),
                "customer_name": item.order.customer_name if item.order else "",
                "created_at": item.order.created_at if item.order else None,
                "target_shop_id": default_shop.id if default_shop else None,
                "target_shop_name": default_shop.name if default_shop else "",
            }
            for item in pending_items_query.all()
        ]

        if order_item_id:
            source_order_item = next(
                (row for row in pending_order_items if row["item_id"] == order_item_id),
                None,
            )

            if source_order_item is None:
                item = (
                    ShopOrderItem.query
                    .join(Product, Product.id == ShopOrderItem.product_id)
                    .join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
                    .filter(ShopOrderItem.id == order_item_id)
                    .filter(Product.factory_id == factory_id)
                    .filter(ShopOrder.status == "pending")
                    .filter(ShopOrderItem.qty_remaining > 0)
                    .first()
                )
                if item:
                    source_order_item = {
                        "item_id": item.id,
                        "order_id": item.order_id,
                        "product_id": item.product_id,
                        "product_name": item.product.name if item.product else f"Product #{item.product_id}",
                        "category": item.product.category if item.product else "",
                        "qty_requested": int(item.qty_requested or 0),
                        "qty_from_shop_now": int(item.qty_from_shop_now or 0),
                        "qty_remaining": int(item.qty_remaining or 0),
                        "customer_name": item.order.customer_name if item.order else "",
                        "created_at": item.order.created_at if item.order else None,
                        "target_shop_id": default_shop.id if default_shop else None,
                        "target_shop_name": default_shop.name if default_shop else "",
                    }

    if selected_product_id is None and source_order_item:
        selected_product_id = source_order_item["product_id"]

    if target_qty is None and source_order_item:
        target_qty = max(int(source_order_item["qty_remaining"] or 0), 1)
    if target_qty is None or target_qty < 1:
        target_qty = 1
    if not source_order_item:
        order_item_id = None

    selected_product = None
    plan_rows = []
    blocking_rows = []
    shortage_rows = []
    purchase_rows = []
    max_producible_units = 0
    can_fulfill_plan = False

    if selected_product_id and factory_id:
        selected_product = (
            Product.query
            .filter(Product.id == selected_product_id, Product.factory_id == factory_id)
            .first()
        )

        if selected_product:
            composition_items = (
                ProductComposition.query
                .join(Fabric, Fabric.id == ProductComposition.fabric_id)
                .filter(ProductComposition.product_id == selected_product.id)
                .order_by(Fabric.material_type.asc(), Fabric.name.asc())
                .all()
            )

            max_units_candidates = []

            for item in composition_items:
                available = float(getattr(item.fabric, "quantity", 0) or 0)
                required_per_unit = float(item.quantity_required or 0)
                total_required = required_per_unit * int(target_qty or 0)
                shortage_qty = max(total_required - available, 0.0)
                min_stock_quantity = float(getattr(item.fabric, "min_stock_quantity", 0) or 0)
                suggested_purchase_qty = max((total_required + min_stock_quantity) - available, 0.0)
                unit = item.unit or getattr(item.fabric, "unit", "") or ""

                if required_per_unit > 0:
                    supported_units = int(available // required_per_unit)
                    max_units_candidates.append(supported_units)
                else:
                    supported_units = 0

                status = "ready"
                if available <= 0:
                    status = "blocked"
                elif shortage_qty > 0:
                    status = "short"

                row = {
                    "fabric_id": item.fabric_id,
                    "material_name": item.fabric.name if item.fabric else f"Material #{item.fabric_id}",
                    "material_type": getattr(item.fabric, "material_type", "fabric") or "fabric",
                    "available": available,
                    "required_per_unit": required_per_unit,
                    "total_required": total_required,
                    "shortage_qty": shortage_qty,
                    "unit": unit,
                    "status": status,
                    "supported_units": supported_units,
                    "min_stock_quantity": min_stock_quantity,
                    "suggested_purchase_qty": suggested_purchase_qty,
                }
                plan_rows.append(row)

                if status in ("blocked", "short"):
                    blocking_rows.append(row)
                if shortage_qty > 0:
                    shortage_rows.append(row)
                    purchase_rows.append(row)

            if max_units_candidates:
                max_producible_units = min(max_units_candidates)
            elif selected_product:
                max_producible_units = 0

            can_fulfill_plan = bool(plan_rows) and max_producible_units >= int(target_qty or 0) and not blocking_rows

    return {
        "products": products,
        "pending_order_items": pending_order_items,
        "source_order_item": source_order_item,
        "order_item_id": order_item_id,
        "selected_product": selected_product,
        "selected_product_id": selected_product_id,
        "target_qty": target_qty,
        "plan_rows": plan_rows,
        "blocking_rows": blocking_rows,
        "shortage_rows": shortage_rows,
        "purchase_rows": purchase_rows,
        "max_producible_units": max_producible_units,
        "can_fulfill_plan": can_fulfill_plan,
    }


def _recent_production_plans(factory_id: int | None, limit: int = 8):
    if not factory_id:
        return []

    rows = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.factory_id == factory_id)
        .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
        .limit(limit)
        .all()
    )

    result = []
    for row in rows:
        plan_state = _build_production_plan_state(
            factory_id=factory_id,
            selected_product_id=row.product_id,
            target_qty=int(row.target_qty or 0),
            order_item_id=row.order_item_id,
        )
        task_summary = _linked_production_plan_task_summary(factory_id=factory_id, plan_id=row.id)
        run_summary = _production_run_summary(factory_id=factory_id, plan_id=row.id)
        progress = _plan_execution_progress(
            target_qty=int(row.target_qty or 0),
            run_summary=run_summary,
        )
        shortage_preview = [
            {
                "material_name": item.get("material_name") or "",
                "material_type": item.get("material_type") or "fabric",
                "shortage_qty": float(item.get("shortage_qty") or 0),
                "unit": item.get("unit") or "",
            }
            for item in (plan_state.get("purchase_rows") or [])[:3]
        ]
        order_item = getattr(row, "order_item", None)
        order = getattr(order_item, "order", None) if order_item else None
        result.append({
            "id": row.id,
            "product_id": row.product_id,
            "product_name": row.product.name if row.product else f"Product #{row.product_id}",
            "target_qty": int(row.target_qty or 0),
            "max_producible_units": int(row.max_producible_units or 0),
            "shortage_count": int(row.shortage_count or 0),
            "can_fulfill_plan": bool(row.can_fulfill_plan),
            "order_item_id": row.order_item_id,
            "order_id": order.id if order else None,
            "customer_name": order.customer_name if order else "",
            "note": row.note or "",
            "created_at": row.created_at,
            "shortage_preview": shortage_preview,
            "task_summary": task_summary,
            "run_summary": run_summary,
            "progress": progress,
            "movement_history_href": url_for("shop.movement_history", product_id=row.product_id),
            "order_history_href": url_for("shop.history_by_order", order_id=order.id) if order else None,
        })
    return result


def _query_saved_production_plans(
    factory_id: int | None,
    *,
    limit: int = 50,
    product_id: int | None = None,
    status: str | None = None,
    q: str | None = None,
):
    if not factory_id:
        return []

    query = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.factory_id == factory_id)
    )

    if product_id:
        query = query.filter(ProductionPlan.product_id == product_id)

    normalized_status = str(status or "").strip().lower()
    if normalized_status == "ready":
        query = query.filter(ProductionPlan.can_fulfill_plan.is_(True))
    elif normalized_status == "blocked":
        query = query.filter(ProductionPlan.can_fulfill_plan.is_(False))

    q_value = str(q or "").strip()
    if q_value:
        like_value = f"%{q_value}%"
        query = query.filter(
            or_(
                Product.name.ilike(like_value),
                ProductionPlan.note.ilike(like_value),
            )
        )

    rows = (
        query
        .order_by(ProductionPlan.created_at.desc(), ProductionPlan.id.desc())
        .limit(limit)
        .all()
    )

    result = []
    for row in rows:
        plan_state = _build_production_plan_state(
            factory_id=factory_id,
            selected_product_id=row.product_id,
            target_qty=int(row.target_qty or 0),
            order_item_id=row.order_item_id,
        )
        task_summary = _linked_production_plan_task_summary(factory_id=factory_id, plan_id=row.id)
        run_summary = _production_run_summary(factory_id=factory_id, plan_id=row.id)
        progress = _plan_execution_progress(
            target_qty=int(row.target_qty or 0),
            run_summary=run_summary,
        )
        shortage_preview = [
            {
                "material_name": item.get("material_name") or "",
                "material_type": item.get("material_type") or "fabric",
                "shortage_qty": float(item.get("shortage_qty") or 0),
                "unit": item.get("unit") or "",
            }
            for item in (plan_state.get("purchase_rows") or [])[:3]
        ]
        order_item = getattr(row, "order_item", None)
        order = getattr(order_item, "order", None) if order_item else None
        row_payload = {
            "id": row.id,
            "product_id": row.product_id,
            "product_name": row.product.name if row.product else f"Product #{row.product_id}",
            "target_qty": int(row.target_qty or 0),
            "max_producible_units": int(row.max_producible_units or 0),
            "shortage_count": int(row.shortage_count or 0),
            "can_fulfill_plan": bool(row.can_fulfill_plan),
            "order_item_id": row.order_item_id,
            "order_id": order.id if order else None,
            "customer_name": order.customer_name if order else "",
            "note": row.note or "",
            "created_at": row.created_at,
            "shortage_preview": shortage_preview,
            "task_summary": task_summary,
            "run_summary": run_summary,
            "progress": progress,
            "movement_history_href": url_for("shop.movement_history", product_id=row.product_id),
            "order_history_href": url_for("shop.history_by_order", order_id=order.id) if order else None,
        }
        result.append(row_payload)

    if normalized_status in {"not_started", "in_progress", "completed"}:
        result = [row for row in result if (row.get("progress") or {}).get("status") == normalized_status]
    return result


def _production_runs_for_plan(factory_id: int | None, plan_id: int, limit: int = 10):
    if not factory_id:
        return []

    rows = (
        Production.query
        .join(Product, Product.id == Production.product_id)
        .filter(
            Production.production_plan_id == plan_id,
            Product.factory_id == factory_id,
        )
        .order_by(Production.date.desc(), Production.id.desc())
        .limit(limit)
        .all()
    )

    result = []
    for row in rows:
        consumptions = []
        for consumption in sorted((row.consumptions or []), key=lambda item: item.id):
            consumptions.append({
                "material_name": consumption.fabric.name if consumption.fabric else f"Material #{consumption.fabric_id}",
                "material_type": getattr(consumption.fabric, "material_type", "fabric") if consumption.fabric else "fabric",
                "used_amount": float(consumption.used_amount or 0),
                "unit": (consumption.fabric.unit if consumption.fabric else "") or "",
            })

        result.append({
            "id": row.id,
            "date": row.date,
            "quantity": int(row.quantity or 0),
            "note": row.note or "",
            "consumptions": consumptions,
        })
    return result


def _production_run_detail(factory_id: int | None, run_id: int):
    if not factory_id:
        return None

    row = (
        Production.query
        .join(Product, Product.id == Production.product_id)
        .filter(
            Production.id == run_id,
            Product.factory_id == factory_id,
        )
        .first()
    )
    if not row:
        return None

    consumptions = []
    total_material_rows = 0
    for consumption in sorted((row.consumptions or []), key=lambda item: item.id):
        total_material_rows += 1
        fabric = consumption.fabric
        consumptions.append({
            "material_id": consumption.fabric_id,
            "material_name": fabric.name if fabric else f"Material #{consumption.fabric_id}",
            "material_type": getattr(fabric, "material_type", "fabric") if fabric else "fabric",
            "used_amount": float(consumption.used_amount or 0),
            "unit": (fabric.unit if fabric else "") or "",
            "material_stock_now": float(getattr(fabric, "quantity", 0) or 0) if fabric else 0.0,
        })

    plan = getattr(row, "production_plan", None)
    order_item = getattr(plan, "order_item", None) if plan else None
    order = getattr(order_item, "order", None) if order_item else None
    product_stock_now = int(getattr(row.product, "quantity", 0) or 0) if row.product else 0
    can_rollback = product_stock_now >= int(row.quantity or 0)
    rollback_block_reason = None
    if not can_rollback:
        rollback_block_reason = "Factory stock is already below the produced quantity, so this run has likely been transferred, shipped, or sold."

    return {
        "id": row.id,
        "date": row.date,
        "quantity": int(row.quantity or 0),
        "note": row.note or "",
        "product_id": row.product_id,
        "product_name": row.product.name if row.product else f"Product #{row.product_id}",
        "product_stock_now": product_stock_now,
        "production_plan_id": row.production_plan_id,
        "consumptions": consumptions,
        "total_material_rows": total_material_rows,
        "movement_history_href": url_for("shop.movement_history", product_id=row.product_id),
        "order_id": order.id if order else None,
        "order_history_href": url_for("shop.history_by_order", order_id=order.id) if order else None,
        "can_rollback": can_rollback,
        "rollback_block_reason": rollback_block_reason,
    }


def _production_run_summary(factory_id: int | None, plan_id: int):
    runs = _production_runs_for_plan(factory_id=factory_id, plan_id=plan_id, limit=50)
    total_qty = sum(int(run.get("quantity") or 0) for run in runs)
    return {
        "count": len(runs),
        "total_qty": total_qty,
        "last_run": runs[0] if runs else None,
    }


def _shop_stock_breakdown_for_product(factory_id: int | None, product_id: int | None):
    if not factory_id or not product_id:
        return []

    rows = (
        ShopStock.query
        .join(Shop, Shop.id == ShopStock.shop_id)
        .filter(
            ShopStock.source_factory_id == factory_id,
            ShopStock.product_id == product_id,
        )
        .order_by(Shop.name.asc(), ShopStock.id.asc())
        .all()
    )

    result = []
    for row in rows:
        result.append({
            "shop_id": row.shop_id,
            "shop_name": row.shop.name if row.shop else f"Shop #{row.shop_id}",
            "quantity": int(row.quantity or 0),
            "transfer_href": url_for("shop.transfer_to_shop"),
            "orders_href": url_for("shop.list_shop_orders"),
        })
    return result


def _plan_execution_progress(*, target_qty: int, run_summary: dict | None):
    target_qty = max(int(target_qty or 0), 0)
    executed_qty = int((run_summary or {}).get("total_qty") or 0)
    remaining_qty = max(target_qty - executed_qty, 0)
    completion_pct = 0
    if target_qty > 0:
        completion_pct = min(int((executed_qty / target_qty) * 100), 100)

    if executed_qty <= 0:
        status = "not_started"
        label = "Not started"
    elif executed_qty >= target_qty > 0:
        status = "completed"
        label = "Completed"
    else:
        status = "in_progress"
        label = "In progress"

    return {
        "executed_qty": executed_qty,
        "remaining_qty": remaining_qty,
        "completion_pct": completion_pct,
        "status": status,
        "label": label,
        "is_complete": status == "completed",
        "has_started": executed_qty > 0,
    }


def _linked_production_plan_tasks(factory_id: int | None, plan_id: int):
    if not factory_id:
        return []

    rows = (
        OperationalTask.query
        .filter(OperationalTask.factory_id == factory_id)
        .filter(
            or_(
                and_(OperationalTask.source_type == "production_plan_blocked", OperationalTask.source_id == plan_id),
                and_(OperationalTask.source_type == "production_plan_purchase", OperationalTask.source_id == plan_id),
            )
        )
        .order_by(OperationalTask.created_at.desc(), OperationalTask.id.desc())
        .all()
    )
    return [_serialize_operational_task(task) for task in rows]


def _linked_production_plan_task_summary(factory_id: int | None, plan_id: int):
    tasks = _linked_production_plan_tasks(factory_id=factory_id, plan_id=plan_id)
    open_count = sum(1 for task in tasks if task.get("status") in {"open", "in_progress"})
    done_count = sum(1 for task in tasks if task.get("status") == "done")
    blocked_task = next((task for task in tasks if task.get("origin") == "manual" and "production plan" in str(task.get("title") or "").lower()), None)
    purchase_task = next((task for task in tasks if "buy materials" in str(task.get("title") or "").lower()), None)
    return {
        "total": len(tasks),
        "open": open_count,
        "done": done_count,
        "has_tasks": bool(tasks),
        "has_open_tasks": open_count > 0,
        "has_blocked_task": blocked_task is not None,
        "has_purchase_task": purchase_task is not None,
    }


def _sync_production_plan_purchase_tasks(factory_id: int | None, plan_id: int, plan_state: dict | None = None):
    if not factory_id:
        return

    plan = (
        ProductionPlan.query
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first()
    )
    if not plan:
        return

    if plan_state is None:
        plan_state = _build_production_plan_state(
            factory_id=factory_id,
            selected_product_id=plan.product_id,
            target_qty=int(plan.target_qty or 0),
            order_item_id=plan.order_item_id,
        )

    purchase_rows = plan_state.get("purchase_rows") or []
    purchase_tasks = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "production_plan_purchase",
            OperationalTask.source_id == plan_id,
        )
        .all()
    )

    should_be_done = not purchase_rows
    for task in purchase_tasks:
        if should_be_done and task.status in {"open", "in_progress"}:
            task.status = "done"
            task.updated_at = datetime.utcnow()
            task.closed_at = datetime.utcnow()
            task.closed_by_id = getattr(current_user, "id", None)
        elif not should_be_done and task.status == "done":
            task.status = "open"
            task.updated_at = datetime.utcnow()
            task.closed_at = None
            task.closed_by_id = None

    plan.shortage_count = len(plan_state.get("shortage_rows") or [])
    plan.max_producible_units = int(plan_state.get("max_producible_units") or 0)
    plan.can_fulfill_plan = bool(plan_state.get("can_fulfill_plan"))
    db.session.commit()


def _build_production_plan_xlsx(plan_state):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Plan Summary"
    ws_summary.append(["Field", "Value"])
    ws_summary.append(["Product", getattr(plan_state.get("selected_product"), "name", "") or "-"])
    ws_summary.append(["Target quantity", int(plan_state.get("target_qty") or 0)])
    ws_summary.append(["Max buildable", int(plan_state.get("max_producible_units") or 0)])
    ws_summary.append(["Plan status", "Ready" if plan_state.get("can_fulfill_plan") else "Blocked"])

    source_order_item = plan_state.get("source_order_item")
    if source_order_item:
        ws_summary.append(["Source order", f"Order #{source_order_item['order_id']}"])
        ws_summary.append(["Source customer", source_order_item.get("customer_name") or "-"])
        ws_summary.append(["Order remaining qty", int(source_order_item.get("qty_remaining") or 0)])

    _summary_xlsx_style_sheet(ws_summary)
    _summary_xlsx_fit_columns(ws_summary)

    ws_materials = wb.create_sheet("Materials")
    ws_materials.append(["Material", "Type", "Unit", "Available", "Per unit", "Total needed", "Shortage", "Suggested buy", "Supports units", "Status"])
    for row in plan_state.get("plan_rows", []):
        ws_materials.append([
            row.get("material_name") or "",
            row.get("material_type") or "fabric",
            row.get("unit") or "",
            float(row.get("available") or 0),
            float(row.get("required_per_unit") or 0),
            float(row.get("total_required") or 0),
            float(row.get("shortage_qty") or 0),
            float(row.get("suggested_purchase_qty") or 0),
            int(row.get("supported_units") or 0),
            row.get("status") or "",
        ])
    _summary_xlsx_style_sheet(ws_materials)
    _summary_xlsx_fit_columns(ws_materials)

    ws_purchase = wb.create_sheet("Purchase List")
    ws_purchase.append(["Material", "Type", "Unit", "Missing for target", "Min stock", "Suggested buy"])
    for row in plan_state.get("purchase_rows", []):
        ws_purchase.append([
            row.get("material_name") or "",
            row.get("material_type") or "fabric",
            row.get("unit") or "",
            float(row.get("shortage_qty") or 0),
            float(row.get("min_stock_quantity") or 0),
            float(row.get("suggested_purchase_qty") or 0),
        ])
    _summary_xlsx_style_sheet(ws_purchase)
    _summary_xlsx_fit_columns(ws_purchase)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@main_bp.route("/inventory/production-plan")
@login_required
def production_plan():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    selected_product_id = request.args.get("product_id", type=int)
    target_qty = request.args.get("target_qty", type=int)
    order_item_id = request.args.get("order_item_id", type=int)

    plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=selected_product_id,
        target_qty=target_qty,
        order_item_id=order_item_id,
    )
    plan_state["recent_saved_plans"] = _recent_production_plans(factory_id=factory_id, limit=6)

    return render_template(
        "inventory/production_plan.html",
        **plan_state,
    )


@main_bp.route("/inventory/production-plan/save", methods=["POST"])
@login_required
def production_plan_save():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    if not factory_id:
        flash("A factory workspace is required to save a production plan.", "warning")
        return redirect(url_for("main.production_plan"))

    selected_product_id = request.form.get("product_id", type=int)
    target_qty = request.form.get("target_qty", type=int)
    order_item_id = request.form.get("order_item_id", type=int)
    note = (request.form.get("note") or "").strip()[:255]

    plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=selected_product_id,
        target_qty=target_qty,
        order_item_id=order_item_id,
    )

    selected_product = plan_state.get("selected_product")
    if not selected_product:
        flash("Run a plan with a valid product before saving it.", "warning")
        return redirect(url_for("main.production_plan"))

    plan = ProductionPlan(
        factory_id=factory_id,
        product_id=selected_product.id,
        order_item_id=plan_state.get("order_item_id"),
        created_by_id=getattr(current_user, "id", None),
        target_qty=int(plan_state.get("target_qty") or 0),
        max_producible_units=int(plan_state.get("max_producible_units") or 0),
        shortage_count=len(plan_state.get("shortage_rows") or []),
        can_fulfill_plan=bool(plan_state.get("can_fulfill_plan")),
        note=note or None,
    )
    db.session.add(plan)
    db.session.commit()

    flash("Production plan saved.", "success")
    return redirect(url_for(
        "main.production_plan",
        product_id=selected_product.id,
        target_qty=int(plan_state.get("target_qty") or 1),
        order_item_id=plan_state.get("order_item_id"),
    ))


@main_bp.route("/inventory/production-plans")
@login_required
def production_plan_history():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)
    selected_product_id = request.args.get("product_id", type=int)
    status = (request.args.get("status") or "").strip().lower()
    q = (request.args.get("q") or "").strip()
    products = (
        Product.query
        .filter(Product.factory_id == factory_id)
        .order_by(Product.name.asc())
        .all()
        if factory_id else []
    )
    saved_plans = _query_saved_production_plans(
        factory_id=factory_id,
        limit=100,
        product_id=selected_product_id,
        status=status,
        q=q,
    )
    history_stats = {
        "total": len(saved_plans),
        "ready": sum(1 for row in saved_plans if row.get("can_fulfill_plan")),
        "blocked": sum(1 for row in saved_plans if not row.get("can_fulfill_plan")),
        "with_orders": sum(1 for row in saved_plans if row.get("order_id")),
        "with_tasks": sum(1 for row in saved_plans if (row.get("task_summary") or {}).get("has_tasks")),
        "open_tasks": sum(int((row.get("task_summary") or {}).get("open") or 0) for row in saved_plans),
    }

    return render_template(
        "inventory/production_plan_history.html",
        saved_plans=saved_plans,
        products=products,
        selected_product_id=selected_product_id,
        selected_status=status,
        search_query=q,
        history_stats=history_stats,
    )


@main_bp.route("/inventory/production-plans/<int:plan_id>")
@login_required
def production_plan_detail(plan_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=plan.product_id,
        target_qty=int(plan.target_qty or 0),
        order_item_id=plan.order_item_id,
    )
    _sync_production_plan_purchase_tasks(factory_id=factory_id, plan_id=plan.id, plan_state=plan_state)
    plan_state["saved_plan"] = {
        "id": plan.id,
        "note": plan.note or "",
        "created_at": plan.created_at,
        "shortage_count": int(plan.shortage_count or 0),
        "max_producible_units": int(plan.max_producible_units or 0),
        "can_fulfill_plan": bool(plan.can_fulfill_plan),
        "target_qty": int(plan.target_qty or 0),
    }
    plan_state["linked_tasks"] = _linked_production_plan_tasks(factory_id=factory_id, plan_id=plan.id)
    plan_state["linked_task_summary"] = _linked_production_plan_task_summary(factory_id=factory_id, plan_id=plan.id)
    plan_state["execution_runs"] = _production_runs_for_plan(factory_id=factory_id, plan_id=plan.id, limit=12)
    plan_state["execution_summary"] = _production_run_summary(factory_id=factory_id, plan_id=plan.id)
    plan_state["execution_progress"] = _plan_execution_progress(
        target_qty=int(plan.target_qty or 0),
        run_summary=plan_state["execution_summary"],
    )
    plan_state["shop_stock_rows"] = _shop_stock_breakdown_for_product(factory_id=factory_id, product_id=plan.product_id)
    plan_state["shop_links"] = (
        Shop.query
        .join(ShopFactoryLink, ShopFactoryLink.shop_id == Shop.id)
        .filter(ShopFactoryLink.factory_id == factory_id)
        .order_by(Shop.name.asc())
        .all()
        if factory_id else []
    )
    default_order_shop = _find_default_linked_shop(factory_id) if plan.order_item_id else None
    plan_state["shop_handoff_links"] = {
        "transfer_href": url_for("shop.transfer_to_shop"),
        "pending_orders_href": url_for("shop.factory_pending_orders"),
        "ready_orders_href": url_for("shop.list_shop_orders", status="ready"),
        "all_orders_href": url_for("shop.list_shop_orders"),
        "movement_history_href": url_for("shop.movement_history", product_id=plan.product_id),
        "order_history_href": url_for("shop.history_by_order", order_id=plan.order_item.order_id) if getattr(plan, "order_item", None) and getattr(plan.order_item, "order_id", None) else None,
        "default_order_shop_id": getattr(default_order_shop, "id", None),
        "default_order_shop_name": getattr(default_order_shop, "name", None),
    }
    plan_state["purchase_summary"] = {
        "items_to_buy": len(plan_state.get("purchase_rows") or []),
        "blocked_items": len(plan_state.get("blocking_rows") or []),
        "total_suggested_buy": sum(float(row.get("suggested_purchase_qty") or 0) for row in (plan_state.get("purchase_rows") or [])),
    }

    return render_template(
        "inventory/production_plan_detail.html",
        **plan_state,
    )


@main_bp.route("/inventory/production-runs/<int:run_id>")
@login_required
def production_run_detail(run_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    run_detail = _production_run_detail(factory_id=factory_id, run_id=run_id)
    if not run_detail:
        return abort(404)

    return render_template(
        "inventory/production_run_detail.html",
        run=run_detail,
    )


@main_bp.route("/inventory/production-runs/<int:run_id>/rollback", methods=["POST"])
@login_required
def production_run_rollback(run_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    run = (
        Production.query
        .join(Product, Product.id == Production.product_id)
        .filter(
            Production.id == run_id,
            Product.factory_id == factory_id,
        )
        .first_or_404()
    )

    rollback_reason = (request.form.get("rollback_reason") or "").strip()
    if not rollback_reason:
        flash("Rollback reason is required.", "warning")
        return redirect(url_for("main.production_run_detail", run_id=run_id))

    product = run.product
    if not product:
        flash("Run product could not be found.", "danger")
        return redirect(url_for("main.production_run_detail", run_id=run_id))

    run_qty = int(run.quantity or 0)
    current_stock = int(product.quantity or 0)
    if current_stock < run_qty:
        flash("Rollback is blocked because factory stock is already below the run quantity. Some of this production likely already left the factory.", "danger")
        return redirect(url_for("main.production_run_detail", run_id=run_id))

    plan_id = run.production_plan_id
    rollback_note = f"Rollback run #{run.id}: {rollback_reason}"

    try:
        product.quantity = current_stock - run_qty

        for consumption in list(run.consumptions or []):
            fabric = consumption.fabric
            if fabric:
                fabric.quantity = float(fabric.quantity or 0) + float(consumption.used_amount or 0)

        if plan_id:
            plan = ProductionPlan.query.filter(
                ProductionPlan.id == plan_id,
                ProductionPlan.factory_id == factory_id,
            ).first()
            if plan:
                existing_note = (plan.note or "").strip()
                merged_note = rollback_note if not existing_note else f"{existing_note} | {rollback_note}"
                plan.note = merged_note[:255]

        db.session.delete(run)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f"Rollback failed: {exc}", "danger")
        return redirect(url_for("main.production_run_detail", run_id=run_id))

    if plan_id:
        plan = ProductionPlan.query.filter_by(id=plan_id, factory_id=factory_id).first()
        refreshed_plan_state = _build_production_plan_state(
            factory_id=factory_id,
            selected_product_id=product.id,
            target_qty=int(plan.target_qty or 0) if plan else 0,
            order_item_id=plan.order_item_id if plan else None,
        )
        _sync_production_plan_purchase_tasks(
            factory_id=factory_id,
            plan_id=plan_id,
            plan_state=refreshed_plan_state,
        )
        flash("Production run rolled back and saved-plan progress was refreshed.", "success")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    flash("Production run rolled back.", "success")
    return redirect(url_for("main.production_plan_history"))


@main_bp.route("/inventory/production-plans/<int:plan_id>/note", methods=["POST"])
@login_required
def production_plan_update_note(plan_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )
    plan.note = (request.form.get("note") or "").strip()[:255] or None
    db.session.commit()
    flash("Plan note updated.", "success")
    return redirect(url_for("main.production_plan_history"))


@main_bp.route("/inventory/production-plans/<int:plan_id>/delete", methods=["POST"])
@login_required
def production_plan_delete(plan_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )
    db.session.delete(plan)
    db.session.commit()
    flash("Saved plan removed.", "success")
    return redirect(url_for("main.production_plan_history"))


@main_bp.route("/inventory/production-plans/<int:plan_id>/task", methods=["POST"])
@login_required
def production_plan_create_task(plan_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    if plan.can_fulfill_plan:
        flash("This saved plan is already ready, so no blocker task was created.", "info")
        return redirect(url_for("main.production_plan_history"))

    existing_task = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "production_plan_blocked",
            OperationalTask.source_id == plan.id,
            OperationalTask.status.in_(("open", "in_progress")),
        )
        .first()
    )
    if existing_task:
        flash("A follow-up task already exists for this blocked plan.", "info")
        return redirect(url_for("main.command_center"))

    title = f"Resolve blocked production plan for {plan.product.name if plan.product else f'Product #{plan.product_id}'}"
    description = (
        f"Target {int(plan.target_qty or 0)} units. "
        f"Current plan shows {int(plan.shortage_count or 0)} material shortage(s)."
    )
    task = OperationalTask(
        factory_id=factory_id,
        created_by_id=getattr(current_user, "id", None),
        task_type="production_plan_followup",
        source_type="production_plan_blocked",
        source_id=plan.id,
        title=title[:160],
        description=description[:255],
        action_url=url_for(
            "main.production_plan",
            product_id=plan.product_id,
            target_qty=plan.target_qty,
            order_item_id=plan.order_item_id,
        ),
        target_role="manager",
        priority="high" if int(plan.shortage_count or 0) >= 2 else "medium",
        status="open",
        due_date=_default_operational_task_due_date("high" if int(plan.shortage_count or 0) >= 2 else "medium"),
        is_system_generated=False,
    )
    db.session.add(task)
    db.session.commit()

    flash("Follow-up task created for this blocked plan.", "success")
    return redirect(url_for("main.command_center"))


@main_bp.route("/inventory/production-plans/<int:plan_id>/purchase-task", methods=["POST"])
@login_required
def production_plan_create_purchase_task(plan_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=plan.product_id,
        target_qty=int(plan.target_qty or 0),
        order_item_id=plan.order_item_id,
    )
    purchase_rows = plan_state.get("purchase_rows") or []

    if not purchase_rows:
        flash("This saved plan does not need a purchase follow-up right now.", "info")
        return redirect(url_for("main.production_plan_detail", plan_id=plan.id))

    existing_task = (
        OperationalTask.query
        .filter(
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_type == "production_plan_purchase",
            OperationalTask.source_id == plan.id,
            OperationalTask.status.in_(("open", "in_progress")),
        )
        .first()
    )
    if existing_task:
        flash("A purchase follow-up task already exists for this saved plan.", "info")
        return redirect(url_for("main.command_center"))

    top_items = []
    for row in purchase_rows[:3]:
        material_name = str(row.get("material_name") or "").strip()
        qty = float(row.get("suggested_purchase_qty") or 0)
        unit = str(row.get("unit") or "").strip()
        top_items.append(f"{material_name}: {qty:.2f} {unit}".strip())

    description = f"Buy for target {int(plan.target_qty or 0)} units."
    if top_items:
        description = f"{description} Priority items: {', '.join(top_items)}."

    task = OperationalTask(
        factory_id=factory_id,
        created_by_id=getattr(current_user, "id", None),
        task_type="purchase_followup",
        source_type="production_plan_purchase",
        source_id=plan.id,
        title=f"Buy materials for {plan.product.name if plan.product else f'Product #{plan.product_id}'}"[:160],
        description=description[:255],
        action_url=url_for("main.production_plan_detail", plan_id=plan.id),
        target_role="manager",
        priority="high" if len(purchase_rows) >= 2 else "medium",
        status="open",
        due_date=_default_operational_task_due_date("high" if len(purchase_rows) >= 2 else "medium"),
        is_system_generated=False,
    )
    db.session.add(task)
    db.session.commit()

    flash("Purchase follow-up task created from this saved plan.", "success")
    return redirect(url_for("main.command_center"))


@main_bp.route("/inventory/production-plans/<int:plan_id>/tasks/<int:task_id>/status", methods=["POST"])
@login_required
def production_plan_update_task_status(plan_id: int, task_id: int):
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    task = (
        OperationalTask.query
        .filter(
            OperationalTask.id == task_id,
            OperationalTask.factory_id == factory_id,
            OperationalTask.source_id == plan_id,
            OperationalTask.source_type.in_(("production_plan_blocked", "production_plan_purchase")),
        )
        .first_or_404()
    )

    if not _can_update_operational_task(task, workspace):
        flash("You do not have permission to update this task.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    next_status = (request.form.get("next_status") or "").strip().lower()
    allowed_statuses = {"open", "in_progress", "done", "dismissed"}
    if next_status not in allowed_statuses:
        flash("Invalid task status requested.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    task.status = next_status
    task.updated_at = datetime.utcnow()
    if next_status in {"done", "dismissed"}:
        task.closed_at = datetime.utcnow()
        task.closed_by_id = current_user.id
    else:
        task.closed_at = None
        task.closed_by_id = None

    db.session.commit()
    flash("Linked task updated.", "success")
    return redirect(url_for("main.production_plan_detail", plan_id=plan_id))


@main_bp.route("/inventory/production-plans/<int:plan_id>/receive-material", methods=["POST"])
@login_required
def production_plan_receive_material(plan_id: int):
    from app.models import Fabric

    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    fabric_id = request.form.get("fabric_id", type=int)
    received_qty = request.form.get("received_qty", type=float)
    if not fabric_id:
        flash("Material not found for this receive action.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))
    if received_qty is None or received_qty <= 0:
        flash("Enter a quantity greater than zero.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    fabric = (
        Fabric.query
        .filter(Fabric.id == fabric_id, Fabric.factory_id == factory_id)
        .first()
    )
    if not fabric:
        flash("Selected material is not available in this workspace.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    fabric.quantity = float(fabric.quantity or 0) + float(received_qty)
    db.session.commit()

    refreshed_plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=plan.product_id,
        target_qty=int(plan.target_qty or 0),
        order_item_id=plan.order_item_id,
    )
    _sync_production_plan_purchase_tasks(
        factory_id=factory_id,
        plan_id=plan_id,
        plan_state=refreshed_plan_state,
    )

    flash(f"Received {received_qty:.2f} {fabric.unit or ''} for {fabric.name}.", "success")
    return redirect(url_for("main.production_plan_detail", plan_id=plan_id))


@main_bp.route("/inventory/production-plans/<int:plan_id>/execute", methods=["POST"])
@login_required
def production_plan_execute(plan_id: int):
    from app.models import Fabric, ProductComposition, FabricConsumption

    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    execute_qty = request.form.get("execute_qty", type=int)
    if execute_qty is None or execute_qty <= 0:
        flash("Execution quantity must be greater than zero.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=plan.product_id,
        target_qty=execute_qty,
        order_item_id=plan.order_item_id,
    )

    selected_product = plan_state.get("selected_product")
    if not selected_product:
        flash("Product not found for this production plan.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    plan_rows = plan_state.get("plan_rows") or []
    if not plan_rows:
        flash("No composition rows are available for this plan, so execution could not start.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    if plan_state.get("blocking_rows"):
        flash("This plan is still blocked. Receive or buy the missing materials before executing production.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    composition_items = (
        ProductComposition.query
        .join(Fabric, Fabric.id == ProductComposition.fabric_id)
        .filter(ProductComposition.product_id == selected_product.id, Fabric.factory_id == factory_id)
        .all()
    )
    if not composition_items:
        flash("This product has no saved composition to consume.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    production_note = f"Executed from saved plan #{plan.id}"
    if plan.order_item_id:
        production_note += f" for order item #{plan.order_item_id}"

    try:
        production = Production(
            product_id=selected_product.id,
            production_plan_id=plan.id,
            date=date.today(),
            quantity=int(execute_qty),
            note=production_note[:255],
        )
        db.session.add(production)
        db.session.flush()

        selected_product.quantity = int(selected_product.quantity or 0) + int(execute_qty)

        for row in plan_rows:
            fabric_id = int(row.get("fabric_id") or 0)
            total_required = float(row.get("total_required") or 0)
            if fabric_id <= 0 or total_required <= 0:
                continue

            fabric = Fabric.query.filter_by(id=fabric_id, factory_id=factory_id).first()
            if not fabric:
                raise ValueError("A material used in this plan is no longer available.")

            current_qty = float(fabric.quantity or 0)
            if current_qty < total_required:
                raise ValueError(f"Not enough stock for {fabric.name}.")

            fabric.quantity = current_qty - total_required
            db.session.add(
                FabricConsumption(
                    factory_id=factory_id,
                    fabric_id=fabric.id,
                    production_id=production.id,
                    used_amount=total_required,
                )
            )

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f"Production execution failed: {exc}", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    refreshed_plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=plan.product_id,
        target_qty=int(plan.target_qty or 0),
        order_item_id=plan.order_item_id,
    )
    plan.max_producible_units = int(refreshed_plan_state.get("max_producible_units") or 0)
    plan.shortage_count = len(refreshed_plan_state.get("shortage_rows") or [])
    plan.can_fulfill_plan = bool(refreshed_plan_state.get("can_fulfill_plan"))
    db.session.commit()

    flash("Production run executed and materials were consumed.", "success")
    return redirect(url_for("main.production_plan_detail", plan_id=plan_id))


@main_bp.route("/inventory/production-plans/<int:plan_id>/transfer-to-shop", methods=["POST"])
@login_required
def production_plan_transfer_to_shop(plan_id: int):
    from ..services.shop_service import ShopService

    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    shop_id = request.form.get("shop_id", type=int)
    transfer_qty = request.form.get("transfer_qty", type=int)

    if not shop_id:
        flash("Choose a shop before transferring stock.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))
    if transfer_qty is None or transfer_qty <= 0:
        flash("Transfer quantity must be greater than zero.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    selected_product = (
        Product.query
        .filter(Product.id == plan.product_id, Product.factory_id == factory_id)
        .first()
    )
    if not selected_product:
        flash("Product not found for this transfer.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    if int(selected_product.quantity or 0) < int(transfer_qty):
        flash("Not enough finished product stock in the factory for this transfer.", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    try:
        ShopService().transfer_factory_to_shop(
            product_id=selected_product.id,
            quantity=int(transfer_qty),
            factory_id=factory_id,
            shop_id=shop_id,
            created_by=current_user,
        )
    except Exception as exc:
        flash(f"Shop transfer failed: {exc}", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    flash("Finished goods transferred to shop stock.", "success")
    return redirect(url_for("main.production_plan_detail", plan_id=plan_id))


@main_bp.route("/inventory/production-plans/<int:plan_id>/ship-order", methods=["POST"])
@login_required
def production_plan_ship_order(plan_id: int):
    from ..services.shop_service import ShopService

    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    plan = (
        ProductionPlan.query
        .join(Product, Product.id == ProductionPlan.product_id)
        .filter(ProductionPlan.id == plan_id, ProductionPlan.factory_id == factory_id)
        .first_or_404()
    )

    if not plan.order_item_id:
        flash("This saved plan is not linked to a shop order item.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    ship_qty = request.form.get("ship_qty", type=int)
    if ship_qty is None or ship_qty <= 0:
        flash("Ship quantity must be greater than zero.", "warning")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    try:
        ShopService().ship_order_item_to_shop(
            item_id=plan.order_item_id,
            ship_qty=int(ship_qty),
            factory_id=factory_id,
            created_by=current_user,
        )
    except Exception as exc:
        flash(f"Order shipment failed: {exc}", "danger")
        return redirect(url_for("main.production_plan_detail", plan_id=plan_id))

    flash("Order quantity shipped to shop stock.", "success")
    return redirect(url_for("main.production_plan_detail", plan_id=plan_id))


@main_bp.route("/inventory/production-plan/xlsx")
@login_required
def production_plan_xlsx():
    workspace = _resolve_current_workspace()
    factory_id = getattr(workspace, "id", None) if workspace else getattr(current_user, "factory_id", None)

    selected_product_id = request.args.get("product_id", type=int)
    target_qty = request.args.get("target_qty", type=int)
    order_item_id = request.args.get("order_item_id", type=int)

    plan_state = _build_production_plan_state(
        factory_id=factory_id,
        selected_product_id=selected_product_id,
        target_qty=target_qty,
        order_item_id=order_item_id,
    )

    if not plan_state.get("selected_product"):
        flash("Choose a product first to export a production plan.", "warning")
        return redirect(url_for("main.production_plan"))

    workbook_buffer = _build_production_plan_xlsx(plan_state)
    workspace_name = getattr(workspace, "name", None) or getattr(workspace, "title", None)
    return send_file(
        workbook_buffer,
        as_attachment=True,
        download_name=_summary_export_filename(workspace_name, "production_plan", "xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
@main_bp.route("/inventory/low-stock")
@login_required
def low_stock():
    from app.models import Fabric, Product, ShopStock, Factory

    LOW_MATERIAL_THRESHOLD = 5
    LOW_SHOP_THRESHOLD = 5

    all_materials = (
        Fabric.query
        .order_by(Fabric.quantity.asc(), Fabric.name.asc())
        .all()
    )

    low_materials = []
    for item in all_materials:
        threshold = float(getattr(item, "min_stock_quantity", 0) or 0) or float(LOW_MATERIAL_THRESHOLD)
        qty = float(getattr(item, "quantity", 0) or 0)
        if qty < threshold:
            low_materials.append(item)

    low_shop_rows = (
        ShopStock.query
        .join(Product, Product.id == ShopStock.product_id)
        .outerjoin(Factory, Factory.id == ShopStock.source_factory_id)
        .filter(ShopStock.quantity > 0, ShopStock.quantity < LOW_SHOP_THRESHOLD)
        .order_by(ShopStock.quantity.asc(), Product.name.asc())
        .all()
    )

    zero_shop_rows = (
        ShopStock.query
        .join(Product, Product.id == ShopStock.product_id)
        .outerjoin(Factory, Factory.id == ShopStock.source_factory_id)
        .filter(ShopStock.quantity == 0)
        .order_by(Product.name.asc())
        .all()
    )

    low_shop_items = []
    for row in low_shop_rows:
        low_shop_items.append({
            "name": row.product.name if row.product else f"Product #{row.product_id}",
            "qty": row.quantity or 0,
            "source_factory_name": row.source_factory.name if getattr(row, "source_factory", None) else None,
            "image_path": row.product.image_path if row.product else None,
        })

    zero_shop_items = []
    for row in zero_shop_rows:
        zero_shop_items.append({
            "name": row.product.name if row.product else f"Product #{row.product_id}",
            "qty": 0,
            "source_factory_name": row.source_factory.name if getattr(row, "source_factory", None) else None,
        })

    return render_template(
        "inventory/low_stock.html",
        low_materials=low_materials,
        low_shop_items=low_shop_items,
        zero_shop_items=zero_shop_items,
    )
@main_bp.route("/inventory/composition", methods=["GET", "POST"])
@login_required
def composition():
    from app.models import Product, Fabric, ProductComposition
    from app import db
    factory_id = getattr(current_user, "factory_id", None)

    if request.method == "POST":
        action = (request.form.get("action") or "save").strip()

        if action == "delete":
            composition_id = request.form.get("composition_id", type=int)
            if not composition_id:
                flash("Composition item id is required.", "warning")
                return redirect(url_for("main.composition"))

            row = (
                ProductComposition.query
                .join(Product, Product.id == ProductComposition.product_id)
                .filter(
                    ProductComposition.id == composition_id,
                    Product.factory_id == factory_id,
                )
                .first()
            )
            if not row:
                flash("Composition item not found.", "warning")
                return redirect(url_for("main.composition"))

            db.session.delete(row)
            db.session.commit()
            flash("Composition item deleted.", "success")
            return redirect(url_for("main.composition"))

        composition_id = request.form.get("composition_id", type=int)
        product_id = request.form.get("product_id", type=int)
        fabric_id = request.form.get("fabric_id", type=int)
        quantity_required = request.form.get("quantity_required", type=float)
        unit = (request.form.get("unit") or "").strip() or "m"
        note = (request.form.get("note") or "").strip() or None

        if not product_id or not fabric_id or quantity_required is None:
            flash("Please fill all required composition fields.", "warning")
            return redirect(url_for("main.composition"))

        product = Product.query.filter_by(id=product_id, factory_id=factory_id).first()
        fabric = Fabric.query.filter_by(id=fabric_id, factory_id=factory_id).first()
        if not product or not fabric:
            flash("Selected product or material is not available in this workspace.", "warning")
            return redirect(url_for("main.composition"))

        duplicate = (
            ProductComposition.query
            .join(Product, Product.id == ProductComposition.product_id)
            .filter(
                Product.factory_id == factory_id,
                ProductComposition.product_id == product_id,
                ProductComposition.fabric_id == fabric_id,
            )
            .first()
        )

        editing_row = None
        if composition_id:
            editing_row = (
                ProductComposition.query
                .join(Product, Product.id == ProductComposition.product_id)
                .filter(
                    ProductComposition.id == composition_id,
                    Product.factory_id == factory_id,
                )
                .first()
            )
            if not editing_row:
                flash("Composition item not found.", "warning")
                return redirect(url_for("main.composition"))

        if duplicate and (not editing_row or duplicate.id != editing_row.id):
            duplicate.quantity_required = quantity_required
            duplicate.unit = unit
            duplicate.note = note
            flash("Composition item updated.", "success")
        elif editing_row:
            editing_row.product_id = product_id
            editing_row.fabric_id = fabric_id
            editing_row.quantity_required = quantity_required
            editing_row.unit = unit
            editing_row.note = note
            flash("Composition item updated.", "success")
        else:
            db.session.add(ProductComposition(
                product_id=product_id,
                fabric_id=fabric_id,
                quantity_required=quantity_required,
                unit=unit,
                note=note,
            ))
            flash("Composition item added.", "success")

        db.session.commit()
        return redirect(url_for("main.composition"))

    selected_product_id = request.args.get("product_id", type=int)
    selected_material_type = (request.args.get("material_type") or "").strip().lower()
    q = (request.args.get("q") or "").strip()
    edit_id = request.args.get("edit", type=int)

    products = (
        Product.query
        .filter(Product.factory_id == factory_id)
        .order_by(Product.name.asc())
        .all()
    )
    fabrics = (
        Fabric.query
        .filter(Fabric.factory_id == factory_id)
        .order_by(Fabric.material_type.asc(), Fabric.name.asc())
        .all()
    )
    composition_query = (
        ProductComposition.query
        .join(Product, Product.id == ProductComposition.product_id)
        .join(Fabric, Fabric.id == ProductComposition.fabric_id)
        .filter(Product.factory_id == factory_id)
    )

    if selected_product_id:
        composition_query = composition_query.filter(ProductComposition.product_id == selected_product_id)
    if selected_material_type:
        composition_query = composition_query.filter(db.func.lower(Fabric.material_type) == selected_material_type)
    if q:
        pattern = f"%{q.lower()}%"
        composition_query = composition_query.filter(
            db.or_(
                db.func.lower(Product.name).like(pattern),
                db.func.lower(Fabric.name).like(pattern),
                db.func.lower(db.func.coalesce(Fabric.category, "")).like(pattern),
                db.func.lower(db.func.coalesce(ProductComposition.note, "")).like(pattern),
            )
        )

    composition_items = composition_query.order_by(Product.name.asc(), Fabric.material_type.asc(), Fabric.name.asc()).all()

    grouped_items = []
    grouped_map = {}

    for item in composition_items:
        product_id = item.product_id
        if product_id not in grouped_map:
            grouped_map[product_id] = {
                "product": item.product,
                "items": [],
            }
            grouped_items.append(grouped_map[product_id])

        grouped_map[product_id]["items"].append(item)

    edit_item = None
    if edit_id:
        edit_item = (
            ProductComposition.query
            .join(Product, Product.id == ProductComposition.product_id)
            .join(Fabric, Fabric.id == ProductComposition.fabric_id)
            .filter(
                ProductComposition.id == edit_id,
                Product.factory_id == factory_id,
            )
            .first()
        )

    material_types = sorted({
        str(getattr(fabric, "material_type", None) or "fabric").strip().lower()
        for fabric in fabrics
        if getattr(fabric, "material_type", None) or True
    })

    return render_template(
        "inventory/composition.html",
        products=products,
        fabrics=fabrics,
        composition_items=composition_items,
        grouped_items=grouped_items,
        selected_product_id=selected_product_id,
        selected_material_type=selected_material_type,
        q=q,
        edit_item=edit_item,
        material_types=material_types,
    )
def _build_more_cards():
    is_admin = current_user.role == "admin"
    is_superadmin = bool(getattr(current_user, "is_superadmin", False))
    factory_id = getattr(current_user, "factory_id", None)

    managed_users_count = 0
    managed_shops_count = 0
    managed_factories_count = 0

    if is_superadmin:
        managed_users_count = User.query.count()
        managed_shops_count = Shop.query.count()
        managed_factories_count = Factory.query.count()
    elif is_admin and factory_id:
        managed_users_count = User.query.filter_by(factory_id=factory_id).count()
        managed_shops_count = (
            db.session.query(ShopFactoryLink.shop_id)
            .filter(ShopFactoryLink.factory_id == factory_id)
            .distinct()
            .count()
        )
        managed_factories_count = 1

    primary_cards = [
        {
            "title": "Business summary",
            "subtitle": "Owner-facing health, branch, product, and customer intelligence",
            "href": url_for("main.business_summary"),
            "status": "live",
        },
        {
            "title": "Profile",
            "subtitle": "Account settings, password, Telegram link",
            "href": url_for("main.profile"),
            "status": "live",
        },
        {
            "title": "Orders",
            "subtitle": "Customer and shop orders",
            "href": url_for("shop.list_shop_orders"),
            "status": "live",
        },
        {
            "title": "Activity",
            "subtitle": "Movement history and stock actions",
            "href": url_for("shop.movement_history"),
            "status": "live",
        },
        {
            "title": "Reports",
            "subtitle": "Business summaries and reporting pages",
            "href": url_for("manager_report.manager_report"),
            "status": "live",
        },
        {
            "title": "Export shop report",
            "subtitle": "Download current shop XLSX report",
            "href": url_for("shop.export_shop"),
            "status": "live",
        },
        {
            "title": "Cash",
            "subtitle": "Cash and finance records",
            "href": url_for("cash.list_cash"),
            "status": "live",
        },
    ]

    admin_cards = [
        {
            "title": "Team",
            "subtitle": "Users and access control",
            "href": url_for("auth.list_users") if is_admin else None,
            "status": "live" if is_admin else "locked",
        },
        {
            "title": "Branches",
            "subtitle": "Shops and linked factories",
            "href": url_for("auth.list_shops") if is_admin else None,
            "status": "live" if is_admin else "locked",
        },
        {
            "title": "Factories",
            "subtitle": "Manage factory workspaces",
            "href": url_for("auth.list_factories") if is_superadmin else None,
            "status": "live" if is_superadmin else "locked",
        },
        {
            "title": "Create user",
            "subtitle": "Add manager, shop user, or accountant",
            "href": url_for("auth.create_user") if is_admin else None,
            "status": "live" if is_admin else "locked",
        },
        {
            "title": "Create shop",
            "subtitle": "Add and connect a new shop",
            "href": url_for("auth.create_shop") if is_admin else None,
            "status": "live" if is_admin else "locked",
        },
        {
            "title": "Create factory",
            "subtitle": "Register a new factory workspace",
            "href": url_for("auth.create_factory") if is_superadmin else None,
            "status": "live" if is_superadmin else "locked",
        },
    ]

    language_cards = [
        {
            "title": "English",
            "subtitle": "Switch interface language to EN",
            "href": url_for("switch_language", lang_code="en"),
            "status": "active" if session.get("lang_code", "ru") == "en" else "live",
        },
        {
            "title": "Русский",
            "subtitle": "Переключить язык интерфейса на RU",
            "href": url_for("switch_language", lang_code="ru"),
            "status": "active" if session.get("lang_code", "ru") == "ru" else "live",
        },
        {
            "title": "O‘zbek",
            "subtitle": "Interfeys tilini UZ ga o‘zgartirish",
            "href": url_for("switch_language", lang_code="uz"),
            "status": "active" if session.get("lang_code", "ru") == "uz" else "live",
        },
    ]

    return {
        "primary_cards": primary_cards,
        "admin_cards": admin_cards,
        "language_cards": language_cards,
        "managed_users_count": managed_users_count,
        "managed_shops_count": managed_shops_count,
        "managed_factories_count": managed_factories_count,
    }
REPORT_DEFINITIONS = {
    "factory_stock": {"title": "Factory Stock Report", "subtitle": "Current product quantities, stock value, and factory-side health.", "icon": "FS"},
    "shop_stock": {"title": "Shop Stock Report", "subtitle": "Branch-facing inventory, low-stock pressure, and sell-side coverage.", "icon": "SS"},
    "sales": {"title": "Sales Report", "subtitle": "Recorded sales performance across recent activity windows.", "icon": "SL"},
    "orders": {"title": "Orders Report", "subtitle": "Pending, ready, completed, and cancelled order flow.", "icon": "OR"},
    "cutting_orders": {"title": "Cutting Orders", "subtitle": "Track sets cut, estimated material cost, and order status.", "icon": "CO"},
    "movements": {"title": "Movement History Report", "subtitle": "Stock transfers, sales-linked movements, and adjustments.", "icon": "MV"},
    "low_stock": {"title": "Low Stock Report", "subtitle": "Priority products that need replenishment or urgent attention.", "icon": "LS"},
    "cash": {"title": "Cash / Finance Summary", "subtitle": "Cash records, inflow, outflow, and current balance.", "icon": "CA"},
    "production": {"title": "Production Report", "subtitle": "Production output, active models, and recent execution rows.", "icon": "PR"},
    "period_summary": {"title": "Annual / Period Summary", "subtitle": "A compact owner-level view of the current business period.", "icon": "YR"},
}


def _report_parse_date_range(default_days: int | None = None):
    date_fmt = "%Y-%m-%d"
    date_from_str = (request.args.get("from") or "").strip()
    date_to_str = (request.args.get("to") or "").strip()
    date_from = None
    date_to = None

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, date_fmt).date()
        except ValueError:
            date_from = None
            date_from_str = ""

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, date_fmt).date()
        except ValueError:
            date_to = None
            date_to_str = ""

    if default_days and not date_from and not date_to:
        date_to = date.today()
        date_from = date_to - timedelta(days=max(default_days - 1, 0))
        date_from_str = date_from.strftime(date_fmt)
        date_to_str = date_to.strftime(date_fmt)

    return {
        "from": date_from,
        "to": date_to,
        "from_str": date_from_str,
        "to_str": date_to_str,
    }


def _report_money(amount, currency="UZS") -> str:
    currency_code = str(currency or "UZS").upper()
    value = float(amount or 0)
    if currency_code == "USD":
        return f"{value:,.2f} USD"
    return f"{value:,.0f} UZS"


def _report_tone_from_status(status_value: str | None) -> str:
    status = str(status_value or "").strip().lower()
    if status in {"out", "cancelled", "danger"}:
        return "danger"
    if status in {"low", "pending", "warning"}:
        return "warn"
    if status in {"ready", "completed", "healthy", "positive", "income", "produced"}:
        return "up"
    return "neutral"


def _report_workspace_or_redirect():
    workspace = _resolve_current_workspace()
    if not workspace:
        flash("No workspace is linked to this account yet.", "danger")
        return None
    return workspace


def _report_summary_card(label: str, value: str, sub: str = "", tone: str = "neutral"):
    return {"label": label, "value": value, "sub": sub, "tone": tone}


def _report_base_state(report_key: str):
    report_meta = REPORT_DEFINITIONS[report_key]
    return {
        "report_key": report_key,
        "report_title": report_meta["title"],
        "report_subtitle": report_meta["subtitle"],
        "report_icon": report_meta["icon"],
        "table_columns": [],
        "rows": [],
        "summary_cards": [],
        "filter_fields": [],
        "empty_title": "No report rows yet",
        "empty_copy": "Adjust filters or wait for new activity to appear.",
    }


def _report_product_secondary(product) -> list[str]:
    return [f"SKU #{getattr(product, 'id', '-')}", getattr(product, "category", None) or "No category"]


def _report_export_lookup(row: dict, label: str):
    key = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(label or ""))
    while "__" in key:
        key = key.replace("__", "_")
    key = key.strip("_")
    return row.get(key, row.get(label, ""))


def _build_factory_stock_report(factory_id: int):
    state = _report_base_state("factory_stock")
    q = (request.args.get("q") or "").strip()
    selected_category = (request.args.get("category") or "").strip()
    status_filter = (request.args.get("status") or "").strip().lower()

    query = (
        db.session.query(
            Product,
            func.coalesce(func.sum(ShopStock.quantity), 0).label("qty_shop"),
        )
        .outerjoin(
            ShopStock,
            (ShopStock.product_id == Product.id)
            & (ShopStock.source_factory_id == Product.factory_id),
        )
        .filter(Product.factory_id == factory_id)
        .group_by(Product.id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(or_(Product.name.ilike(like), Product.category.ilike(like)))

    if selected_category:
        query = query.filter(Product.category == selected_category)

    rows = query.order_by(Product.quantity.desc(), Product.name.asc()).all()
    categories = sorted({getattr(product, "category", None) for product, _ in rows if getattr(product, "category", None)})

    filtered_rows = []
    total_qty = 0
    total_value = 0.0
    low_count = 0
    for product, qty_shop in rows:
        qty_factory = int(getattr(product, "quantity", 0) or 0)
        if status_filter == "out" and qty_factory != 0:
            continue
        if status_filter == "low" and not (0 < qty_factory <= 5):
            continue
        if status_filter == "healthy" and qty_factory <= 5:
            continue

        stock_value = qty_factory * float(getattr(product, "cost_price_per_item", 0) or 0)
        total_qty += qty_factory
        total_value += stock_value
        if qty_factory <= 5:
            low_count += 1

        status_label = "Out" if qty_factory <= 0 else ("Low" if qty_factory <= 5 else "Healthy")
        filtered_rows.append(
            {
                "primary": product.name,
                "secondary_lines": _report_product_secondary(product),
                "image_path": getattr(product, "image_path", None),
                "cells": [
                    str(qty_factory),
                    _report_money(stock_value, getattr(product, "currency", "UZS") or "UZS"),
                    status_label,
                ],
                "mobile_value": str(qty_factory),
                "mobile_subvalue": _report_money(stock_value, getattr(product, "currency", "UZS") or "UZS"),
                "badge_label": status_label,
                "badge_tone": _report_tone_from_status(status_label),
                "export": {
                    "sku": getattr(product, "id", None),
                    "name": product.name,
                    "category": getattr(product, "category", None) or "",
                    "factory_qty": qty_factory,
                    "shop_qty": int(qty_shop or 0),
                    "stock_value": round(stock_value, 2),
                    "currency": getattr(product, "currency", "UZS") or "UZS",
                    "status": status_label,
                },
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Factory quantity", str(total_qty), "Units currently in factory", "info"),
        _report_summary_card("Models", str(len(filtered_rows)), "Tracked products in this report", "neutral"),
        _report_summary_card("Stock value", _report_money(total_value, "UZS"), "Based on cost price", "up"),
        _report_summary_card("Low stock", str(low_count), "Units at 5 pcs or lower", "warn" if low_count else "up"),
    ]
    state["table_columns"] = ["Qty", "Value", "Status"]
    state["rows"] = filtered_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Product or category"},
        {"name": "category", "label": "Category", "type": "select", "value": selected_category, "options": [{"value": "", "label": "All"}, *[{"value": value, "label": value} for value in categories]]},
        {"name": "status", "label": "Status", "type": "select", "value": status_filter, "options": [{"value": "", "label": "All"}, {"value": "healthy", "label": "Healthy"}, {"value": "low", "label": "Low"}, {"value": "out", "label": "Out"}]},
    ]
    state["export_columns"] = ["SKU", "Product", "Category", "Factory qty", "Shop qty", "Stock value", "Currency", "Status"]
    state["export_rows"] = [row["export"] for row in filtered_rows]
    state["empty_title"] = "No factory stock rows"
    state["empty_copy"] = "No products matched the current stock filters."
    return state


def _build_shop_stock_report(factory_id: int):
    state = _report_base_state("shop_stock")
    q = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "").strip().lower()
    shop_filter = request.args.get("shop_id", type=int)

    query = (
        ShopStock.query
        .join(Product, Product.id == ShopStock.product_id)
        .join(Shop, Shop.id == ShopStock.shop_id)
        .filter(ShopStock.source_factory_id == factory_id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(or_(Product.name.ilike(like), Product.category.ilike(like), Shop.name.ilike(like)))

    if shop_filter:
        query = query.filter(ShopStock.shop_id == shop_filter)

    rows = query.order_by(ShopStock.quantity.asc(), Product.name.asc()).all()
    shop_options = [{"value": "", "label": "All"}] + [{"value": str(shop.id), "label": shop.name} for shop in _get_workspace_shops(factory_id)]

    report_rows = []
    total_qty = 0
    total_value = 0.0
    low_count = 0
    for row in rows:
        qty = int(row.quantity or 0)
        if status_filter == "out" and qty != 0:
            continue
        if status_filter == "low" and not (0 < qty <= 5):
            continue
        if status_filter == "healthy" and qty <= 5:
            continue

        unit_value = float(getattr(row.product, "sell_price_per_item", 0) or 0)
        row_value = qty * unit_value
        total_qty += qty
        total_value += row_value
        if qty <= 5:
            low_count += 1

        status_label = "Out" if qty <= 0 else ("Low" if qty <= 5 else "Healthy")
        report_rows.append(
            {
                "primary": getattr(row.product, "name", "Product"),
                "secondary_lines": [f"SKU #{getattr(row.product, 'id', '-')}", getattr(row.shop, "name", "Shop"), getattr(row.product, "category", None) or "No category"],
                "image_path": getattr(row.product, "image_path", None),
                "cells": [getattr(row.shop, "name", "Shop"), str(qty), _report_money(row_value, getattr(row.product, "currency", "UZS") or "UZS"), status_label],
                "mobile_value": str(qty),
                "mobile_subvalue": getattr(row.shop, "name", "Shop"),
                "badge_label": status_label,
                "badge_tone": _report_tone_from_status(status_label),
                "export": {
                    "sku": getattr(row.product, "id", None),
                    "product": getattr(row.product, "name", ""),
                    "shop": getattr(row.shop, "name", ""),
                    "category": getattr(row.product, "category", None) or "",
                    "qty": qty,
                    "value": round(row_value, 2),
                    "currency": getattr(row.product, "currency", "UZS") or "UZS",
                    "status": status_label,
                },
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Shop units", str(total_qty), "Units currently placed in shops", "info"),
        _report_summary_card("SKUs", str(len(report_rows)), "Rows in this shop report", "neutral"),
        _report_summary_card("Shop value", _report_money(total_value, "UZS"), "Using sell price", "up"),
        _report_summary_card("Low stock", str(low_count), "Rows at 5 pcs or lower", "warn" if low_count else "up"),
    ]
    state["table_columns"] = ["Shop", "Qty", "Value", "Status"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Product, category, or shop"},
        {"name": "shop_id", "label": "Shop", "type": "select", "value": str(shop_filter or ""), "options": shop_options},
        {"name": "status", "label": "Status", "type": "select", "value": status_filter, "options": [{"value": "", "label": "All"}, {"value": "healthy", "label": "Healthy"}, {"value": "low", "label": "Low"}, {"value": "out", "label": "Out"}]},
    ]
    state["export_columns"] = ["SKU", "Product", "Shop", "Category", "Qty", "Value", "Currency", "Status"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No shop stock rows"
    state["empty_copy"] = "No shop stock matched the current filters."
    return state


def _build_sales_report(factory_id: int):
    state = _report_base_state("sales")
    q = (request.args.get("q") or "").strip()
    date_range = _report_parse_date_range(default_days=30)

    retail_query = Sale.query.join(Product).filter(Product.factory_id == factory_id)
    wholesale_query = (
        db.session.query(WholesaleSaleItem, WholesaleSale, Product)
        .join(WholesaleSale, WholesaleSale.id == WholesaleSaleItem.wholesale_sale_id)
        .join(Product, Product.id == WholesaleSaleItem.product_id)
        .filter(WholesaleSaleItem.source_factory_id == factory_id)
    )

    if date_range["from"]:
        retail_query = retail_query.filter(Sale.date >= date_range["from"])
        wholesale_query = wholesale_query.filter(WholesaleSale.sale_date >= date_range["from"])
    if date_range["to"]:
        retail_query = retail_query.filter(Sale.date <= date_range["to"])
        wholesale_query = wholesale_query.filter(WholesaleSale.sale_date <= date_range["to"])

    if q:
        like = f"%{q}%"
        retail_query = retail_query.filter(or_(Product.name.ilike(like), Sale.customer_name.ilike(like), Sale.customer_phone.ilike(like)))
        wholesale_query = wholesale_query.filter(or_(Product.name.ilike(like), WholesaleSale.customer_name.ilike(like), WholesaleSale.customer_phone.ilike(like)))

    report_rows = []
    sale_count = 0
    today_total = 0.0
    week_total = 0.0
    month_total = 0.0
    today_value = date.today()
    week_start = today_value - timedelta(days=6)
    month_start = today_value.replace(day=1)

    for sale in retail_query.order_by(Sale.date.desc(), Sale.id.desc()).all():
        amount = float(getattr(sale, "total_sell", 0) or 0)
        sale_date = getattr(sale, "date", None)
        sale_count += 1
        if sale_date == today_value:
            today_total += amount
        if sale_date and week_start <= sale_date <= today_value:
            week_total += amount
        if sale_date and month_start <= sale_date <= today_value:
            month_total += amount
        report_rows.append(
            {
                "primary": getattr(sale.product, "name", "Product"),
                "secondary_lines": [getattr(sale, "customer_name", None) or "Walk-in customer", sale_date.strftime("%Y-%m-%d") if sale_date else "-"],
                "image_path": getattr(sale.product, "image_path", None),
                "cells": [str(int(getattr(sale, "quantity", 0) or 0)), _report_money(amount, getattr(sale, "currency", "UZS") or "UZS"), "Retail"],
                "mobile_value": _report_money(amount, getattr(sale, "currency", "UZS") or "UZS"),
                "mobile_subvalue": sale_date.strftime("%Y-%m-%d") if sale_date else "-",
                "badge_label": "Retail",
                "badge_tone": "info",
                "export": {"date": sale_date.strftime("%Y-%m-%d") if sale_date else "", "channel": "Retail", "product": getattr(sale.product, "name", ""), "customer": getattr(sale, "customer_name", None) or "", "qty": int(getattr(sale, "quantity", 0) or 0), "total": round(amount, 2), "currency": getattr(sale, "currency", "UZS") or "UZS"},
            }
        )

    for item, wholesale_sale, product in wholesale_query.order_by(WholesaleSale.sale_date.desc(), WholesaleSale.id.desc()).all():
        amount = float(getattr(item, "line_total", 0) or 0)
        sale_date = getattr(wholesale_sale, "sale_date", None)
        sale_count += 1
        if sale_date == today_value:
            today_total += amount
        if sale_date and week_start <= sale_date <= today_value:
            week_total += amount
        if sale_date and month_start <= sale_date <= today_value:
            month_total += amount
        report_rows.append(
            {
                "primary": getattr(product, "name", "Product"),
                "secondary_lines": [getattr(wholesale_sale, "customer_name", None) or "Wholesale customer", sale_date.strftime("%Y-%m-%d") if sale_date else "-"],
                "image_path": getattr(product, "image_path", None),
                "cells": [str(int(getattr(item, "quantity", 0) or 0)), _report_money(amount, getattr(item, "currency", "UZS") or "UZS"), "Wholesale"],
                "mobile_value": _report_money(amount, getattr(item, "currency", "UZS") or "UZS"),
                "mobile_subvalue": sale_date.strftime("%Y-%m-%d") if sale_date else "-",
                "badge_label": "Wholesale",
                "badge_tone": "neutral",
                "export": {"date": sale_date.strftime("%Y-%m-%d") if sale_date else "", "channel": "Wholesale", "product": getattr(product, "name", ""), "customer": getattr(wholesale_sale, "customer_name", None) or "", "qty": int(getattr(item, "quantity", 0) or 0), "total": round(amount, 2), "currency": getattr(item, "currency", "UZS") or "UZS"},
            }
        )

    report_rows.sort(key=lambda row: (row["export"]["date"], row["primary"]), reverse=True)
    state["summary_cards"] = [
        _report_summary_card("Today sales", _report_money(today_total, "UZS"), "Current day revenue", "up"),
        _report_summary_card("7 day sales", _report_money(week_total, "UZS"), "Recent weekly revenue", "info"),
        _report_summary_card("Month sales", _report_money(month_total, "UZS"), "Month-to-date revenue", "up"),
        _report_summary_card("Sale rows", str(sale_count), "Retail + wholesale rows", "neutral"),
    ]
    state["table_columns"] = ["Qty", "Total", "Channel"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Product or customer"},
        {"name": "from", "label": "From", "type": "date", "value": date_range["from_str"]},
        {"name": "to", "label": "To", "type": "date", "value": date_range["to_str"]},
    ]
    state["export_columns"] = ["Date", "Channel", "Product", "Customer", "Qty", "Total", "Currency"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No sales in this range"
    state["empty_copy"] = "No retail or wholesale sales matched the current filters."
    return state


def _build_orders_report(factory_id: int):
    state = _report_base_state("orders")
    q = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "").strip().lower()
    query = ShopOrder.query.filter(ShopOrder.factory_id == factory_id)
    if status_filter:
        query = query.filter(ShopOrder.status == status_filter)
    rows = query.order_by(ShopOrder.created_at.desc(), ShopOrder.id.desc()).all()

    report_rows = []
    counts = {"pending": 0, "ready": 0, "completed": 0, "cancelled": 0}
    for order in rows:
        if q:
            order_text = " ".join([str(order.id), getattr(order, "customer_name", None) or "", getattr(order, "customer_phone", None) or "", " ".join(getattr(item.product, "name", "") for item in getattr(order, "items", []))]).lower()
            if q.lower() not in order_text:
                continue

        order_status = str(getattr(order, "status", "pending") or "pending").lower()
        counts[order_status] = counts.get(order_status, 0) + 1
        total_qty = sum(int(getattr(item, "qty_requested", 0) or 0) for item in getattr(order, "items", []))
        remaining_qty = sum(int(getattr(item, "qty_remaining", 0) or 0) for item in getattr(order, "items", []))
        first_item = next((item for item in getattr(order, "items", []) if getattr(item, "product", None)), None)
        item_preview = ", ".join(getattr(item.product, "name", "Item") for item in getattr(order, "items", [])[:2])
        if len(getattr(order, "items", [])) > 2:
            item_preview += f" +{len(order.items) - 2} more"
        status_label = order_status.replace("_", " ").title()
        report_rows.append(
            {
                "primary": f"Order #{order.id}",
                "secondary_lines": [getattr(order, "customer_name", None) or "No customer name", item_preview or "No items"],
                "image_path": getattr(getattr(first_item, "product", None), "image_path", None),
                "cells": [str(total_qty), str(remaining_qty), status_label],
                "mobile_value": status_label,
                "mobile_subvalue": order.created_at.strftime("%Y-%m-%d") if getattr(order, "created_at", None) else "-",
                "badge_label": status_label,
                "badge_tone": _report_tone_from_status(status_label),
                "export": {"order_id": order.id, "customer": getattr(order, "customer_name", None) or "", "status": status_label, "created_at": order.created_at.strftime("%Y-%m-%d %H:%M") if getattr(order, "created_at", None) else "", "items": item_preview, "qty_requested": total_qty, "qty_remaining": remaining_qty},
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Pending", str(counts.get("pending", 0)), "Needs action", "warn" if counts.get("pending", 0) else "up"),
        _report_summary_card("Ready", str(counts.get("ready", 0)), "Ready to move", "info"),
        _report_summary_card("Completed", str(counts.get("completed", 0)), "Closed orders", "up"),
        _report_summary_card("Cancelled", str(counts.get("cancelled", 0)), "Dropped orders", "neutral"),
    ]
    state["table_columns"] = ["Requested", "Remaining", "Status"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Order, customer, or product"},
        {"name": "status", "label": "Status", "type": "select", "value": status_filter, "options": [{"value": "", "label": "All"}, {"value": "pending", "label": "Pending"}, {"value": "ready", "label": "Ready"}, {"value": "completed", "label": "Completed"}, {"value": "cancelled", "label": "Cancelled"}]},
    ]
    state["export_columns"] = ["Order ID", "Customer", "Status", "Created at", "Items", "Qty requested", "Qty remaining"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No orders matched"
    state["empty_copy"] = "No shop orders matched the current filters."
    return state


def _build_cutting_orders_report(factory_id: int):
    state = _report_base_state("cutting_orders")
    q = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "").strip().lower()
    date_range = _report_parse_date_range(default_days=30)

    query = (
        CuttingOrder.query
        .options(
            joinedload(CuttingOrder.product),
            joinedload(CuttingOrder.created_by),
            joinedload(CuttingOrder.materials).joinedload(CuttingOrderMaterial.material),
        )
        .filter(CuttingOrder.factory_id == factory_id)
    )

    if status_filter:
        query = query.filter(func.lower(CuttingOrder.status) == status_filter)
    if date_range["from"]:
        query = query.filter(CuttingOrder.cut_date >= date_range["from"])
    if date_range["to"]:
        query = query.filter(CuttingOrder.cut_date <= date_range["to"])

    orders = query.order_by(CuttingOrder.cut_date.desc(), CuttingOrder.id.desc()).all()

    if q:
        q_lower = q.lower()
        filtered_orders = []
        for order in orders:
            search_blob = " ".join(
                [
                    str(getattr(order, "id", "") or ""),
                    getattr(getattr(order, "product", None), "name", None) or "",
                    getattr(order, "status", None) or "",
                    getattr(order, "notes", None) or "",
                    getattr(getattr(order, "created_by", None), "full_name", None) or "",
                    getattr(getattr(order, "created_by", None), "username", None) or "",
                ]
            ).lower()
            if q_lower in search_blob:
                filtered_orders.append(order)
        orders = filtered_orders

    known_statuses = ["open", "in_progress", "closed"]
    discovered_statuses = {
        str(value or "").strip().lower()
        for (value,) in db.session.query(CuttingOrder.status)
        .filter(CuttingOrder.factory_id == factory_id)
        .distinct()
        .all()
        if str(value or "").strip()
    }
    ordered_statuses = [status for status in known_statuses if status in discovered_statuses]
    ordered_statuses.extend(sorted(status for status in discovered_statuses if status not in known_statuses))

    total_orders = len(orders)
    total_sets = 0
    total_estimated_material_cost = 0.0
    status_counts = {}
    product_rollup = {}
    report_rows = []

    for order in orders:
        sets_cut = int(getattr(order, "sets_cut", 0) or 0)
        material_rows = getattr(order, "materials", None) or []
        material_cost = sum(float(getattr(material, "total_cost_snapshot", 0) or 0) for material in material_rows)
        material_lines = len(material_rows)
        total_sets += sets_cut
        total_estimated_material_cost += material_cost

        status_key = str(getattr(order, "status", None) or "open").strip().lower() or "open"
        status_counts[status_key] = status_counts.get(status_key, 0) + 1

        product = getattr(order, "product", None)
        product_id = getattr(order, "product_id", None)
        product_name = getattr(product, "name", None) or f"Product #{product_id or '-'}"
        product_rollup[product_id or product_name] = {
            "name": product_name,
            "sets": product_rollup.get(product_id or product_name, {}).get("sets", 0) + sets_cut,
        }

        cut_date = getattr(order, "cut_date", None)
        created_by = getattr(order, "created_by", None)
        created_by_label = getattr(created_by, "full_name", None) or getattr(created_by, "username", None) or "Unknown user"
        status_label = status_key.replace("_", " ").title()

        report_rows.append(
            {
                "primary": product_name,
                "secondary_lines": [
                    f"Order #{getattr(order, 'id', '-')}",
                    cut_date.strftime("%Y-%m-%d") if cut_date else "-",
                    f"{material_lines} material lines",
                    created_by_label,
                ],
                "image_path": getattr(product, "image_path", None),
                "cells": [
                    cut_date.strftime("%Y-%m-%d") if cut_date else "-",
                    str(sets_cut),
                    f"{material_cost:,.2f}",
                    status_label,
                ],
                "mobile_value": str(sets_cut),
                "mobile_subvalue": f"Est. cost {material_cost:,.2f}",
                "badge_label": status_label,
                "badge_tone": "up" if status_key == "closed" else _report_tone_from_status(status_key),
                "export": {
                    "order_id": getattr(order, "id", None),
                    "cut_date": cut_date.strftime("%Y-%m-%d") if cut_date else "",
                    "product": product_name,
                    "sets_cut": sets_cut,
                    "status": status_label,
                    "estimated_material_cost": round(material_cost, 2),
                    "material_lines": material_lines,
                    "created_by": created_by_label,
                    "notes": getattr(order, "notes", None) or "",
                },
            }
        )

    avg_sets_per_order = (total_sets / total_orders) if total_orders else 0.0
    status_breakdown = ", ".join(
        f"{status.replace('_', ' ').title()}: {status_counts[status]}"
        for status in ordered_statuses
        if status_counts.get(status)
    ) or "No status rows in this view"
    top_products = sorted(
        product_rollup.items(),
        key=lambda item: (-int(item[1]["sets"] or 0), item[1]["name"].lower()),
    )
    product_breakdown = ", ".join(
        f"{row['name']}: {row['sets']} sets"
        for _, row in top_products[:3]
    ) or "No product rows in this view"

    state["summary_cards"] = [
        _report_summary_card("Cutting orders", str(total_orders), "Orders matching the current filters", "info"),
        _report_summary_card("Sets cut", str(total_sets), "Total sets recorded in this report", "up"),
        _report_summary_card("Avg sets / order", f"{avg_sets_per_order:,.1f}", "Average sets cut per cutting order", "neutral"),
        _report_summary_card("Est. material cost", f"{total_estimated_material_cost:,.2f}", "Sum of stored material cost snapshots", "info"),
        _report_summary_card("Status breakdown", str(sum(1 for count in status_counts.values() if count)), status_breakdown, "neutral"),
        _report_summary_card("Sets by product", str(len(product_rollup)), product_breakdown, "neutral"),
    ]
    state["table_columns"] = ["Cut date", "Sets", "Est. material cost", "Status"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Order, product, note, or user"},
        {"name": "status", "label": "Status", "type": "select", "value": status_filter, "options": [{"value": "", "label": "All"}, *[{"value": status, "label": status.replace("_", " ").title()} for status in ordered_statuses]]},
        {"name": "from", "label": "From", "type": "date", "value": date_range["from_str"]},
        {"name": "to", "label": "To", "type": "date", "value": date_range["to_str"]},
    ]
    state["export_columns"] = ["Order ID", "Cut date", "Product", "Sets cut", "Status", "Estimated material cost", "Material lines", "Created by", "Notes"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No cutting orders in this view"
    state["empty_copy"] = "No cutting orders matched the current search, status, or date filters."
    return state


def _build_movements_report(factory_id: int):
    state = _report_base_state("movements")
    q = (request.args.get("q") or "").strip()
    movement_type = (request.args.get("type") or "").strip()
    date_range = _report_parse_date_range(default_days=30)
    query = StockMovement.query.join(Product).filter(Product.factory_id == factory_id)
    if q:
        query = query.filter(Product.name.ilike(f"%{q}%"))
    if movement_type:
        query = query.filter(StockMovement.movement_type == movement_type)
    if date_range["from"]:
        query = query.filter(StockMovement.timestamp >= datetime.combine(date_range["from"], datetime.min.time()))
    if date_range["to"]:
        query = query.filter(StockMovement.timestamp < datetime.combine(date_range["to"] + timedelta(days=1), datetime.min.time()))

    rows = query.order_by(StockMovement.timestamp.desc()).all()
    total_in = 0
    total_out = 0
    order_linked = 0
    report_rows = []
    for row in rows:
        qty = int(getattr(row, "qty_change", 0) or 0)
        if qty > 0:
            total_in += qty
        elif qty < 0:
            total_out += abs(qty)
        if getattr(row, "order_id", None):
            order_linked += 1
        move_label = str(getattr(row, "movement_type", "") or "movement").replace("_", " ").title()
        report_rows.append(
            {
                "primary": getattr(getattr(row, "product", None), "name", "Product"),
                "secondary_lines": [move_label, row.timestamp.strftime("%Y-%m-%d %H:%M") if getattr(row, "timestamp", None) else "-"],
                "image_path": getattr(getattr(row, "product", None), "image_path", None),
                "cells": [str(qty), _report_money(getattr(row, "total_value", 0) or 0, getattr(row, "currency", "UZS") or "UZS"), (f"Order #{row.order_id}" if getattr(row, "order_id", None) else "No order")],
                "mobile_value": str(qty),
                "mobile_subvalue": move_label,
                "badge_label": move_label,
                "badge_tone": "info",
                "export": {"timestamp": row.timestamp.strftime("%Y-%m-%d %H:%M") if getattr(row, "timestamp", None) else "", "product": getattr(getattr(row, "product", None), "name", ""), "movement_type": move_label, "qty_change": qty, "total_value": round(float(getattr(row, "total_value", 0) or 0), 2), "currency": getattr(row, "currency", "UZS") or "UZS", "order_id": getattr(row, "order_id", None) or ""},
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Movements", str(len(report_rows)), "Rows in this range", "info"),
        _report_summary_card("Inbound qty", str(total_in), "Positive movement units", "up"),
        _report_summary_card("Outbound qty", str(total_out), "Negative movement units", "warn"),
        _report_summary_card("Order linked", str(order_linked), "Rows tied to orders", "neutral"),
    ]
    state["table_columns"] = ["Qty", "Value", "Order"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Product"},
        {"name": "from", "label": "From", "type": "date", "value": date_range["from_str"]},
        {"name": "to", "label": "To", "type": "date", "value": date_range["to_str"]},
        {"name": "type", "label": "Type", "type": "select", "value": movement_type, "options": [{"value": "", "label": "All"}, {"value": "factory_to_shop", "label": "Factory to shop"}, {"value": "factory_to_shop_for_order", "label": "Factory to shop for order"}, {"value": "shop_sale", "label": "Shop sale"}, {"value": "adjustment", "label": "Adjustment"}]},
    ]
    state["export_columns"] = ["Timestamp", "Product", "Movement type", "Qty change", "Total value", "Currency", "Order ID"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No movement rows"
    state["empty_copy"] = "No stock movements matched the current filters."
    return state


def _build_low_stock_report(factory_id: int):
    state = _report_base_state("low_stock")
    q = (request.args.get("q") or "").strip()
    report_rows = []
    low_count = 0
    out_count = 0

    factory_products = Product.query.filter(Product.factory_id == factory_id).order_by(Product.quantity.asc(), Product.name.asc()).all()
    for product in factory_products:
        qty = int(getattr(product, "quantity", 0) or 0)
        if qty > 5:
            continue
        if q and q.lower() not in f"{product.name} {getattr(product, 'category', '')}".lower():
            continue
        low_count += 1
        if qty <= 0:
            out_count += 1
        status_label = "Out" if qty <= 0 else "Low"
        report_rows.append(
            {
                "primary": product.name,
                "secondary_lines": [f"Factory · SKU #{product.id}", getattr(product, "category", None) or "No category"],
                "image_path": getattr(product, "image_path", None),
                "cells": ["Factory", str(qty), status_label],
                "mobile_value": str(qty),
                "mobile_subvalue": "Factory",
                "badge_label": status_label,
                "badge_tone": _report_tone_from_status(status_label),
                "export": {"product": product.name, "location": "Factory", "sku": product.id, "qty": qty, "status": status_label},
            }
        )

    for row in (
        ShopStock.query.join(Product, Product.id == ShopStock.product_id).join(Shop, Shop.id == ShopStock.shop_id).filter(ShopStock.source_factory_id == factory_id, ShopStock.quantity <= 5).order_by(ShopStock.quantity.asc(), Product.name.asc()).all()
    ):
        text = f"{row.product.name} {getattr(row.product, 'category', '')} {getattr(row.shop, 'name', '')}"
        if q and q.lower() not in text.lower():
            continue
        qty = int(row.quantity or 0)
        low_count += 1
        if qty <= 0:
            out_count += 1
        status_label = "Out" if qty <= 0 else "Low"
        report_rows.append(
            {
                "primary": row.product.name,
                "secondary_lines": [f"{row.shop.name} · SKU #{row.product.id}", getattr(row.product, "category", None) or "No category"],
                "image_path": getattr(row.product, "image_path", None),
                "cells": [row.shop.name, str(qty), status_label],
                "mobile_value": str(qty),
                "mobile_subvalue": row.shop.name,
                "badge_label": status_label,
                "badge_tone": _report_tone_from_status(status_label),
                "export": {"product": row.product.name, "location": row.shop.name, "sku": row.product.id, "qty": qty, "status": status_label},
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Low stock rows", str(low_count), "Factory + shop rows", "warn" if low_count else "up"),
        _report_summary_card("Out of stock", str(out_count), "Rows at zero", "danger" if out_count else "up"),
        _report_summary_card("Factory low rows", str(sum(1 for row in report_rows if row["cells"][0] == "Factory")), "Factory-side pressure", "info"),
        _report_summary_card("Shop low rows", str(sum(1 for row in report_rows if row["cells"][0] != "Factory")), "Branch-side pressure", "info"),
    ]
    state["table_columns"] = ["Location", "Qty", "Status"]
    state["rows"] = report_rows
    state["filter_fields"] = [{"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Product or location"}]
    state["export_columns"] = ["Product", "Location", "SKU", "Qty", "Status"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No low stock rows"
    state["empty_copy"] = "No factory or shop rows are currently at low stock."
    return state


def _build_cash_report(factory_id: int):
    state = _report_base_state("cash")
    date_range = _report_parse_date_range(default_days=30)
    currency_filter = (request.args.get("currency") or "").strip().upper()
    query = CashRecord.query.filter(CashRecord.factory_id == factory_id)
    if date_range["from"]:
        query = query.filter(CashRecord.date >= date_range["from"])
    if date_range["to"]:
        query = query.filter(CashRecord.date <= date_range["to"])
    if currency_filter in {"UZS", "USD"}:
        query = query.filter(CashRecord.currency == currency_filter)

    rows = query.order_by(CashRecord.date.desc(), CashRecord.id.desc()).all()
    income = 0.0
    expense = 0.0
    report_rows = []
    for row in rows:
        amount = float(getattr(row, "amount", 0) or 0)
        direction = "Income" if amount >= 0 else "Expense"
        if amount >= 0:
            income += amount
        else:
            expense += abs(amount)
        report_rows.append(
            {
                "primary": getattr(row, "note", None) or direction,
                "secondary_lines": [row.date.strftime("%Y-%m-%d") if getattr(row, "date", None) else "-", getattr(row, "currency", "UZS") or "UZS"],
                "image_path": None,
                "cells": [_report_money(amount, getattr(row, "currency", "UZS") or "UZS"), getattr(row, "currency", "UZS") or "UZS", direction],
                "mobile_value": _report_money(amount, getattr(row, "currency", "UZS") or "UZS"),
                "mobile_subvalue": direction,
                "badge_label": direction,
                "badge_tone": "up" if amount >= 0 else "warn",
                "export": {"date": row.date.strftime("%Y-%m-%d") if getattr(row, "date", None) else "", "note": getattr(row, "note", None) or "", "amount": round(amount, 2), "currency": getattr(row, "currency", "UZS") or "UZS", "direction": direction},
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Income", _report_money(income, currency_filter or "UZS"), "Positive cash rows", "up"),
        _report_summary_card("Expense", _report_money(expense, currency_filter or "UZS"), "Outgoing cash rows", "warn"),
        _report_summary_card("Balance", _report_money(income - expense, currency_filter or "UZS"), "Net over current range", "info"),
        _report_summary_card("Records", str(len(report_rows)), "Cash rows in this report", "neutral"),
    ]
    state["table_columns"] = ["Amount", "Currency", "Direction"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "from", "label": "From", "type": "date", "value": date_range["from_str"]},
        {"name": "to", "label": "To", "type": "date", "value": date_range["to_str"]},
        {"name": "currency", "label": "Currency", "type": "select", "value": currency_filter, "options": [{"value": "", "label": "All"}, {"value": "UZS", "label": "UZS"}, {"value": "USD", "label": "USD"}]},
    ]
    state["export_columns"] = ["Date", "Note", "Amount", "Currency", "Direction"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No cash rows"
    state["empty_copy"] = "No cash records matched the current date or currency filters."
    return state


def _build_production_report(factory_id: int):
    state = _report_base_state("production")
    q = (request.args.get("q") or "").strip()
    date_range = _report_parse_date_range(default_days=30)
    query = Production.query.join(Product).filter(Product.factory_id == factory_id)
    if q:
        query = query.filter(Product.name.ilike(f"%{q}%"))
    if date_range["from"]:
        query = query.filter(Production.date >= date_range["from"])
    if date_range["to"]:
        query = query.filter(Production.date <= date_range["to"])

    rows = query.order_by(Production.date.desc(), Production.id.desc()).all()
    today_value = date.today()
    week_start = today_value - timedelta(days=6)
    today_qty = 0
    week_qty = 0
    models = set()
    report_rows = []
    for row in rows:
        qty = int(getattr(row, "quantity", 0) or 0)
        models.add(getattr(row, "product_id", None))
        if getattr(row, "date", None) == today_value:
            today_qty += qty
        if getattr(row, "date", None) and week_start <= row.date <= today_value:
            week_qty += qty
        report_rows.append(
            {
                "primary": getattr(getattr(row, "product", None), "name", "Product"),
                "secondary_lines": [row.date.strftime("%Y-%m-%d") if getattr(row, "date", None) else "-", getattr(row, "note", None) or "Production row"],
                "image_path": getattr(getattr(row, "product", None), "image_path", None),
                "cells": [str(qty), str(getattr(row, "production_plan_id", None) or "-"), "Produced"],
                "mobile_value": str(qty),
                "mobile_subvalue": row.date.strftime("%Y-%m-%d") if getattr(row, "date", None) else "-",
                "badge_label": "Produced",
                "badge_tone": "up",
                "export": {"date": row.date.strftime("%Y-%m-%d") if getattr(row, "date", None) else "", "product": getattr(getattr(row, "product", None), "name", ""), "qty": qty, "plan_id": getattr(row, "production_plan_id", None) or "", "note": getattr(row, "note", None) or ""},
            }
        )

    state["summary_cards"] = [
        _report_summary_card("Produced today", str(today_qty), "Units made today", "up"),
        _report_summary_card("Produced 7 days", str(week_qty), "Recent weekly output", "info"),
        _report_summary_card("Production rows", str(len(report_rows)), "Rows in this range", "neutral"),
        _report_summary_card("Models", str(len([model for model in models if model])), "Distinct products", "neutral"),
    ]
    state["table_columns"] = ["Qty", "Plan", "Status"]
    state["rows"] = report_rows
    state["filter_fields"] = [
        {"name": "q", "label": "Search", "type": "search", "value": q, "placeholder": "Product"},
        {"name": "from", "label": "From", "type": "date", "value": date_range["from_str"]},
        {"name": "to", "label": "To", "type": "date", "value": date_range["to_str"]},
    ]
    state["export_columns"] = ["Date", "Product", "Qty", "Plan ID", "Note"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No production rows"
    state["empty_copy"] = "No production rows matched the current filters."
    return state


def _build_period_summary_report(workspace_id: int):
    state = _report_base_state("period_summary")
    summary_state = _build_business_summary(workspace_id)
    report_rows = []
    for branch in summary_state.get("branch_rows", [])[:6]:
        report_rows.append(
            {
                "primary": branch["name"],
                "secondary_lines": [branch.get("note", ""), "Branch performance"],
                "image_path": None,
                "cells": [branch["today_sales_display"]["primary"], branch["month_sales_display"]["primary"], branch["month_profit_display"]["primary"]],
                "mobile_value": branch["month_sales_display"]["primary"],
                "mobile_subvalue": "Month sales",
                "badge_label": f"{branch['stock_units']} in stock",
                "badge_tone": "info",
                "export": {"type": "Branch", "name": branch["name"], "today_sales": branch["today_sales_display"]["primary"], "month_sales": branch["month_sales_display"]["primary"], "month_profit": branch["month_profit_display"]["primary"]},
            }
        )

    for product in summary_state.get("top_product_rows", [])[:6]:
        report_rows.append(
            {
                "primary": product["name"],
                "secondary_lines": [product.get("category", ""), "Top product"],
                "image_path": None,
                "cells": [str(product["sold_qty_30"]), product["revenue_display"]["primary"], product["profit_display"]["primary"]],
                "mobile_value": product["revenue_display"]["primary"],
                "mobile_subvalue": "Revenue",
                "badge_label": product.get("pace_label", "Moving"),
                "badge_tone": "up",
                "export": {"type": "Product", "name": product["name"], "today_sales": "", "month_sales": product["revenue_display"]["primary"], "month_profit": product["profit_display"]["primary"]},
            }
        )

    state["report_subtitle"] = "Current period owner summary, branch performance, and best sellers in one compact view."
    state["summary_cards"] = summary_state.get("summary_cards", [])[:4]
    state["table_columns"] = ["Today / Units", "Month / Revenue", "Profit"]
    state["rows"] = report_rows
    state["export_columns"] = ["Type", "Name", "Today sales / units", "Month sales / revenue", "Month profit"]
    state["export_rows"] = [row["export"] for row in report_rows]
    state["empty_title"] = "No summary rows"
    state["empty_copy"] = "Branch and product summary rows are not available yet."
    state["custom_export_pdf_href"] = url_for("main.business_summary_pdf")
    state["custom_export_xlsx_href"] = url_for("main.business_summary_xlsx")
    return state


def _build_report_state(report_key: str, workspace):
    factory_id = getattr(workspace, "id", None)
    builders = {
        "factory_stock": lambda: _build_factory_stock_report(factory_id),
        "shop_stock": lambda: _build_shop_stock_report(factory_id),
        "sales": lambda: _build_sales_report(factory_id),
        "orders": lambda: _build_orders_report(factory_id),
        "cutting_orders": lambda: _build_cutting_orders_report(factory_id),
        "movements": lambda: _build_movements_report(factory_id),
        "low_stock": lambda: _build_low_stock_report(factory_id),
        "cash": lambda: _build_cash_report(factory_id),
        "production": lambda: _build_production_report(factory_id),
        "period_summary": lambda: _build_period_summary_report(factory_id),
    }
    if report_key not in builders:
        abort(404)
    state = builders[report_key]()
    state["report_home_href"] = url_for("main.business_summary")
    state["filter_action"] = url_for("main.report_detail", report_key=report_key)
    state["export_pdf_href"] = state.get("custom_export_pdf_href") or url_for("main.report_export_pdf", report_key=report_key, **request.args)
    state["export_xlsx_href"] = state.get("custom_export_xlsx_href") or url_for("main.report_export_xlsx", report_key=report_key, **request.args)
    return state


def _build_reports_hub_state(workspace_id: int):
    sales_stats = _get_sales_dashboard_stats(workspace_id)
    factory_products = Product.query.filter(Product.factory_id == workspace_id).all()
    shop_stock_rows = ShopStock.query.filter(ShopStock.source_factory_id == workspace_id).all()
    order_rows = ShopOrder.query.filter(ShopOrder.factory_id == workspace_id).all()
    cutting_order_rows = CuttingOrder.query.filter(CuttingOrder.factory_id == workspace_id).all()
    movement_count = StockMovement.query.filter(StockMovement.factory_id == workspace_id).count()
    cash_rows = CashRecord.query.filter(CashRecord.factory_id == workspace_id).all()
    production_rows = Production.query.join(Product, Product.id == Production.product_id).filter(Product.factory_id == workspace_id).all()
    factory_low = sum(1 for product in factory_products if int(getattr(product, "quantity", 0) or 0) <= 5)
    shop_low = sum(1 for row in shop_stock_rows if int(getattr(row, "quantity", 0) or 0) <= 5)
    cash_balance = sum(float(getattr(row, "amount", 0) or 0) for row in cash_rows if (getattr(row, "currency", "UZS") or "UZS") == "UZS")
    produced_today = sum(int(getattr(row, "quantity", 0) or 0) for row in production_rows if getattr(row, "date", None) == date.today())

    return {
        "hub_title": "Reports",
        "hub_subtitle": "Choose the exact report you want, then drill into a compact export-ready page.",
        "report_cards": [
            {"title": "Factory Stock Report", "description": "Current products, quantities, and factory stock value.", "icon": "FS", "stat": f"{sum(int(getattr(p, 'quantity', 0) or 0) for p in factory_products)} units", "href": url_for("main.report_detail", report_key="factory_stock")},
            {"title": "Shop Stock Report", "description": "Live branch stock and current low-stock pressure in shops.", "icon": "SS", "stat": f"{sum(int(getattr(row, 'quantity', 0) or 0) for row in shop_stock_rows)} units", "href": url_for("main.report_detail", report_key="shop_stock")},
            {"title": "Sales Report", "description": "Sales totals, sale rows, and recent performance windows.", "icon": "SL", "stat": _report_money(sales_stats.get("today_sales_uzs", 0), "UZS"), "href": url_for("main.report_detail", report_key="sales")},
            {"title": "Orders Report", "description": "Pending, ready, completed, and cancelled customer orders.", "icon": "OR", "stat": f"{sum(1 for row in order_rows if getattr(row, 'status', '') == 'pending')} pending", "href": url_for("main.report_detail", report_key="orders")},
            {"title": "Cutting Orders", "description": "Track sets cut, estimated material cost, and order status.", "icon": "CO", "stat": f"{len(cutting_order_rows)} orders", "href": url_for("main.report_detail", report_key="cutting_orders")},
            {"title": "Movement History Report", "description": "Factory-to-shop, sales-linked, and adjustment movement rows.", "icon": "MV", "stat": f"{movement_count} rows", "href": url_for("main.report_detail", report_key="movements")},
            {"title": "Low Stock Report", "description": "Priority items at low stock or already out in factory or shops.", "icon": "LS", "stat": f"{factory_low + shop_low} alerts", "href": url_for("main.report_detail", report_key="low_stock")},
            {"title": "Cash / Finance Summary", "description": "Cash inflow, outflow, and current balance across recent records.", "icon": "CA", "stat": _report_money(cash_balance, "UZS"), "href": url_for("main.report_detail", report_key="cash")},
            {"title": "Production Report", "description": "Produced quantities, active models, and recent production rows.", "icon": "PR", "stat": f"{produced_today} today", "href": url_for("main.report_detail", report_key="production")},
            {"title": "Annual / Period Summary", "description": "Compact owner-level summary with branch and product performance.", "icon": "YR", "stat": f"{len(factory_products)} models", "href": url_for("main.report_detail", report_key="period_summary")},
        ],
    }


def _build_report_pdf(report_state: dict) -> bytes:
    regular_font, bold_font = _summary_pdf_fonts()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName=bold_font, fontSize=18, leading=22, textColor=colors.HexColor("#0f172a"), alignment=TA_LEFT, spaceAfter=6)
    body_style = ParagraphStyle(name="ReportBody", parent=styles["Normal"], fontName=regular_font, fontSize=9, leading=12, textColor=colors.HexColor("#334155"))
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    story = [
        Paragraph(report_state["report_title"], title_style),
        Paragraph(f"{report_state['report_subtitle']} Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", body_style),
        Spacer(1, 8),
    ]

    summary_rows = [["Metric", "Value", "Note"]]
    for card in report_state.get("summary_cards", []):
        summary_rows.append([card["label"], card["value"], card.get("sub", "")])
    if len(summary_rows) > 1:
        story.append(_summary_pdf_table(summary_rows, col_widths=[50 * mm, 45 * mm, 85 * mm]))
        story.append(Spacer(1, 8))

    export_columns = report_state.get("export_columns") or ["Item", *report_state.get("table_columns", [])]
    export_rows = report_state.get("export_rows") or []
    table_rows = [export_columns]
    if export_rows:
        for row in export_rows:
            table_rows.append([str(_report_export_lookup(row, key)) for key in export_columns])
    else:
        for row in report_state.get("rows", []):
            primary_block = row["primary"]
            secondary = [item for item in row.get("secondary_lines", []) if item]
            if secondary:
                primary_block += "\n" + " | ".join(secondary)
            table_rows.append([primary_block, *row.get("cells", [])])
    if len(table_rows) > 1:
        story.append(_summary_pdf_table(table_rows))

    doc.build(story)
    return buffer.getvalue()


def _build_report_xlsx(report_state: dict):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Report"
    export_columns = report_state.get("export_columns") or ["Item", *report_state.get("table_columns", [])]
    export_rows = report_state.get("export_rows") or []
    ws.append(export_columns)
    if export_rows:
        for row in export_rows:
            ws.append([_report_export_lookup(row, key) for key in export_columns])
    else:
        for row in report_state.get("rows", []):
            primary_block = row["primary"]
            secondary = [item for item in row.get("secondary_lines", []) if item]
            if secondary:
                primary_block += " | " + " | ".join(secondary)
            ws.append([primary_block, *row.get("cells", [])])
    _summary_xlsx_style_sheet(ws)
    _summary_xlsx_fit_columns(ws)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@main_bp.route("/reports/<report_key>")
@login_required
def report_detail(report_key: str):
    workspace = _report_workspace_or_redirect()
    if not workspace:
        return redirect(url_for("main.dashboard"))
    state = _build_report_state(report_key, workspace)
    return render_template("dashboard/report_detail.html", **state)


@main_bp.route("/reports/<report_key>/pdf")
@login_required
def report_export_pdf(report_key: str):
    workspace = _report_workspace_or_redirect()
    if not workspace:
        return redirect(url_for("main.dashboard"))
    state = _build_report_state(report_key, workspace)
    buffer = BytesIO(_build_report_pdf(state))
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=_summary_export_filename(getattr(workspace, "name", None), report_key, "pdf"), mimetype="application/pdf", max_age=0)


@main_bp.route("/reports/<report_key>/xlsx")
@login_required
def report_export_xlsx(report_key: str):
    workspace = _report_workspace_or_redirect()
    if not workspace:
        return redirect(url_for("main.dashboard"))
    state = _build_report_state(report_key, workspace)
    workbook_buffer = _build_report_xlsx(state)
    return send_file(workbook_buffer, as_attachment=True, download_name=_summary_export_filename(getattr(workspace, "name", None), report_key, "xlsx"), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", max_age=0)
