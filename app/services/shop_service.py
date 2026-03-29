from typing import Optional, Dict, Any
from datetime import date

from sqlalchemy import or_, case
from flask import current_app
from openpyxl import load_workbook
from io import BytesIO
from copy import copy
from pathlib import Path

from ..shop_utils import get_or_create_default_shop
from ..extensions import db
from ..models import (
    Product,
    ShopStock,
    Sale,
    ShopOrder,
    ShopOrderItem,
    StockMovement,
    CashRecord,
    Shop,
    Movement,
    ShopFactoryLink,
    Factory,
)


class ShopService:
    """Business logic for shop stock (магазин)."""

    # =========================
    # Internal helpers
    # =========================

    def _get_product_or_raise(
        self,
        product_id: int,
        factory_id: Optional[int] = None,
    ) -> Product:
        query = Product.query.filter(Product.id == product_id)

        if factory_id is not None:
            query = query.filter(Product.factory_id == factory_id)

        product = query.first()
        if not product:
            raise ValueError("Товар не найден.")

        return product

    def _get_shop_stock_row(
        self,
        product_id: int,
        factory_id: Optional[int] = None,
        shop_id: Optional[int] = None,
    ) -> Optional[ShopStock]:
        query = ShopStock.query.filter(ShopStock.product_id == product_id)

        if factory_id is not None:
            query = query.filter(ShopStock.source_factory_id == factory_id)

        if shop_id is not None:
            query = query.filter(ShopStock.shop_id == shop_id)

        return query.first()

    @staticmethod
    def _safe_float(value) -> float:
        try:
            return float(value or 0)
        except Exception:
            return 0.0

    @staticmethod
    def _safe_int(value) -> int:
        try:
            return int(value or 0)
        except Exception:
            return 0

    def _add_stock_movement(
        self,
        *,
        factory_id: int,
        product_id: int,
        qty_change: int,
        source: str,
        destination: str,
        movement_type: str,
        order_id: Optional[int] = None,
        comment: Optional[str] = None,
    ) -> StockMovement:
        mv = StockMovement(
            factory_id=factory_id,
            product_id=product_id,
            qty_change=qty_change,
            source=source,
            destination=destination,
            movement_type=movement_type,
            order_id=order_id,
            comment=comment,
        )
        db.session.add(mv)
        return mv

    def _add_cash_record_for_sale(
        self,
        *,
        factory_id: int,
        sale: Sale,
        product: Product,
        customer_name: Optional[str] = None,
        note: Optional[str] = None,
    ) -> Optional[CashRecord]:
        qty_sold = self._safe_int(sale.quantity)
        currency = getattr(sale, "currency", None) or getattr(
            product, "currency", "UZS"
        )

        if getattr(sale, "total_sell", None) is not None:
            total_sell = self._safe_float(sale.total_sell)
        else:
            price = getattr(sale, "sell_price_per_item", None)
            if price is None:
                price = getattr(product, "sell_price_per_item", 0) or 0
            total_sell = self._safe_float(qty_sold) * self._safe_float(price)

        sale_date = getattr(sale, "date", None) or date.today()

        cash_note = f"Продажа (магазин) #{sale.id}: {product.name} x{qty_sold}"
        if customer_name:
            cash_note += f" — {customer_name}"
        if note:
            cash_note += f" ({note})"

        existing_cash = (
            CashRecord.query.filter_by(factory_id=factory_id, currency=currency)
            .filter(CashRecord.date == sale_date)
            .filter(CashRecord.amount == total_sell)
            .filter(CashRecord.note.ilike(f"%#{sale.id}%"))
            .first()
        )

        if existing_cash:
            return existing_cash

        cash = CashRecord(
            factory_id=factory_id,
            date=sale_date,
            amount=total_sell,
            currency=currency,
            note=cash_note,
        )
        db.session.add(cash)
        return cash

    # =========================
    # 1) List shop items
    # =========================

    def list_items(
        self,
        q: Optional[str] = None,
        sort: str = "name",
        factory_id: Optional[int] = None,
        shop_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        query = (
            ShopStock.query
            .join(Product, Product.id == ShopStock.product_id)
            .join(Factory, Factory.id == ShopStock.source_factory_id)
        )

        if factory_id is not None:
            query = query.filter(ShopStock.source_factory_id == factory_id)

        if shop_id is not None:
            query = query.filter(ShopStock.shop_id == shop_id)

        if q:
            like_pattern = f"%{q.strip()}%"
            query = query.filter(
                or_(
                    Product.name.ilike(like_pattern),
                    Product.category.ilike(like_pattern),
                    Factory.name.ilike(like_pattern),
                )
            )

        if sort == "name":
            query = query.order_by(Product.name.asc(), Factory.name.asc())

        elif sort == "category":
            query = query.order_by(
                case(
                    (Product.category.is_(None), 1),
                    (Product.category == "", 1),
                    else_=0,
                ).asc(),
                Product.category.asc(),
                Product.name.asc(),
                Factory.name.asc(),
            )

        elif sort == "qty_desc":
            query = query.order_by(ShopStock.quantity.desc(), Product.name.asc())

        elif sort == "qty_asc":
            query = query.order_by(ShopStock.quantity.asc(), Product.name.asc())

        elif sort == "factory":
            query = query.order_by(Factory.name.asc(), Product.name.asc())

        else:
            query = query.order_by(Product.name.asc(), Factory.name.asc())

        items = query.all()

        total_qty = sum(self._safe_int(item.quantity) for item in items)
        total_value_uzs = sum(
            self._safe_int(item.quantity) * self._safe_float(item.product.sell_price_per_item)
            for item in items
        )

        return {
            "items": items,
            "total_qty": total_qty,
            "total_value_uzs": total_value_uzs,
        }

    # =========================
    # 2) Transfer factory -> shop
    # =========================

    def transfer_to_shop(
        self,
        product_id: int,
        quantity: int,
        sell_price_per_item: Optional[float] = None,
        created_by=None,
        factory_id: Optional[int] = None,
        shop_id: Optional[int] = None,
    ) -> ShopStock:
        if quantity <= 0:
            raise ValueError("Количество должно быть больше нуля.")

        if not factory_id:
            raise ValueError("Не указана фабрика-источник.")

        if not shop_id:
            raise ValueError("Не указан магазин назначения.")

        product = self._get_product_or_raise(
            product_id=product_id,
            factory_id=factory_id,
        )

        shop = Shop.query.get(shop_id)
        if not shop:
            raise ValueError("Магазин не найден.")

        link = ShopFactoryLink.query.filter_by(
            shop_id=shop.id,
            factory_id=factory_id,
        ).first()
        if not link:
            raise ValueError("Этот магазин не привязан к выбранной фабрике.")

        if self._safe_int(product.quantity) < quantity:
            raise ValueError("На фабрике недостаточно остатка для передачи.")

        if sell_price_per_item is not None:
            if sell_price_per_item < 0:
                raise ValueError("Цена продажи не может быть отрицательной.")
            product.sell_price_per_item = sell_price_per_item

        shop_row = ShopStock.query.filter_by(
            shop_id=shop.id,
            product_id=product.id,
            source_factory_id=factory_id,
        ).first()

        if not shop_row:
            shop_row = ShopStock(
                shop_id=shop.id,
                product_id=product.id,
                source_factory_id=factory_id,
                quantity=0,
            )
            db.session.add(shop_row)

        shop_row.quantity = self._safe_int(shop_row.quantity) + quantity
        product.quantity = self._safe_int(product.quantity) - quantity

        self._add_stock_movement(
            factory_id=factory_id,
            product_id=product.id,
            qty_change=quantity,
            source=f"factory:{factory_id}",
            destination=f"shop:{shop.id}",
            movement_type="factory_to_shop",
            comment=f"Transferred {quantity} pcs to shop_id={shop.id}",
        )

        if created_by:
            mv = Movement(
                factory_id=factory_id,
                product_id=product.id,
                source=f"factory:{factory_id}",
                destination=f"shop:{shop.id}",
                change=quantity,
                note=f"Transferred {quantity} pcs to {shop.name}",
                created_by_id=created_by.id,
            )
            db.session.add(mv)

        db.session.commit()
        return shop_row

    # =========================
    # 3) Current quantity in shop
    # =========================

    def get_stock_quantity(
        self,
        product_id: int,
        factory_id: Optional[int] = None,
        shop_id: Optional[int] = None,
    ) -> int:
        stock = self._get_shop_stock_row(
            product_id=product_id,
            factory_id=factory_id,
            shop_id=shop_id,
        )
        return self._safe_int(stock.quantity) if stock else 0

    # =========================
    # 4) Sell from shop or create order
    # =========================

    def sell_from_shop_or_create_order(
        self,
        product_id: int,
        requested_qty: int,
        customer_name: Optional[str] = None,
        customer_phone: Optional[str] = None,
        note: Optional[str] = None,
        allow_partial_sale: bool = True,
        created_by=None,
        shop_stock_id: Optional[int] = None,
        factory_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        if requested_qty <= 0:
            raise ValueError("Количество должно быть больше нуля.")

        product = self._get_product_or_raise(
            product_id=product_id,
            factory_id=factory_id,
        )

        if shop_stock_id:
            stock = ShopStock.query.get(shop_stock_id)
            if not stock:
                raise ValueError("Строка shop stock не найдена.")

            if stock.product_id != product.id:
                raise ValueError("Shop stock не соответствует выбранному товару.")

            if factory_id is not None and stock.source_factory_id != factory_id:
                raise ValueError("Неверная фабрика-источник для shop stock.")

            effective_factory_id = stock.source_factory_id
        else:
            default_shop = get_or_create_default_shop(product.factory_id)

            stock = self._get_shop_stock_row(
                product_id=product_id,
                factory_id=factory_id,
                shop_id=default_shop.id,
            )

            effective_factory_id = product.factory_id

        available = self._safe_int(stock.quantity) if stock else 0

        sale_obj: Optional[Sale] = None
        order_obj: Optional[ShopOrder] = None

        # ---------- full sale ----------
        if available > 0 and requested_qty <= available:
            sale_obj = Sale(
                product_id=product.id,
                shop_id=stock.shop_id if stock else None,
                created_by_id=getattr(created_by, "id", None),
                date=date.today(),
                customer_name=customer_name,
                customer_phone=customer_phone,
                quantity=requested_qty,
                sell_price_per_item=self._safe_float(product.sell_price_per_item),
                cost_price_per_item=self._safe_float(product.cost_price_per_item),
                currency=product.currency or "UZS",
            )

            stock.quantity = available - requested_qty

            db.session.add(sale_obj)
            db.session.add(stock)
            db.session.flush()

            self._add_stock_movement(
                factory_id=effective_factory_id,
                product_id=product.id,
                qty_change=-requested_qty,
                source="shop",
                destination="customer",
                movement_type="shop_sale",
                order_id=None,
                comment=f"Продажа {requested_qty} шт. клиенту {customer_name or ''}".strip(),
            )

            self._add_cash_record_for_sale(
                factory_id=effective_factory_id,
                sale=sale_obj,
                product=product,
                customer_name=customer_name,
                note=note,
            )

            db.session.commit()

            return {
                "sale": sale_obj,
                "order": None,
                "missing": 0,
                "sold_now": requested_qty,
                "available_before": available,
            }

        # ---------- not enough stock ----------
        missing = max(requested_qty - available, 0)
        sold_now = 0

        if allow_partial_sale and available > 0:
            sold_now = available

            sale_obj = Sale(
                product_id=product.id,
                shop_id=stock.shop_id if stock else None,
                created_by_id=getattr(created_by, "id", None),
                date=date.today(),
                customer_name=customer_name,
                customer_phone=customer_phone,
                quantity=sold_now,
                sell_price_per_item=self._safe_float(product.sell_price_per_item),
                cost_price_per_item=self._safe_float(product.cost_price_per_item),
                currency=product.currency or "UZS",
            )

            stock.quantity = 0

            db.session.add(sale_obj)
            db.session.add(stock)
            db.session.flush()

            self._add_stock_movement(
                factory_id=effective_factory_id,
                product_id=product.id,
                qty_change=-sold_now,
                source="shop",
                destination="customer",
                movement_type="shop_sale",
                order_id=None,
                comment=f"Продажа {sold_now} шт. клиенту {customer_name or ''} (partial)".strip(),
            )

            self._add_cash_record_for_sale(
                factory_id=effective_factory_id,
                sale=sale_obj,
                product=product,
                customer_name=customer_name,
                note=note,
            )

        if available <= 0 or missing > 0:
            order_obj = ShopOrder(
                factory_id=effective_factory_id,
                customer_name=customer_name,
                customer_phone=customer_phone,
                note=note,
                status="pending",
                created_by=created_by,
            )

            order_item = ShopOrderItem(
                order=order_obj,
                product=product,
                qty_requested=requested_qty,
                qty_from_shop_now=sold_now,
                qty_remaining=missing if missing > 0 else requested_qty,
            )

            db.session.add(order_obj)
            db.session.add(order_item)

        db.session.commit()

        return {
            "sale": sale_obj,
            "order": order_obj,
            "missing": missing,
            "sold_now": sold_now,
            "available_before": available,
        }

    # =========================
    # 4.5) Ship order item from factory to shop
    # =========================

    def ship_order_item_to_shop(
        self,
        *,
        item_id: int,
        ship_qty: int,
        factory_id: int,
        created_by=None,
    ) -> Dict[str, Any]:
        item = (
            ShopOrderItem.query.join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
            .join(Product, Product.id == ShopOrderItem.product_id)
            .filter(
                ShopOrderItem.id == item_id,
                Product.factory_id == factory_id,
                ShopOrder.factory_id == factory_id,
            )
            .first()
        )

        if not item:
            raise ValueError("Позиция заказа не найдена.")

        order = item.order
        product = item.product

        if ship_qty <= 0:
            raise ValueError("Количество должно быть больше нуля.")

        if ship_qty > self._safe_int(item.qty_remaining):
            raise ValueError("Нельзя отправить больше, чем осталось по заказу.")

        if self._safe_int(product.quantity) < ship_qty:
            raise ValueError("На фабрике нет такого количества на складе.")

        product.quantity = self._safe_int(product.quantity) - ship_qty

        default_shop = get_or_create_default_shop(product.factory_id)

        link = ShopFactoryLink.query.filter_by(
            shop_id=default_shop.id,
            factory_id=factory_id,
        ).first()
        if not link:
            raise ValueError("Магазин не привязан к выбранной фабрике.")

        shop_row = ShopStock.query.filter_by(
            shop_id=default_shop.id,
            product_id=product.id,
            source_factory_id=factory_id,
        ).first()

        if not shop_row:
            shop_row = ShopStock(
                shop_id=default_shop.id,
                product_id=product.id,
                source_factory_id=factory_id,
                quantity=0,
            )
            db.session.add(shop_row)

        shop_row.quantity = self._safe_int(shop_row.quantity) + ship_qty

        item.qty_from_shop_now = self._safe_int(item.qty_from_shop_now) + ship_qty
        item.qty_remaining = self._safe_int(item.qty_remaining) - ship_qty

        if order:
            order.recalc_status()

        self._add_stock_movement(
            factory_id=factory_id,
            product_id=product.id,
            qty_change=ship_qty,
            source=f"factory:{factory_id}",
            destination=f"shop:{default_shop.id}",
            movement_type="factory_to_shop_for_order",
            order_id=order.id if order else None,
            comment=f"Shipped {ship_qty} pcs for order #{order.id} from factory to shop",
        )

        if created_by:
            mv = Movement(
                factory_id=factory_id,
                product_id=product.id,
                source=f"factory:{factory_id}",
                destination=f"shop:{default_shop.id}",
                change=ship_qty,
                note=f"Shipped {ship_qty} pcs for order #{order.id}",
                created_by_id=created_by.id,
            )
            db.session.add(mv)

        db.session.commit()

        return {
            "item": item,
            "order": order,
            "product": product,
            "ship_qty": ship_qty,
        }

    # =========================
    # 5) Pending order allocation
    # =========================

    def allocate_transfer_to_pending_orders(
        self,
        *,
        product_id: int,
        transferred_qty: int,
        factory_id: int,
    ) -> Dict[str, Any]:
        """
        After factory -> shop transfer, allocate newly arrived quantity
        to oldest pending shop orders for this product.

        Returns:
            {
                "fulfilled_order_ids": set[int],
                "remaining_unallocated": int,
            }
        """
        if transferred_qty <= 0:
            return {
                "fulfilled_order_ids": set(),
                "remaining_unallocated": 0,
            }

        product = self._get_product_or_raise(
            product_id=product_id,
            factory_id=factory_id,
        )

        remaining_to_allocate = self._safe_int(transferred_qty)
        fulfilled_order_ids = set()

        pending_items = (
            ShopOrderItem.query.join(ShopOrder, ShopOrder.id == ShopOrderItem.order_id)
            .filter(ShopOrderItem.product_id == product.id)
            .filter(ShopOrder.factory_id == factory_id)
            .filter(ShopOrder.status == "pending")
            .filter(ShopOrderItem.qty_remaining > 0)
            .order_by(ShopOrder.created_at.asc(), ShopOrderItem.id.asc())
            .all()
        )

        for item in pending_items:
            if remaining_to_allocate <= 0:
                break

            need = self._safe_int(item.qty_remaining)
            if need <= 0:
                continue

            shipped = min(remaining_to_allocate, need)

            item.qty_from_shop_now = self._safe_int(item.qty_from_shop_now) + shipped
            item.qty_remaining = self._safe_int(item.qty_remaining) - shipped

            if item.order:
                item.order.recalc_status()
                fulfilled_order_ids.add(item.order.id)

            remaining_to_allocate -= shipped

        db.session.commit()

        return {
            "fulfilled_order_ids": fulfilled_order_ids,
            "remaining_unallocated": remaining_to_allocate,
        }

    def transfer_factory_to_shop(
        self,
        *,
        product_id: int,
        quantity: int,
        factory_id: int,
        shop_id: int,
        sell_price_per_item: Optional[float] = None,
        created_by=None,
    ) -> Dict[str, Any]:
        """
        Full transfer flow:
        1. move stock from source factory to selected shop
        2. allocate transferred quantity to pending shop orders
        3. return summary for UI
        """
        shop_row = self.transfer_to_shop(
            product_id=product_id,
            quantity=quantity,
            sell_price_per_item=sell_price_per_item,
            created_by=created_by,
            factory_id=factory_id,
            shop_id=shop_id,
        )

        allocation_result = self.allocate_transfer_to_pending_orders(
            product_id=product_id,
            transferred_qty=quantity,
            factory_id=factory_id,
        )

        return {
            "shop_row": shop_row,
            "fulfilled_order_ids": allocation_result["fulfilled_order_ids"],
            "remaining_unallocated": allocation_result["remaining_unallocated"],
        }

    # =========================
    # 6) XLSX export helpers
    # =========================

    def _copy_row_style(self, ws, source_row: int, target_row: int, max_col: int = 12):
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

        for col in range(1, max_col + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)

            if source.has_style:
                target._style = copy(source._style)

            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format

    def export_full_report_xlsx(
        self,
        shop_id: int,
        q: Optional[str] = None,
        sort: str = "name",
    ):
        q = (q or "").strip()
        sort = (sort or "name").strip()

        query = (
            db.session.query(ShopStock)
            .join(Product, Product.id == ShopStock.product_id)
            .join(Factory, Factory.id == ShopStock.source_factory_id)
            .filter(ShopStock.shop_id == shop_id)
        )

        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    Product.name.ilike(like),
                    Product.category.ilike(like),
                    Factory.name.ilike(like),
                )
            )

        if sort == "qty":
            query = query.order_by(ShopStock.quantity.desc(), Product.name.asc())
        elif sort == "factory":
            query = query.order_by(Factory.name.asc(), Product.name.asc())
        else:
            query = query.order_by(Product.name.asc(), Factory.name.asc())

        stock_rows = query.all()

        template_path = (
            Path(current_app.root_path)
            / "templates"
            / "xlsx"
            / "dad_report_template.xlsx"
        )
        wb = load_workbook(template_path)
        ws = wb.active

        start_row = 4
        template_row = 4

        if len(stock_rows) > 1:
            ws.insert_rows(start_row + 1, amount=len(stock_rows) - 1)
            for i in range(1, len(stock_rows)):
                self._copy_row_style(ws, template_row, start_row + i, max_col=12)

        total_shop_value = 0
        total_units = 0

        for i, stock in enumerate(stock_rows):
            row = start_row + i
            product = stock.product
            source_factory = stock.source_factory

            qty = stock.quantity or 0
            price = 0

            if product:
                if getattr(product, "sell_price_per_item", None) is not None:
                    price = product.sell_price_per_item or 0
                elif getattr(product, "sale_price", None) is not None:
                    price = product.sale_price or 0

            total = qty * price
            total_units += qty
            total_shop_value += total

            ws[f"A{row}"] = source_factory.name if source_factory else ""
            ws[f"B{row}"] = product.name if product else ""
            ws[f"C{row}"] = qty
            ws[f"D{row}"] = price
            ws[f"E{row}"] = total

            ws[f"C{row}"].number_format = "#,##0"
            ws[f"D{row}"].number_format = "#,##0"
            ws[f"E{row}"].number_format = "#,##0"

        ws["I3"] = total_shop_value
        ws["I3"].number_format = "#,##0"

        ws["J4"] = 12750
        ws["J4"].number_format = "#,##0"

        ws["I5"] = ""
        ws["J6"] = total_units
        ws["J6"].number_format = "#,##0"

        ws["J8"] = total_shop_value - total_units if total_units else total_shop_value
        ws["J8"].number_format = "#,##0"

        ws["I10"] = "скидка"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()