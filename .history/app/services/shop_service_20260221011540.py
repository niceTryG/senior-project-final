from typing import Optional, Dict, Any
from datetime import date
from sqlalchemy import or_
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..extensions import db
from ..models import (
    Product,
    ShopStock,
    Sale,
    Movement,
)


class ShopService:
    # ---------- SHOP STOCK LIST ----------

    def list_items(
        self,
        q: Optional[str] = None,
        sort: str = "name",
        factory_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        query = ShopStock.query.join(ShopStock.product)

        if factory_id:
            query = query.filter(Product.factory_id == factory_id)

        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    Product.name.ilike(like),
                    Product.category.ilike(like),
                )
            )

        if sort == "qty_desc":
            query = query.order_by(ShopStock.quantity.desc())
        else:
            query = query.order_by(Product.name.asc())

        items = query.all()

        return {
            "items": items,
            "total_qty": sum(i.quantity for i in items),
            "total_value": sum(i.total_value for i in items),
        }

    # ---------- DAILY REPORT ----------

    def get_daily_report(self, factory_id: int, day: date):
        sales = (
            Sale.query
            .join(Product)
            .filter(
                Product.factory_id == factory_id,
                Sale.date == day,
            )
            .all()
        )

        total_sell = sum(s.total_sell or 0 for s in sales)
        total_cost = sum(s.total_cost or 0 for s in sales)
        profit = total_sell - total_cost

        cash_in = total_sell
        cash_out = 0  # extend later if needed

        return {
            "sales_count": len(sales),
            "total_sell": total_sell,
            "total_cost": total_cost,
            "profit": profit,
            "cash_in": cash_in,
            "cash_out": cash_out,
            "balance": cash_in - cash_out,
        }

    # ---------- FINAL XLSX EXPORT ----------

    def export_full_report_xlsx(
        self,
        factory_id: int,
        q: Optional[str] = None,
        sort: str = "name",
    ) -> bytes:
        wb = Workbook()

        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)

        # ===== SHEET 1: SHOP STOCK =====
        ws = wb.active
        ws.title = "Shop Stock"

        headers = ["ID", "Name", "Category", "Quantity", "Total Value (UZS)"]
        ws.append(headers)

        for c in range(1, 6):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        data = self.list_items(q=q, sort=sort, factory_id=factory_id)

        row_i = 2
        for item in data["items"]:
            p = item.product
            ws.append([
                p.id,
                p.name,
                p.category or "",
                item.quantity,
                item.total_value,
            ])
            row_i += 1

        ws.append([
            "",
            "TOTAL",
            "",
            data["total_qty"],
            data["total_value"],
        ])

        for c in range(1, 6):
            ws.cell(row=row_i, column=c).font = bold

        ws.freeze_panes = "A2"

        for c in range(1, 6):
            col = get_column_letter(c)
            ws.column_dimensions[col].width = 20

        # ===== SHEET 2: DAILY REPORT =====
        ws2 = wb.create_sheet("Daily Report")

        report = self.get_daily_report(factory_id, date.today())

        ws2["A1"] = "Mini Moda — Daily Report"
        ws2["A1"].font = Font(bold=True, size=14)

        ws2["A3"] = f"Date: {date.today()}"

        ws2.append([])
        ws2.append(["SALES"])
        ws2.append(["Sales count", report["sales_count"]])
        ws2.append(["Total revenue", report["total_sell"]])
        ws2.append(["Total cost", report["total_cost"]])
        ws2.append(["Profit", report["profit"]])

        ws2.append([])
        ws2.append(["CASH"])
        ws2.append(["Cash in", report["cash_in"]])
        ws2.append(["Cash out", report["cash_out"]])
        ws2.append(["Balance", report["balance"]])

        for row in (5, 10):
            ws2[f"A{row}"].font = bold

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 20

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()